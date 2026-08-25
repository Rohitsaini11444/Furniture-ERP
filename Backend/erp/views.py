import uuid
import csv
from datetime import datetime, timedelta
from decimal import Decimal
import io
from io import BytesIO
import json
import math
import os
import random
import re
import string
import tempfile
import zipfile

from num2words import num2words
import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import coordinate_to_tuple, get_column_letter
from PIL import Image as PILImage, ImageDraw
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as rl_canvas
from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from django.conf import settings
from django.core.files.base import ContentFile
from django.db.models import Case, Count, DecimalField, IntegerField, OuterRef, Q, Subquery, Sum, Value, When
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone

from rest_framework import generics, status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken

from .db_diagram_pdf import generate_db_relationships_pdf
from .models import (
    Buyer, BuyerMaster, BuyerMasterFinishingImage, BuyerPI, BuyerPIItem,
    BuyerUnitAllocation, ContractorPerson, Finish, GateInwardReceipt,
    Notification, POExtensionLog, POSupplierHistory, PerformaInvoice,
    PerformaInvoiceItem, ProductionJob, ProductionQCLog, ProductionUnit,
    Sample, SampleImage, StockItem, StoreDailyIssue, StoreItem,
    StoreItemCategory, StoreItemRateHistory, StoreItemStatus, StoreMaterialIn, StoreMaterialReturn,
    StorePurchaseOrder, StorePurchaseOrderItem, Supplier, SupplierDebitNote,
    StoreRequisition, StoreStockAdjustment,
    SupplierDebitNoteItem, SupplierPO, SupplierPOItem, SupplierPOItemDefect,
    SupplierPOItemDefectImage, SupplierTaxInvoice, SupplierTaxInvoiceItem,
    UnitWorkReallocation, User, UserSession, AuditLog, AuditAction
)
from .pagination import OptionalPagination
from .permissions import (
    IsAdmin, IsAdminOrSandingSupervisor, IsAdminOrSupervisor,
    IsContractor, IsSandingSupervisor, IsSupervisor, IsSupplierManager
)
from .presentation_generator import (
    find_image_path, generate_brand_pptx_presentation,
    generate_pptx_presentation, generate_vendor_inspection_pptx
)
from .serializers import (
    BuyerDropdownSerializer, BuyerMasterFinishingImageSerializer,
    BuyerMasterListSerializer, BuyerMasterSerializer, BuyerPIItemSerializer,
    BuyerPIListSerializer, BuyerPISerializer, BuyerSerializer,
    BuyerUnitAllocationSerializer, ContractorPersonSerializer,
    FinishDropdownSerializer, FinishSerializer, GateInwardReceiptSerializer,
    LoginSerializer, NotificationSerializer, POSupplierHistorySerializer,
    PerformaInvoiceItemSerializer, PerformaInvoiceSerializer,
    ProductionJobSerializer, ProductionQCLogSerializer, ProductionUnitSerializer,
    SampleCompactSerializer, SampleDropdownSerializer, SampleImageSerializer,
    SampleListSerializer, SampleSerializer, StockItemSerializer,
    StoreDailyIssueSerializer, StoreItemCategorySerializer,
    StoreItemRateHistorySerializer, StoreItemSerializer, StoreMaterialInSerializer, StoreMaterialReturnSerializer,
    StorePurchaseOrderItemSerializer, StorePurchaseOrderSerializer,
    StoreRequisitionSerializer, StoreStockAdjustmentSerializer,
    SupplierDebitNoteItemSerializer, SupplierDebitNoteSerializer,
    SupplierPOItemDefectSerializer, SupplierPOItemSerializer,
    SupplierPOListSerializer, SupplierPOSerializer, SupplierSerializer,
    SupplierTaxInvoiceItemSerializer, SupplierTaxInvoiceSerializer,
    UnitWorkReallocationSerializer, UserMinimalSerializer, UserSerializer,
    UserSessionSerializer, AuditLogSerializer
)


def add_centered_image(ws, cell_address, xl_img):
    """
    Embeds `xl_img` into `ws` at `cell_address` (e.g. 'F2' or 'B3'),
    centering the image both horizontally and vertically within the target cell.
    """
    row_idx, col_idx = coordinate_to_tuple(cell_address)
    col_letter = get_column_letter(col_idx)

    anchor_row = row_idx - 1
    anchor_col = col_idx - 1

    row_height_pt = ws.row_dimensions[row_idx].height or 15.0
    col_width_char = ws.column_dimensions[col_letter].width or 8.43

    cell_height_emu = int(row_height_pt * 12700)
    cell_width_emu = int((col_width_char * 7.5 + 5) * 9525)

    img_width_emu = int(xl_img.width * 9525)
    img_height_emu = int(xl_img.height * 9525)

    col_off = max(0, (cell_width_emu - img_width_emu) // 2)
    row_off = max(0, (cell_height_emu - img_height_emu) // 2)

    marker = AnchorMarker(col=anchor_col, colOff=col_off, row=anchor_row, rowOff=row_off)
    size = XDRPositiveSize2D(cx=img_width_emu, cy=img_height_emu)

    xl_img.anchor = OneCellAnchor(_from=marker, ext=size)
    ws.add_image(xl_img)



# ─── Auth Views ───────────────────────────────────────────────────────────────

class LoginView(APIView):
    """
    POST /api/auth/login/
    Returns JWT access + refresh tokens along with user profile.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = LoginSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']

        refresh = RefreshToken.for_user(user)

        # Track Session & Security Notification
        ip_address = request.META.get('REMOTE_ADDR') or '127.0.0.1'
        user_agent = request.META.get('HTTP_USER_AGENT', '')[:512]
        
        # Deactivate any previous active sessions for this user on the same IP and user agent
        UserSession.objects.filter(
            user=user,
            ip_address=ip_address,
            user_agent=user_agent,
            is_active=True
        ).update(is_active=False)

        session = UserSession.objects.create(
            user=user,
            ip_address=ip_address,
            user_agent=user_agent
        )

        from .serializers import parse_user_agent
        parsed_ua = parse_user_agent(user_agent)
        
        Notification.objects.create(
            user=user,
            title="New login detected",
            message=f"from {ip_address}\n{user_agent[:120]}",
            category="security",
            link="/active-devices"
        )

        return Response({
            'access': str(refresh.access_token),
            'refresh': str(refresh),
            'session_id': session.id,
            'user': {
                'id': user.id,
                'username': user.username,
                'full_name': user.get_full_name() or user.username,
                'email': user.email,
                'role': user.role,
                'batch_category': user.batch_category,
                'supervisor_id': user.supervisor_id,
                'profile_image': request.build_absolute_uri(user.profile_image.url) if user.profile_image else None,
            }
        }, status=status.HTTP_200_OK)

class LogoutView(APIView):
    """
    POST /api/auth/logout/
    Blacklists the refresh token and deactivates the user session.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        session_id = request.data.get('session_id')
        if session_id:
            try:
                session_id_int = int(session_id)
                UserSession.objects.filter(id=session_id_int, user=request.user).update(is_active=False)
            except (ValueError, TypeError):
                UserSession.objects.filter(id=session_id, user=request.user).update(is_active=False)
        else:
            UserSession.objects.filter(user=request.user, is_active=True).update(is_active=False)

        try:
            refresh_token = request.data.get('refresh')
            if refresh_token:
                token = RefreshToken(refresh_token)
                token.blacklist()
        except Exception:
            pass  # Token blacklisting optional

        return Response({'detail': 'Logged out successfully.'}, status=status.HTTP_200_OK)


class ActiveDevicesView(APIView):
    """
    GET /api/auth/devices/
    Returns active devices.
    - Admins see sessions across all users (or filter by ?user_id=...).
    - Standard users see their own active sessions.
    POST /api/auth/devices/
    DELETE /api/auth/devices/
    Deactivates/Revokes a specific session.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        user_id = request.query_params.get('user_id')
        
        if user.role == 'admin':
            if user_id:
                sessions = UserSession.objects.filter(user_id=user_id, is_active=True).select_related('user').order_by('-last_activity')
            else:
                sessions = UserSession.objects.filter(is_active=True).select_related('user').order_by('-last_activity')
        else:
            sessions = UserSession.objects.filter(user=user, is_active=True).select_related('user').order_by('-last_activity')
            
        serializer = UserSessionSerializer(sessions, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        return self._revoke(request)

    def delete(self, request):
        return self._revoke(request)

    def _revoke(self, request):
        session_id = request.data.get('session_id') or request.query_params.get('session_id')
        if not session_id:
            return Response({'detail': 'session_id is required.'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            session_id_int = int(session_id)
        except (ValueError, TypeError):
            session_id_int = session_id

        qs = UserSession.objects.filter(id=session_id_int)
        if request.user.role != 'admin':
            qs = qs.filter(user=request.user)
            
        updated = qs.update(is_active=False)
        if updated:
            return Response({'detail': 'Session revoked successfully.'}, status=status.HTTP_200_OK)
        return Response({'detail': 'Session not found or permission denied.'}, status=status.HTTP_404_NOT_FOUND)


# ─── User Management (Admin Only) ─────────────────────────────────────────────

class UserViewSet(viewsets.ModelViewSet):
    """
    Admin-only CRUD for managing all users.
    GET /api/users/?role=supervisor  — filter by role
    """
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.action in ('list', 'retrieve', 'supervisors', 'contractors'):
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsAdmin()]

    def get_queryset(self):
        user = self.request.user
        qs = User.objects.all().order_by('role', 'username')
        role = self.request.query_params.get('role')
        if role:
            qs = qs.filter(role=role)
        supervisor_id = self.request.query_params.get('supervisor')
        if supervisor_id:
            qs = qs.filter(supervisor_id=supervisor_id)
        if user.role == 'supervisor' and role == 'contractor':
            qs = qs.filter(Q(supervisor=user) | Q(supervisor__isnull=True))
        return qs

    def perform_update(self, serializer):
        instance = serializer.save()
        if 'is_active' in serializer.validated_data and not serializer.validated_data['is_active']:
            UserSession.objects.filter(user=instance, is_active=True).update(is_active=False)

    @action(detail=False, methods=['get'], url_path='supervisors')
    def supervisors(self, request):
        """GET /api/users/supervisors/ — list all supervisors (for contractor assignment dropdown)"""
        qs = User.objects.filter(role='supervisor', is_active=True)
        return Response(UserMinimalSerializer(qs, many=True).data)

    @action(detail=True, methods=['get'], url_path='contractors')
    def contractors(self, request, pk=None):
        """GET /api/users/<id>/contractors/ — list contractors under a supervisor"""
        user = self.get_object()
        contractors = User.objects.filter(supervisor=user, role='contractor', is_active=True)
        return Response(UserMinimalSerializer(contractors, many=True).data)


class CurrentUserView(APIView):
    """GET /api/auth/me/ — returns logged-in user profile."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = UserSerializer(request.user, context={'request': request})
        return Response(serializer.data)

    def patch(self, request):
        serializer = UserSerializer(request.user, data=request.data, partial=True, context={'request': request})
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


# ─── ERP Core ViewSets ────────────────────────────────────────────────────────

class FinishViewSet(viewsets.ModelViewSet):
    """
    Finishes / Polish Catalog ViewSet.
    Accessible to all authenticated users. Admins can CRUD.
    """
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list' and self.request.query_params.get('nopage') == 'true':
            return FinishDropdownSerializer
        return FinishSerializer

    def get_queryset(self):
        qs = Finish.objects.all().order_by('-created_at')
        q = self.request.query_params.get('search')
        wood_type = self.request.query_params.get('wood_type')
        color = self.request.query_params.get('color')

        if q:
            q = q.strip()
            qs = qs.filter(
                Q(name__icontains=q) |
                Q(finish_code__icontains=q) |
                Q(color__icontains=q) |
                Q(wood_type__icontains=q)
            )
        if wood_type:
            qs = qs.filter(wood_type__icontains=wood_type)
        if color:
            qs = qs.filter(color__icontains=color)

        ordering = self.request.query_params.get('ordering')
        if ordering:
            qs = qs.order_by(ordering)
        return qs

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [IsAuthenticated(), IsAdmin()]
        return [IsAuthenticated()]


class SampleViewSet(viewsets.ModelViewSet):
    """
    Samples — accessible to all authenticated users.
    Admins & Supervisors can create/edit; Contractors read-only.
    """
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list':
            # Support compact/minimal response either via dedicated endpoint
            # `/api/samples/compact/` or by adding `compact=true` to the
            # regular list endpoint: `/api/samples/?compact=true`.
            if self.request.query_params.get('compact') == 'true':
                return SampleCompactSerializer
            if self.request.query_params.get('nopage') == 'true':
                return SampleDropdownSerializer
            return SampleListSerializer
        return SampleSerializer

    # def get_queryset(self):
    #     qs = Sample.objects.select_related('buyer').prefetch_related('images').all()
    #     buyer = self.request.query_params.get('buyer')
    #     material = self.request.query_params.get('material')
    #     if buyer:
    #         qs = qs.filter(buyer_id=buyer)
    #     if material:
    #         qs = qs.filter(material__icontains=material)
    #     return qs

    def get_queryset(self):
        qs = Sample.objects.select_related('buyer', 'finish').prefetch_related('images').all()

        buyer = self.request.query_params.get('buyer')
        material = self.request.query_params.get('material')
        q = self.request.query_params.get('search')

        if buyer:
            qs = qs.filter(buyer_id=buyer)
        if material:
            qs = qs.filter(material__icontains=material)

        if q:
            q = q.strip()
            # 1. Filter ONLY by style_no or product_name
            qs = qs.filter(
                Q(style_no__icontains=q) |
                Q(product_name__icontains=q)
            )

            # 2. Prioritize exact matches at the top of the list
            qs = qs.annotate(
                match_priority=Case(
                    When(style_no__iexact=q, then=Value(1)),      # Exact style_no match first
                    When(product_name__iexact=q, then=Value(2)),  # Exact product_name match second
                    When(style_no__istartswith=q, then=Value(3)), # Starts with search query third
                    default=Value(4),
                    output_field=IntegerField(),
                )
            ).order_by('match_priority', 'style_no')

        return qs   

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy', 'import_excel'):
            return [IsAuthenticated(), IsAdmin()]
        return [IsAuthenticated()]

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        """
        GET /api/samples/download-template/
        Returns an empty Excel template formatted with expected headers & example data.
        """

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sample_Import_Template"
        ws.views.sheetView[0].showGridLines = True

        headers = [
            'Style No.*', 'Product Name*', 'Sample Photo / Image', 'Buyer Code', 'Material', 
            'Finish / Color', 'Size Length (cm)', 'Size Breadth (cm)', 
            'Size Height (cm)', 'USD ($)', 'CBM', 'Vendor Name', 'Remark'
        ]

        ws.append(headers)

        header_fill = PatternFill(start_color="8b5a2b", end_color="8b5a2b", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        sample_row = [
            'STY-101', 'Handcrafted Dining Chair', '[Photo Embedded Below]', 'BUY-01', 'Solid Sheesham Wood',
            'Honey Finish', 55.0, 50.0, 95.0,
            120.00, 0.26, 'Raj Artisans', 'Insert sample photos into worksheet cells for auto-import'
        ]
        ws.append(sample_row)

        # Embed demo sample image into Cell C2
        ws.column_dimensions['C'].width = 22
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 60
        try:
            
            img_buf = io.BytesIO()
            demo_img = PILImage.new('RGB', (140, 140), color='#8b5a2b')
            draw = ImageDraw.Draw(demo_img)
            draw.rectangle([(12, 12), (128, 128)], outline='#ffffff', width=3)
            draw.text((22, 55), "SAMPLE PHOTO", fill='#ffffff')
            demo_img.save(img_buf, format='PNG')
            img_buf.seek(0)
            
            excel_img = OpenpyxlImage(img_buf)
            excel_img.width = 65
            excel_img.height = 65
            add_centered_image(ws, 'C2', excel_img)
        except Exception as e:
            print(f"Error creating template image: {e}")

        for col in ws.columns:
            if get_column_letter(col[0].column) == 'C':
                continue
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Samples_Import_Template.xlsx"'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return response

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        """
        POST /api/samples/import-excel/
        Uploads an .xlsx file containing sample data and optional embedded cell images.
        """

        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({
                "error_type": "No File Uploaded",
                "detail": "Please select a valid .xlsx file to upload."
            }, status=status.HTTP_400_BAD_REQUEST)

        if not (excel_file.name.lower().endswith('.xlsx')):
            return Response({
                "error_type": "Invalid Extension",
                "detail": "File must be an Excel (.xlsx) spreadsheet. Please download the expected template below."
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
        except Exception:
            return Response({
                "error_type": "Unreadable File",
                "detail": "Unable to read Excel file. Please ensure the file is not corrupt."
            }, status=status.HTTP_400_BAD_REQUEST)

        header_row = [str(cell.value or '').strip() for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]
        
        header_map = {}
        for idx, col in enumerate(header_row):
            col_norm = col.lower().replace('.', '').replace('*', '').replace('_', ' ').strip()
            header_map[col_norm] = idx

        style_idx = next((header_map[k] for k in ['style no', 'style', 'style #', 'style_no', 'sample id', 'sample_id'] if k in header_map), None)
        product_name_idx = next((header_map[k] for k in ['product name', 'product', 'item name', 'name', 'product_name'] if k in header_map), None)

        if style_idx is None or product_name_idx is None:
            return Response({
                "error_type": "Header / Schema Mismatch",
                "detail": "Required headers 'Style No.' or 'Product Name' are missing. Please download the expected template below."
            }, status=status.HTTP_400_BAD_REQUEST)

        buyer_idx = next((header_map[k] for k in ['buyer code', 'buyer', 'buyer name'] if k in header_map), None)
        material_idx = next((header_map[k] for k in ['material', 'wood'] if k in header_map), None)
        finish_idx = next((header_map[k] for k in ['finish / color', 'finish', 'color', 'finish color'] if k in header_map), None)
        len_idx = next((header_map[k] for k in ['size length (cm)', 'length (cm)', 'length', 'size length'] if k in header_map), None)
        breadth_idx = next((header_map[k] for k in ['size breadth (cm)', 'breadth (cm)', 'width (cm)', 'breadth', 'width', 'size breadth'] if k in header_map), None)
        height_idx = next((header_map[k] for k in ['size height (cm)', 'height (cm)', 'height', 'size height'] if k in header_map), None)
        usd_idx = next((header_map[k] for k in ['usd ($)', 'usd', 'price (usd)', 'price'] if k in header_map), None)
        cbm_idx = next((header_map[k] for k in ['cbm', 'total cbm'] if k in header_map), None)
        vendor_idx = next((header_map[k] for k in ['vendor name', 'vendor', 'supplier'] if k in header_map), None)
        remark_idx = next((header_map[k] for k in ['remark', 'remarks', 'note', 'notes'] if k in header_map), None)

        row_images = {}
        if hasattr(ws, '_images'):
            for img in ws._images:
                try:
                    img_row = img.anchor._from.row + 1
                    if img_row not in row_images:
                        row_images[img_row] = []
                    row_images[img_row].append(img)
                except Exception:
                    pass

        imported_count = 0
        updated_count = 0
        images_extracted = 0

        buyers_by_code = {b.code.lower(): b for b in Buyer.objects.all()}
        buyers_by_name = {b.name.lower(): b for b in Buyer.objects.all()}

        for excel_row_num, row_cells in enumerate(ws.iter_rows(min_row=2), start=2):
            cells = [cell.value for cell in row_cells]
            if not any(cells):
                continue

            style_no_val = str(cells[style_idx]).strip() if style_idx < len(cells) and cells[style_idx] is not None else ''
            product_name_val = str(cells[product_name_idx]).strip() if product_name_idx < len(cells) and cells[product_name_idx] is not None else ''

            if not style_no_val or not product_name_val:
                continue

            buyer_obj = None
            if buyer_idx is not None and buyer_idx < len(cells) and cells[buyer_idx] is not None:
                b_str = str(cells[buyer_idx]).strip().lower()
                buyer_obj = buyers_by_code.get(b_str) or buyers_by_name.get(b_str)

            def parse_dec(val):
                if val is None or val == '': return None
                try:
                    return float(str(val).replace('$', '').replace('₹', '').replace(',', '').strip())
                except ValueError:
                    return None

            material_val = str(cells[material_idx]).strip() if material_idx is not None and material_idx < len(cells) and cells[material_idx] is not None else ''
            finish_val = str(cells[finish_idx]).strip() if finish_idx is not None and finish_idx < len(cells) and cells[finish_idx] is not None else ''
            vendor_val = str(cells[vendor_idx]).strip() if vendor_idx is not None and vendor_idx < len(cells) and cells[vendor_idx] is not None else ''
            remark_val = str(cells[remark_idx]).strip() if remark_idx is not None and remark_idx < len(cells) and cells[remark_idx] is not None else ''

            size_len = parse_dec(cells[len_idx] if len_idx is not None and len_idx < len(cells) else None)
            size_brd = parse_dec(cells[breadth_idx] if breadth_idx is not None and breadth_idx < len(cells) else None)
            size_hgt = parse_dec(cells[height_idx] if height_idx is not None and height_idx < len(cells) else None)
            usd_val = parse_dec(cells[usd_idx] if usd_idx is not None and usd_idx < len(cells) else None)
            cbm_val = parse_dec(cells[cbm_idx] if cbm_idx is not None and cbm_idx < len(cells) else None)

            sample_obj, created = Sample.objects.get_or_create(
                style_no=style_no_val,
                defaults={
                    'sample_id': style_no_val,
                    'product_name': product_name_val,
                    'buyer': buyer_obj,
                    'material': material_val,
                    'finish_color': finish_val,
                    'size_length': size_len,
                    'size_breadth': size_brd,
                    'size_height': size_hgt,
                    'usd': usd_val,
                    'cbm': cbm_val,
                    'vendor_name': vendor_val,
                    'remark': remark_val,
                }
            )

            if not created:
                sample_obj.product_name = product_name_val
                if buyer_obj: sample_obj.buyer = buyer_obj
                if material_val: sample_obj.material = material_val
                if finish_val: sample_obj.finish_color = finish_val
                if size_len is not None: sample_obj.size_length = size_len
                if size_brd is not None: sample_obj.size_breadth = size_brd
                if size_hgt is not None: sample_obj.size_height = size_hgt
                if usd_val is not None: sample_obj.usd = usd_val
                if cbm_val is not None: sample_obj.cbm = cbm_val
                if vendor_val: sample_obj.vendor_name = vendor_val
                if remark_val: sample_obj.remark = remark_val
                sample_obj.save()
                updated_count += 1
            else:
                imported_count += 1

            if excel_row_num in row_images and row_images[excel_row_num]:
                if not created:
                    # Clear previous images on update to prevent duplicate stacked images
                    sample_obj.images.all().delete()
                    sample_obj.image = None

                for img_idx, img_obj in enumerate(row_images[excel_row_num]):
                    try:
                        image_bytes = img_obj._data()
                        ext = img_obj.format if hasattr(img_obj, 'format') and img_obj.format else 'png'
                        file_name = f"{style_no_val.replace('/', '_')}_{img_idx+1}.{ext}"
                        content_file = ContentFile(image_bytes, name=file_name)

                        s_img = SampleImage.objects.create(sample=sample_obj, image=content_file)
                        images_extracted += 1

                        if not sample_obj.image and s_img.image:
                            sample_obj.image = s_img.image
                            sample_obj.save(update_fields=['image'])
                    except Exception as img_err:
                        print(f"Error saving image for row {excel_row_num}: {img_err}")

        return Response({
            "detail": f"Import complete! {imported_count} new sample(s) created, {updated_count} updated. {images_extracted} high-quality image(s) extracted.",
            "imported_count": imported_count,
            "updated_count": updated_count,
            "images_extracted": images_extracted,
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='compact')
    def compact(self, request):
        """GET /api/samples/compact/?search=... — returns minimal fields for quick search/dropdowns."""

        qs = self.get_queryset()
        q = request.query_params.get('search')
        if q:
            qs = qs.filter(
                Q(sample_id__icontains=q) |
                Q(style_no__icontains=q) |
                Q(product_name__icontains=q)
            )

        page = self.paginate_queryset(qs)
        serializer = SampleCompactSerializer(page or qs, many=True, context={'request': request})
        if page is not None:
            return self.get_paginated_response(serializer.data)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='dropdown')
    def dropdown(self, request):
        """GET /api/samples/dropdown/ — returns lightweight sample fields for dropdowns."""
        qs = self.filter_queryset(self.get_queryset())
        serializer = SampleDropdownSerializer(qs, many=True, context={'request': request})
        return Response(serializer.data)


class SampleImageViewSet(viewsets.ModelViewSet):
    """
    Manage images for a sample.
    POST /api/sample-images/  — upload an image (pass `sample` id in body)
    DELETE /api/sample-images/<id>/  — delete a specific image
    """
    serializer_class = SampleImageSerializer
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'delete', 'head', 'options']

    def get_queryset(self):
        qs = SampleImage.objects.select_related('sample')
        sample_id = self.request.query_params.get('sample')
        if sample_id:
            qs = qs.filter(sample_id=sample_id)
        return qs

    def get_permissions(self):
        if self.action in ('create', 'destroy'):
            return [IsAuthenticated(), IsAdmin()]
        return [IsAuthenticated()]

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

class BuyerViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list' and self.request.query_params.get('nopage') == 'true':
            return BuyerDropdownSerializer
        return BuyerSerializer

    def get_queryset(self):
        return Buyer.objects.filter(is_deleted=False).order_by('name')

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [IsAuthenticated(), IsAdmin()]
        return [IsAuthenticated()]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        note = request.query_params.get('note', '')
        instance.is_deleted = True
        instance.deleted_at = timezone.now()
        instance.deletion_note = note
        instance.deleted_by = request.user
        instance.save()
        return Response({"detail": "Buyer soft-deleted and logged successfully."}, status=status.HTTP_200_OK)


class BuyerMasterFinishingImageViewSet(viewsets.ModelViewSet):
    """
    Manage finishing images for a buyer master.
    POST /api/buyer-master-finishing-images/  — upload an image
    DELETE /api/buyer-master-finishing-images/<id>/  — delete a specific image
    """
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', 'post', 'delete', 'head', 'options']

    def get_serializer_class(self):
        return BuyerMasterFinishingImageSerializer

    def get_queryset(self):
        qs = BuyerMasterFinishingImage.objects.select_related('buyer_master')
        bm_id = self.request.query_params.get('buyer_master')
        if bm_id:
            qs = qs.filter(buyer_master_id=bm_id)
        return qs


class BuyerMasterViewSet(viewsets.ModelViewSet):
    queryset = BuyerMaster.objects.select_related('buyer', 'sample', 'sample__finish', 'sample__buyer').prefetch_related('finishing_images').all()
    permission_classes = [IsAuthenticated]
    pagination_class = OptionalPagination

    def get_serializer_class(self):
        if self.action == 'list':
            return BuyerMasterListSerializer
        return BuyerMasterSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        buyer_id = self.request.query_params.get('buyer')
        if buyer_id:
            qs = qs.filter(buyer_id=buyer_id)
        return qs

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy', 'import_excel'):
            return [IsAuthenticated(), IsAdmin()]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        buyer_master = serializer.save()
        images = self.request.FILES.getlist('finishing_images')
        for img in images:
            BuyerMasterFinishingImage.objects.create(buyer_master=buyer_master, image=img)

    def perform_update(self, serializer):
        if self.request.data.get('clear_packaging_image') in ('true', True, '1'):
            serializer.instance.packaging_image = None
        buyer_master = serializer.save()
        images = self.request.FILES.getlist('finishing_images')
        for img in images:
            BuyerMasterFinishingImage.objects.create(buyer_master=buyer_master, image=img)

    @action(detail=True, methods=['get'], url_path='download-packaging-image')
    def download_packaging_image(self, request, pk=None):
        bm = self.get_object()
        if not bm.packaging_image or not os.path.exists(bm.packaging_image.path):
            return HttpResponse("Packaging image not found", status=404)
        
        sample_name = bm.style_no or (bm.sample.sample_id if bm.sample else "Style")
        safe_name = "".join(c for c in sample_name if c.isalnum() or c in ('-', '_')).strip()
        ext = os.path.splitext(bm.packaging_image.path)[1] or '.png'
        filename = f"{safe_name}_Packaging_Image{ext}"

        with open(bm.packaging_image.path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/octet-stream')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

    @action(detail=True, methods=['get'], url_path='download-finishing-images')
    def download_finishing_images(self, request, pk=None):
        bm = self.get_object()
        images = bm.finishing_images.all()
        if not images.exists():
            return HttpResponse("No finishing images found", status=404)
        
        sample_name = bm.style_no or (bm.sample.sample_id if bm.sample else "Style")
        safe_name = "".join(c for c in sample_name if c.isalnum() or c in ('-', '_')).strip()
        zip_filename = f"{safe_name}_Finishing_images.zip"

        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for idx, img_obj in enumerate(images, 1):
                if img_obj.image and os.path.exists(img_obj.image.path):
                    ext = os.path.splitext(img_obj.image.path)[1] or '.png'
                    arcname = f"{safe_name}_Finishing_Image_{idx}{ext}"
                    zf.write(img_obj.image.path, arcname=arcname)
        
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
        return response

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        """
        GET /api/buyer-masters/download-template/?with_details=true|false
        Generates empty Excel template for Buyer Master.
        """

        with_details = request.query_params.get('with_details') == 'true'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Buyer_Master_Template"
        ws.views.sheetView[0].showGridLines = True

        headers = [
            'Buyer Code*', 'Buyer Name*', 'Style No.*', 'Product Name*', 'Finishing Photo / Image',
            'Material / Wood', 'Finish Color', 'Size Length (cm)', 'Size Breadth (cm)', 
            'Size Height (cm)', 'Price USD ($)', 'Units', 'CBM', 'Remark'
        ]

        if with_details:
            headers.extend([
                'Vendor Details', 'Vendor Price', 'Costing', 'Purchase Price', 
                'Net Weight', 'Gross Weight', 'Box Length (cm)', 'Box Breadth (cm)', 'Box Height (cm)'
            ])

        ws.append(headers)

        header_fill = PatternFill(start_color="7c3aed", end_color="7c3aed", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        sample_row = [
            'BUY-01', 'Nkuku UK', 'STY-201', 'Solid Wood Bookshelf', '[Photo Embedded Below]',
            'Mango Wood', 'Walnut Stain', 120.0, 40.0, 180.0,
            245.00, 10, 0.864, 'Insert finishing photos into worksheet cells for auto-import'
        ]
        if with_details:
            sample_row.extend([
                'Rajesh Artisans, Jodhpur', 180.00, 210.00, 195.00,
                32.5, 36.0, 125.0, 45.0, 185.0
            ])

        ws.append(sample_row)

        # Embed demo finishing image into Cell E2
        ws.column_dimensions['E'].width = 22
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 60
        try:
            
            img_buf = io.BytesIO()
            demo_img = PILImage.new('RGB', (140, 140), color='#7c3aed')
            draw = ImageDraw.Draw(demo_img)
            draw.rectangle([(12, 12), (128, 128)], outline='#ffffff', width=3)
            draw.text((18, 55), "FINISHING PHOTO", fill='#ffffff')
            demo_img.save(img_buf, format='PNG')
            img_buf.seek(0)
            
            excel_img = OpenpyxlImage(img_buf)
            excel_img.width = 65
            excel_img.height = 65
            add_centered_image(ws, 'E2', excel_img)
        except Exception as e:
            print(f"Error creating template image: {e}")

        for col in ws.columns:
            if get_column_letter(col[0].column) == 'E':
                continue
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "Buyer_Master_Detailed_Template.xlsx" if with_details else "Buyer_Master_Standard_Template.xlsx"
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return response

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        """
        POST /api/buyer-masters/import-excel/
        Import Buyer Master records & auto-create missing Buyers.
        Does NOT insert into Sample table! Only populates Buyer & BuyerMaster tables.
        Also extracts high-quality finishing images embedded in worksheet cells!
        """

        excel_file = request.FILES.get('file')
        if not excel_file:
            return Response({
                "error_type": "No File Uploaded",
                "detail": "Please select a valid .xlsx file to upload."
            }, status=status.HTTP_400_BAD_REQUEST)

        if not (excel_file.name.lower().endswith('.xlsx')):
            return Response({
                "error_type": "Invalid Extension",
                "detail": "File must be an Excel (.xlsx) spreadsheet. Please download the expected template below."
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
        except Exception:
            return Response({
                "error_type": "Unreadable File",
                "detail": "Unable to read Excel file. Please ensure the file is not corrupt."
            }, status=status.HTTP_400_BAD_REQUEST)

        # Scan rows 1 to 15 to dynamically discover the actual header row
        header_row_idx = 1
        best_match_count = -1
        rows_sample = list(ws.iter_rows(min_row=1, max_row=15))

        keywords_to_check = [
            'buyer code', 'bc code', 'buyer', 'code', 'style no', 'pc style', 'style',
            'désignation', 'designation', 'product name', 'product', 'name', 'picture',
            'material', 'finish', 'size', 'cbm', 'usd factory', 'price', 'qty', 'quantity', 'remark'
        ]

        for r_idx, row_cells in enumerate(rows_sample, start=1):
            row_str = " ".join([str(c.value or '').lower() for c in row_cells])
            matches = sum(1 for kw in keywords_to_check if kw in row_str)
            if matches > best_match_count and matches >= 2:
                best_match_count = matches
                header_row_idx = r_idx

        header_cells = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx))[0]
        header_row = [str(cell.value or '').strip() for cell in header_cells]

        header_map = {}
        for idx, col in enumerate(header_row):
            col_norm = col.lower().replace('.', '').replace('*', '').replace('_', ' ').replace('/', ' ').strip()
            if col_norm:
                header_map[col_norm] = idx

        # Check subheader row (if present)
        subheader_row = []
        if ws.max_row >= header_row_idx + 1:
            sub_cells = list(ws.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 1))[0]
            subheader_row = [str(cell.value or '').strip() for cell in sub_cells]

        def find_col_index(aliases):
            for a in aliases:
                a_norm = a.lower().replace('.', '').replace('*', '').replace('_', ' ').replace('/', ' ').strip()
                if a_norm in header_map:
                    return header_map[a_norm]
            # Partial search
            for k, idx in header_map.items():
                for a in aliases:
                    a_norm = a.lower().replace('.', '').replace('*', '').replace('_', ' ').replace('/', ' ').strip()
                    if a_norm and a_norm in k:
                        return idx
            return None

        buyer_code_idx = find_col_index(['bc code', 'buyer code', 'buyer_code', 'code', 'buyer cde', 'bccode'])
        buyer_name_idx = find_col_index(['buyer name', 'buyer_name', 'buyer', 'client', 'customer'])
        style_idx = find_col_index(['pc style', 'style no', 'style_no', 'style #', 'style', 'sample id', 'sample_id', 'pc_style', 'pcstyle'])
        designation_idx = find_col_index(['désignation', 'designation'])
        product_name_idx = find_col_index(['product name', 'product', 'item name', 'product_name', 'name'])

        if buyer_code_idx is None and buyer_name_idx is not None:
            buyer_code_idx = buyer_name_idx
        elif buyer_name_idx is None and buyer_code_idx is not None:
            buyer_name_idx = buyer_code_idx

        if style_idx is None or (product_name_idx is None and designation_idx is None):
            return Response({
                "error_type": "Header / Schema Mismatch",
                "detail": "Required headers ('PC Style' / 'Style No.' or 'Product Name' / 'Désignation') could not be identified in the Excel file."
            }, status=status.HTTP_400_BAD_REQUEST)

        wood_idx = find_col_index(['material', 'wood type', 'wood', 'material wood type', 'material   wood', 'wood_type'])
        finish_idx = find_col_index(['finish', 'finish color', 'finish_color', 'color', 'finish   color'])
        usd_idx = find_col_index(['usd factory', 'price usd ($)', 'price usd', 'usd ($)', 'usd', 'factory price', 'price', 'usd_factory', 'factory usd'])
        units_idx = find_col_index(['qty.', 'qty', 'units', 'unit', 'quantity'])
        cbm_idx = find_col_index(['cbm', 'total cbm'])
        remark_idx = find_col_index(['remark', 'remarks', 'note', 'notes'])

        vendor_details_idx = find_col_index(['vendor details', 'vendor', 'supplier'])
        vendor_price_idx = find_col_index(['vendor price'])
        costing_idx = find_col_index(['costing'])
        purchase_price_idx = find_col_index(['purchase price'])
        net_wt_idx = find_col_index(['net weight'])
        gross_wt_idx = find_col_index(['gross weight'])
        box_l_idx = find_col_index(['box length (cm)', 'box length'])
        box_b_idx = find_col_index(['box breadth (cm)', 'box breadth'])
        box_h_idx = find_col_index(['box height (cm)', 'box height'])

        # Detailed Size Resolution
        len_idx = find_col_index(['size length (cm)', 'size length', 'length (cm)', 'length', 'size l', 'width (cm)'])
        breadth_idx = find_col_index(['size breadth (cm)', 'size breadth', 'breadth (cm)', 'breadth', 'size b', 'depth (cm)'])
        height_idx = find_col_index(['size height (cm)', 'size height', 'height (cm)', 'height', 'size h'])

        size_composite_idx = find_col_index(['size cms', 'size cm', 'size cms (w x d x h)', 'size (w x d x h)', 'size cms (l x w x h)', 'size'])

        if size_composite_idx is not None and (len_idx is None or breadth_idx is None or height_idx is None):
            len_idx = size_composite_idx
            breadth_idx = size_composite_idx + 1 if size_composite_idx + 1 < len(header_row) else None
            height_idx = size_composite_idx + 2 if size_composite_idx + 2 < len(header_row) else None

        row_images = {}
        if hasattr(ws, '_images'):
            for img in ws._images:
                try:
                    img_row = img.anchor._from.row + 1
                    if img_row not in row_images:
                        row_images[img_row] = []
                    row_images[img_row].append(img)
                except Exception:
                    pass

        imported_count = 0
        updated_count = 0
        buyers_created = 0
        samples_created = 0
        images_extracted = 0

        buyers_by_code = {b.code.lower(): b for b in Buyer.objects.all()}
        buyers_by_name = {b.name.lower(): b for b in Buyer.objects.all()}

        def parse_dec(val):
            if val is None or val == '': return None
            try:
                clean_str = str(val).replace('$', '').replace('₹', '').replace(',', '').strip()
                return float(clean_str)
            except ValueError:
                return None

        # Determine data start row: skip header and subheader if applicable
        start_row = header_row_idx + 1
        if subheader_row and any(w in ' '.join(subheader_row).lower() for w in ['w', 'd', 'h', 'cm', 'l', 'b']):
            start_row = header_row_idx + 2

        for excel_row_num, row_cells in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
            cells = [cell.value for cell in row_cells]
            if not any(cells):
                continue

            b_code_val = str(cells[buyer_code_idx]).strip() if buyer_code_idx is not None and buyer_code_idx < len(cells) and cells[buyer_code_idx] is not None else ''
            style_no_val = str(cells[style_idx]).strip() if style_idx is not None and style_idx < len(cells) and cells[style_idx] is not None else ''

            desig_val = str(cells[designation_idx]).strip() if designation_idx is not None and designation_idx < len(cells) and cells[designation_idx] is not None else ''
            prod_val = str(cells[product_name_idx]).strip() if product_name_idx is not None and product_name_idx < len(cells) and cells[product_name_idx] is not None else ''
            
            product_name_val = desig_val or prod_val or style_no_val

            if not style_no_val or not product_name_val:
                continue

            b_name_val = str(cells[buyer_name_idx]).strip() if buyer_name_idx is not None and buyer_name_idx < len(cells) and cells[buyer_name_idx] is not None else (b_code_val or 'Default Buyer')
            if not b_code_val:
                b_code_val = b_name_val

            buyer_obj = buyers_by_code.get(b_code_val.lower()) or buyers_by_name.get(b_name_val.lower())
            
            if buyer_obj:
                if buyer_obj.is_deleted:
                    buyer_obj.is_deleted = False
                    buyer_obj.save()
            else:
                try:
                    buyer_obj = Buyer.objects.create(code=b_code_val, name=b_name_val)
                    buyers_created += 1
                except Exception:
                    buyer_obj = Buyer.objects.filter(code__iexact=b_code_val).first()
                    if buyer_obj:
                        if buyer_obj.is_deleted:
                            buyer_obj.is_deleted = False
                            buyer_obj.save()
                    else:
                        unique_code = f"{b_code_val}_{uuid.uuid4().hex[:4]}"
                        buyer_obj = Buyer.objects.create(code=unique_code, name=b_name_val)
                        buyers_created += 1

                buyers_by_code[b_code_val.lower()] = buyer_obj
                buyers_by_name[b_name_val.lower()] = buyer_obj

            wood_val = str(cells[wood_idx]).strip() if wood_idx is not None and wood_idx < len(cells) and cells[wood_idx] is not None else ''
            finish_val = str(cells[finish_idx]).strip() if finish_idx is not None and finish_idx < len(cells) and cells[finish_idx] is not None else ''
            remark_val = str(cells[remark_idx]).strip() if remark_idx is not None and remark_idx < len(cells) and cells[remark_idx] is not None else ''
            v_details_val = str(cells[vendor_details_idx]).strip() if vendor_details_idx is not None and vendor_details_idx < len(cells) and cells[vendor_details_idx] is not None else ''

            # Size parsing logic
            size_len = None
            size_brd = None
            size_hgt = None

            raw_len_cell = cells[len_idx] if len_idx is not None and len_idx < len(cells) else None
            if raw_len_cell is not None and isinstance(raw_len_cell, str):
                match = re.search(r'(\d+(?:\.\d+)?)\s*[\s*xX×,-]\s*(\d+(?:\.\d+)?)\s*[\s*xX×,-]\s*(\d+(?:\.\d+)?)', raw_len_cell)
                if match:
                    size_len = parse_dec(match.group(1))
                    size_brd = parse_dec(match.group(2))
                    size_hgt = parse_dec(match.group(3))

            if size_len is None:
                size_len = parse_dec(raw_len_cell)
                size_brd = parse_dec(cells[breadth_idx] if breadth_idx is not None and breadth_idx < len(cells) else None)
                size_hgt = parse_dec(cells[height_idx] if height_idx is not None and height_idx < len(cells) else None)

            price_usd = parse_dec(cells[usd_idx] if usd_idx is not None and usd_idx < len(cells) else None)
            units_val = int(parse_dec(cells[units_idx]) or 1) if units_idx is not None and units_idx < len(cells) else 1
            cbm_val = parse_dec(cells[cbm_idx] if cbm_idx is not None and cbm_idx < len(cells) else None)

            v_price = parse_dec(cells[vendor_price_idx] if vendor_price_idx is not None and vendor_price_idx < len(cells) else None)
            costing_val = parse_dec(cells[costing_idx] if costing_idx is not None and costing_idx < len(cells) else None)
            pur_price = parse_dec(cells[purchase_price_idx] if purchase_price_idx is not None and purchase_price_idx < len(cells) else None)
            net_wt = parse_dec(cells[net_wt_idx] if net_wt_idx is not None and net_wt_idx < len(cells) else None)
            gross_wt = parse_dec(cells[gross_wt_idx] if gross_wt_idx is not None and gross_wt_idx < len(cells) else None)
            box_l = parse_dec(cells[box_l_idx] if box_l_idx is not None and box_l_idx < len(cells) else None)
            box_b = parse_dec(cells[box_b_idx] if box_b_idx is not None and box_b_idx < len(cells) else None)
            box_h = parse_dec(cells[box_h_idx] if box_h_idx is not None and box_h_idx < len(cells) else None)

            # Check if Sample already exists
            sample_obj = Sample.objects.filter(
                Q(style_no__iexact=style_no_val) | Q(sample_id__iexact=style_no_val)
            ).first()

            if not sample_obj:
                base_sid = style_no_val
                unique_sid = base_sid
                if Sample.objects.filter(sample_id=unique_sid).exists():
                    unique_sid = f"{base_sid}-{b_code_val}"

                sample_obj = Sample.objects.create(
                    sample_id=unique_sid,
                    style_no=style_no_val,
                    product_name=product_name_val,
                    buyer=buyer_obj,
                    material=wood_val,
                    finish_color=finish_val,
                    size_length=size_len,
                    size_breadth=size_brd,
                    size_height=size_hgt,
                    usd=price_usd,
                    cbm=cbm_val,
                    remark=remark_val,
                )
                samples_created += 1
            else:
                # Update sample fields if missing
                if not sample_obj.product_name and product_name_val: sample_obj.product_name = product_name_val
                if not sample_obj.buyer and buyer_obj: sample_obj.buyer = buyer_obj
                if not sample_obj.material and wood_val: sample_obj.material = wood_val
                if not sample_obj.finish_color and finish_val: sample_obj.finish_color = finish_val
                if sample_obj.size_length is None and size_len is not None: sample_obj.size_length = size_len
                if sample_obj.size_breadth is None and size_brd is not None: sample_obj.size_breadth = size_brd
                if sample_obj.size_height is None and size_hgt is not None: sample_obj.size_height = size_hgt
                if sample_obj.usd is None and price_usd is not None: sample_obj.usd = price_usd
                if sample_obj.cbm is None and cbm_val is not None: sample_obj.cbm = cbm_val
                sample_obj.save()

            bm_obj, created = BuyerMaster.objects.get_or_create(
                buyer=buyer_obj,
                style_no=style_no_val,
                defaults={
                    'buyer_code': b_code_val,
                    'sample': sample_obj,
                    'product_name': product_name_val,
                    'wood_type': wood_val,
                    'finish_color': finish_val,
                    'size_length': size_len,
                    'size_breadth': size_brd,
                    'size_height': size_hgt,
                    'price_usd': price_usd,
                    'units': units_val,
                    'cbm': cbm_val,
                    'remark': remark_val,
                    'vendor_details': v_details_val,
                    'vendor_price': v_price,
                    'costing': costing_val,
                    'purchase_price': pur_price,
                    'net_weight': net_wt,
                    'gross_weight': gross_wt,
                    'box_length': box_l,
                    'box_breadth': box_b,
                    'box_height': box_h,
                }
            )

            if not created:
                bm_obj.product_name = product_name_val
                bm_obj.buyer_code = b_code_val
                if sample_obj: bm_obj.sample = sample_obj
                if wood_val: bm_obj.wood_type = wood_val
                if finish_val: bm_obj.finish_color = finish_val
                if size_len is not None: bm_obj.size_length = size_len
                if size_brd is not None: bm_obj.size_breadth = size_brd
                if size_hgt is not None: bm_obj.size_height = size_hgt
                if price_usd is not None: bm_obj.price_usd = price_usd
                if units_val is not None: bm_obj.units = units_val
                if cbm_val is not None: bm_obj.cbm = cbm_val
                if remark_val: bm_obj.remark = remark_val
                if v_details_val: bm_obj.vendor_details = v_details_val
                if v_price is not None: bm_obj.vendor_price = v_price
                if costing_val is not None: bm_obj.costing = costing_val
                if pur_price is not None: bm_obj.purchase_price = pur_price
                if net_wt is not None: bm_obj.net_weight = net_wt
                if gross_wt is not None: bm_obj.gross_weight = gross_wt
                if box_l is not None: bm_obj.box_length = box_l
                if box_b is not None: bm_obj.box_breadth = box_b
                if box_h is not None: bm_obj.box_height = box_h
                bm_obj.save()
                updated_count += 1
            else:
                imported_count += 1

            if excel_row_num in row_images:
                for img_idx, img_obj in enumerate(row_images[excel_row_num]):
                    try:
                        image_bytes = img_obj._data()
                        ext = img_obj.format if hasattr(img_obj, 'format') and img_obj.format else 'png'
                        file_name = f"BM_{style_no_val.replace('/', '_')}_{img_idx+1}.{ext}"
                        content_file = ContentFile(image_bytes, name=file_name)

                        BuyerMasterFinishingImage.objects.create(buyer_master=bm_obj, image=content_file)
                        images_extracted += 1

                        if sample_obj:
                            smp_file = ContentFile(image_bytes, name=f"SMP_{style_no_val.replace('/', '_')}_{img_idx+1}.{ext}")
                            SampleImage.objects.create(sample=sample_obj, image=smp_file)
                            if not sample_obj.image and img_idx == 0:
                                sample_obj.image = smp_file
                                sample_obj.save()
                    except Exception as img_err:
                        print(f"Error saving image for row {excel_row_num}: {img_err}")

        return Response({
            "detail": f"Import complete! {imported_count} new Buyer Master style(s) created, {updated_count} updated. {buyers_created} new Buyer(s) created. {samples_created} new Sample(s) registered. {images_extracted} finishing image(s) extracted.",
            "imported_count": imported_count,
            "updated_count": updated_count,
            "buyers_created": buyers_created,
            "samples_created": samples_created,
            "images_extracted": images_extracted,
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        buyer_id = request.query_params.get('buyer')
        if not buyer_id:
            return HttpResponse("Buyer ID is required", status=400)
        
        try:
            buyer = Buyer.objects.get(id=buyer_id)
        except Buyer.DoesNotExist:
            return HttpResponse("Buyer not found", status=404)
        
        masters = self.get_queryset().filter(buyer=buyer)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{buyer.code}_Buyer_Master"
        
        ws.views.sheetView[0].showGridLines = True
        
        with_details = request.query_params.get('with_details') == 'true'
        
        headers = [
            'S. No.', 'Buyer Name', 'Buyer Code', 'Style No', 'Sample ID', 'Picture', 'Product Name', 
            'Material', 'Finish', 'Size Length (cm)', 
            'Size Breadth (cm)', 'Size Height (cm)', 
            'Price USD', 'Units', 'Total CBM', 'Total Amount', 'Remark'
        ]
        
        if with_details:
            headers.extend([
                'Vendor Details', 'Vendor Price', 'Costing', 'Purchase Price', 
                'CBM', 'Net Weight', 'Gross Weight', 'Box Length (cm)', 'Box Breadth (cm)', 'Box Height (cm)', 'Box Size'
            ])
        
        ws.append(headers)
        
        header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="000000")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        border_medium = Side(style='medium', color='000000')
        header_border = Border(left=border_medium, right=border_medium, top=border_medium, bottom=border_medium)
        
        data_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        data_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = header_border
            
        temp_files = []
        
        for idx, bm in enumerate(masters, 1):
            row_idx = idx + 1
            ws.row_dimensions[row_idx].height = 80
            
            sample_id_val = ""
            sample_image_path = ""
            if bm.sample:
                sample_id_val = bm.sample.sample_id
                first_img = bm.sample.images.first()
                if first_img and first_img.image and os.path.exists(first_img.image.path):
                    sample_image_path = first_img.image.path
            
            row_data = [
                idx,
                buyer.name,
                buyer.code,
                bm.style_no or "",
                sample_id_val,
                "", # Picture cell
                bm.product_name or "",
                bm.wood_type or "",
                bm.finish_color or "",
                float(bm.size_length) if bm.size_length else "",
                float(bm.size_breadth) if bm.size_breadth else "",
                float(bm.size_height) if bm.size_height else "",
                float(bm.price_usd) if bm.price_usd else "",
                bm.units or 1,
                float(bm.total_cbm) if bm.total_cbm else "",
                float(bm.total_amount) if bm.total_amount else "",
                bm.remark or ""
            ]
            
            if with_details:
                row_data.extend([
                    bm.vendor_details or "",
                    float(bm.vendor_price) if bm.vendor_price else "",
                    float(bm.costing) if bm.costing else "",
                    float(bm.purchase_price) if bm.purchase_price else "",
                    float(bm.cbm) if bm.cbm else "",
                    float(bm.net_weight) if bm.net_weight else "",
                    float(bm.gross_weight) if bm.gross_weight else "",
                    float(bm.box_length) if bm.box_length else "",
                    float(bm.box_breadth) if bm.box_breadth else "",
                    float(bm.box_height) if bm.box_height else "",
                    bm.box_size or ""
                ])
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = data_align
                cell.border = data_border
                
            if sample_image_path:
                try:
                    pil_img = PILImage.open(sample_image_path)
                    if pil_img.mode in ('RGBA', 'LA', 'P'):
                        pil_img = pil_img.convert('RGB')
                    pil_img.thumbnail((90, 68))
                    
                    tmp_f = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False)
                    pil_img.save(tmp_f.name, format='JPEG', quality=85)
                    tmp_f.close()
                    temp_files.append(tmp_f.name)
                    
                    xl_img = OpenpyxlImage(tmp_f.name)
                    ws.column_dimensions['F'].width = 18
                    add_centered_image(ws, f"F{row_idx}", xl_img)
                except Exception as e:
                    print(f"Error drawing image: {e}")
                    
        ws.column_dimensions['F'].width = 18
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            if col_letter == 'F':
                continue
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{buyer.code}_Buyer_Master.xlsx"'
        wb.save(response)
        
        for f in temp_files:
            try:
                os.remove(f)
            except:
                pass
                
        return response



class SupplierViewSet(viewsets.ModelViewSet):
    """CRUD for Supplier master list."""
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update'):
            return [IsAuthenticated(), IsSupplierManager()]
        if self.action == 'destroy':
            return [IsAuthenticated(), IsAdmin()]
        return [IsAuthenticated()]


class SupplierPOViewSet(viewsets.ModelViewSet):
    """
    CRUD for Supplier Purchase Orders.
    Each PO goes to one supplier and has multiple line items
    referencing different buyer orders.
    """
    queryset = SupplierPO.objects.select_related(
        'supplier', 'supervisor', 'buyer_pi'
    ).prefetch_related(
        'items__buyer',
        'extension_logs',
        'supplier_history',
        'supplier_history__previous_supplier',
        'supplier_history__new_supplier',
        'supplier_history__changed_by'
    ).all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list':
            return SupplierPOListSerializer
        return SupplierPOSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.is_authenticated and user.role == 'supervisor':
            qs = qs.filter(supervisor=user)
        supplier_id = self.request.query_params.get('supplier')
        if supplier_id:
            qs = qs.filter(supplier_id=supplier_id)
        buyer_id = self.request.query_params.get('buyer')
        if buyer_id:
            qs = qs.filter(items__buyer_id=buyer_id).distinct()
        status_f = self.request.query_params.get('status')
        if status_f:
            qs = qs.filter(status=status_f)
        exclude_status = self.request.query_params.get('exclude_status')
        if exclude_status:
            qs = qs.exclude(status=exclude_status)
        return qs

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [IsAuthenticated(), IsAdmin()]
        return [IsAuthenticated()]

    def perform_update(self, serializer):
        old_po = self.get_object()
        new_supplier = serializer.validated_data.get('supplier')
        reason = self.request.data.get('supplier_change_reason', '')

        updated_po = serializer.save()

        if old_po.supplier and new_supplier and old_po.supplier != new_supplier:
            POSupplierHistory.objects.create(
                supplier_po=updated_po,
                previous_supplier=old_po.supplier,
                new_supplier=new_supplier,
                reason=reason or 'Supplier transferred',
                changed_by=self.request.user
            )

    @action(detail=True, methods=['post'], url_path='extend-due-date')
    def extend_due_date(self, request, pk=None):
        po = self.get_object()
        days_added = request.data.get('days_added')
        reason = request.data.get('reason', '')
        custom_new_date = request.data.get('new_due_date')

        if not po.original_due_date and po.due_date:
            po.original_due_date = po.due_date

        prev_due_date = po.due_date

        if custom_new_date:
            try:
                new_due_date = datetime.strptime(str(custom_new_date), '%Y-%m-%d').date()
                if prev_due_date:
                    days_added = (new_due_date - prev_due_date).days
                else:
                    days_added = 0
            except ValueError:
                return Response({'detail': 'Invalid date format. Use YYYY-MM-DD.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            try:
                days_added = int(days_added) if days_added is not None else 5
            except (ValueError, TypeError):
                days_added = 5

            if prev_due_date:
                new_due_date = prev_due_date + timedelta(days=days_added)
            else:
                new_due_date = timezone.now().date() + timedelta(days=days_added)

        po.due_date = new_due_date
        po.save()

        # Log extension
        POExtensionLog.objects.create(
            supplier_po=po,
            extended_by=request.user if request.user.is_authenticated else None,
            days_added=days_added,
            previous_due_date=prev_due_date,
            new_due_date=new_due_date,
            reason=reason
        )

        serializer = SupplierPOSerializer(po, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'], url_path='receive-installment')
    def receive_installment(self, request, pk=None):
        """
        Record a full multi-item delivery installment / round for a Purchase Order.
        A single truck delivery with 1 supplier invoice can contain multiple PO items.
        Generates 1 single combined Debit Note for all rejected SKUs in the round.
        """
        po = self.get_object()
        data = request.data

        raw_items = data.get('items', [])
        if isinstance(raw_items, str):
            try:
                items_payload = json.loads(raw_items)
            except Exception:
                items_payload = []
        else:
            items_payload = raw_items

        if not items_payload:
            return Response({'detail': 'No line items provided for this delivery installment.'}, status=status.HTTP_400_BAD_REQUEST)

        invoice_no = str(data.get('supplier_invoice_no', '')).strip()
        invoice_date = data.get('supplier_invoice_date') or timezone.now().date()
        invoice_amount = Decimal(str(data.get('supplier_invoice_amount', 0))) if data.get('supplier_invoice_amount') else None
        vehicle_no = str(data.get('vehicle_no', '')).strip()
        driver_contact = str(data.get('driver_contact', '')).strip()
        receipt_date = data.get('receipt_date') or timezone.now().date()
        notes = str(data.get('notes', '')).strip()

        # Compute next Round Number and GRN Number for this PO
        existing_grns = list(GateInwardReceipt.objects.filter(supplier_po=po).values_list('grn_number', flat=True).distinct())
        existing_grns = [g for g in existing_grns if g]
        round_number = len(existing_grns) + 1
        grn_number = f"GRN-{po.po_number}-R{round_number}"

        created_receipts = []
        all_rejected_items = []

        for idx, item_data in enumerate(items_payload):
            po_item_id = item_data.get('po_item')
            passed_qty = Decimal(str(item_data.get('passed_qty', 0)))
            rejected_qty = Decimal(str(item_data.get('rejected_qty', 0)))

            if passed_qty == 0 and rejected_qty == 0:
                continue

            po_item = po.items.filter(id=po_item_id).first()
            if not po_item:
                continue

            # Update PO Item passed quantity
            po_item.passed_quantity = (po_item.passed_quantity or Decimal('0')) + passed_qty
            po_item.save()

            # Create Inward Receipt record for this item
            receipt = GateInwardReceipt.objects.create(
                grn_number=grn_number,
                round_number=round_number,
                supplier_po=po,
                po_item=po_item,
                receipt_date=receipt_date,
                challan_no=invoice_no,
                supplier_invoice_no=invoice_no,
                supplier_invoice_date=invoice_date,
                supplier_invoice_amount=invoice_amount,
                vehicle_no=vehicle_no,
                driver_contact=driver_contact,
                received_qty=passed_qty + rejected_qty,
                passed_qty=passed_qty,
                rejected_qty=rejected_qty,
                notes=notes or item_data.get('remark', ''),
                inspected_by=request.user if request.user.is_authenticated else None
            )
            created_receipts.append(receipt)

            # Raw Stock Addition
            if passed_qty > 0:
                words = po_item.description.split()
                style = words[0] if words else "RAW-ITEM"
                item_name = po_item.description[:100] if po_item.description else "Raw Furniture Item"

                StockItem.objects.create(
                    stock_type='raw',
                    po_item=po_item,
                    buyer=po_item.buyer,
                    style_no=style,
                    item_name=item_name,
                    quantity=passed_qty,
                    unit=po_item.unit or 'pcs',
                    unit_price=po_item.rate,
                    location='Main Gate Raw Store',
                    status='In Stock',
                    remarks=f"GRN #{grn_number} (Round #{round_number}) via Inv #{invoice_no or 'N/A'}"
                )

            # Rejections Logging & Defect Images
            if rejected_qty > 0:
                remark = item_data.get('remark') or notes or 'Gate inspection rejected pieces'
                defect_file = request.FILES.get(f'defect_image_{po_item_id}') or request.FILES.get(f'defect_image_{idx}')
                
                defect_obj = SupplierPOItemDefect.objects.create(
                    po_item=po_item,
                    reported_by=request.user if request.user.is_authenticated else None,
                    quantity=rejected_qty,
                    defective_image=defect_file,
                    remark=remark
                )
                if defect_file:
                    SupplierPOItemDefectImage.objects.create(
                        defect=defect_obj,
                        image=defect_file
                    )

                unit_rate = po_item.rate or Decimal('0')
                subtotal = round(rejected_qty * unit_rate, 2)
                all_rejected_items.append({
                    'po_item': po_item,
                    'rejected_qty': rejected_qty,
                    'unit_rate': unit_rate,
                    'subtotal': subtotal,
                    'remark': remark
                })

        # Generate ONE Combined Debit Note for all rejected SKUs in this shipment round
        generated_debit_notes = []
        if all_rejected_items:
            first_receipt_hex = created_receipts[0].id.hex[:4].upper() if created_receipts else "0000"
            dn_number = f"DN-{po.po_number}-{first_receipt_hex}"

            subtotal_all = sum(r['subtotal'] for r in all_rejected_items)
            cartage_gst = round(subtotal_all * Decimal('0.18') * Decimal('0.03'), 2)
            cgst = round(subtotal_all * Decimal('0.09'), 2)
            sgst = round(subtotal_all * Decimal('0.09'), 2)

            raw_total = subtotal_all + cartage_gst + cgst + sgst
            final_amount = Decimal(str(round(raw_total)))
            round_off = final_amount - raw_total
            words_str = num2words(float(final_amount), lang='en_IN').title() + " Rupees Only"

            total_rejected_pcs = sum(r['rejected_qty'] for r in all_rejected_items)

            dn = SupplierDebitNote.objects.create(
                vch_no=dn_number,
                vch_date=receipt_date,
                original_inv_no=invoice_no or po.po_number,
                original_inv_date=invoice_date,
                supplier=po.supplier,
                supplier_po=po,
                po_item=all_rejected_items[0]['po_item'],
                status='Issued',
                item_description=f"Combined Rejection ({len(all_rejected_items)} SKUs) — GRN #{grn_number} (Round #{round_number})",
                rejected_qty=total_rejected_pcs,
                unit='pcs',
                rate=Decimal('0'),
                subtotal_amount=subtotal_all,
                cartage_gst_rate=Decimal('18.0'),
                cartage_gst_amount=cartage_gst,
                cgst_rate=Decimal('9.0'),
                cgst_amount=cgst,
                sgst_rate=Decimal('9.0'),
                sgst_amount=sgst,
                round_off=round_off,
                total_amount=final_amount,
                amount_in_words=words_str,
                remarks=f"BEING AMOUNT DEBITED FOR REJECTED GOODS IN GRN #{grn_number} (ROUND #{round_number})",
                company_pan='ABXPS4077R'
            )

            # Create line items for each rejected SKU in the Debit Note
            for r_item in all_rejected_items:
                SupplierDebitNoteItem.objects.create(
                    debit_note=dn,
                    po_item=r_item['po_item'],
                    description=r_item['po_item'].description,
                    hsn_sac='9403',
                    rejected_qty=r_item['rejected_qty'],
                    unit=r_item['po_item'].unit or 'pcs',
                    rate=r_item['unit_rate'],
                    amount=r_item['subtotal'],
                    reason=r_item['remark']
                )

            generated_debit_notes.append(SupplierDebitNoteSerializer(dn).data)

        # Update PO Status
        all_items = po.items.all()
        total_ordered = sum(it.quantity or Decimal('0') for it in all_items)
        total_passed = sum(it.passed_quantity or Decimal('0') for it in all_items)

        if total_passed >= total_ordered and total_ordered > 0:
            po.status = 'Received'
        elif total_passed > 0:
            po.status = 'Partial Received'
        po.save()

        return Response({
            'detail': f'Recorded Delivery Round #{round_number} ({grn_number}) with {len(created_receipts)} item(s).',
            'grn_number': grn_number,
            'round_number': round_number,
            'po_status': po.status,
            'receipts': GateInwardReceiptSerializer(created_receipts, many=True).data,
            'debit_notes': generated_debit_notes
        }, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'], url_path='pdf')
    def download_pdf(self, request, pk=None):
        """
        Low-level canvas PDF that matches the reference Purchase Order layout.
        Uses absolute positioning exclusively — no Platypus Tables.
        Business logic / queries / serializers are untouched.
        """

        po  = self.get_object()
        buf = BytesIO()

        # ── Page setup ────────────────────────────────────────────────────────
        PW, PH = A4          # 595.28 × 841.89 pt
        ML = 13.0 * mm       # left margin
        MR = 13.0 * mm       # right margin
        MT = 10.0 * mm       # top margin
        MB = 10.0 * mm       # bottom margin
        CW = PW - ML - MR    # content width

        c = rl_canvas.Canvas(buf, pagesize=A4)
        c.setTitle(f'Purchase Order {po.po_number}')

        # ── Drawing primitives ─────────────────────────────────────────────────
        def ds(x, y, text, font='Helvetica', size=8):
            c.setFont(font, size)
            c.setFillColor(colors.black)
            c.drawString(x, y, str(text))

        def dr(x, y, text, font='Helvetica', size=8):
            c.setFont(font, size)
            c.setFillColor(colors.black)
            c.drawRightString(x, y, str(text))

        def dc(x, y, text, font='Helvetica', size=8):
            c.setFont(font, size)
            c.setFillColor(colors.black)
            c.drawCentredString(x, y, str(text))

        def hline(x1, y, x2, lw=0.5):
            c.setLineWidth(lw)
            c.setStrokeColor(colors.black)
            c.line(x1, y, x2, y)

        def vline(x, y1, y2, lw=0.5):
            c.setLineWidth(lw)
            c.setStrokeColor(colors.black)
            c.line(x, y1, x, y2)

        def box(x, y, w, h, lw=0.75):
            c.setLineWidth(lw)
            c.setStrokeColor(colors.black)
            c.rect(x, y, w, h, stroke=1, fill=0)

        def sw(text, font, size):
            c.setFont(font, size)
            return c.stringWidth(str(text), font, size)

        def wrap_line(text, font, size, max_w):
            """Split one paragraph of text into lines that fit within max_w."""
            words = str(text).split()
            lines, cur = [], ''
            for word in words:
                candidate = (cur + ' ' + word).strip()
                if sw(candidate, font, size) <= max_w:
                    cur = candidate
                else:
                    if cur:
                        lines.append(cur)
                    cur = word
            if cur:
                lines.append(cur)
            return lines if lines else ['']

        # ── Business data ──────────────────────────────────────────────────────
        CNAME  = 'PINKCITY ENTERPRISES'
        CADDR1 = 'G-78, EPIP, Indl. Area Sitapura,'
        CADDR2 = 'JAIPUR'
        CIEC   = 'IEC CODE : 1397002620'
        CGSTIN = 'GSTIN/UIN: 08ABXPS4077R1Z8'
        CSTATE = 'State Name : Rajasthan, Code : 08'
        CPAN   = 'ABXPS4077R'

        sup = po.supplier

        def fmt_s(d):   # "15-Jul-26"
            return d.strftime('%d-%b-%y').lstrip('0') if d else ''

        def fmt_l(d):   # "05 Sept 2026"
            return d.strftime('%d %b %Y') if d else ''

        po_date_str  = fmt_s(po.po_date)
        due_date_str = fmt_l(po.due_date)

        items_qs  = list(po.items.select_related('buyer').all())
        total_amt = sum(it.amount or Decimal('0') for it in items_qs)

        try:
            ip = int(total_amt)
            dp = int(round((total_amt - Decimal(str(ip))) * 100))
            ww = num2words(ip, lang='en').replace(',', '').title()
            if dp:
                ww += f' And {num2words(dp, lang="en").title()} Paise'
            words_text = f'INR {ww} Only'
        except Exception:
            words_text = f'Rs. {float(total_amt):,.2f}'

        # ── Layout constants (all in pt) ───────────────────────────────────────
        P   = 2.0 * mm    # general inner padding

        # ─ Bottom sections (built upward from MB) ─
        FOOTER_Y  = MB + 3.0 * mm
        DECL_H    = 44.0 * mm
        DECL_BOT  = MB + 8.0 * mm
        DECL_TOP  = DECL_BOT + DECL_H
        WORDS_H   = 17.0 * mm
        WORDS_BOT = DECL_TOP
        WORDS_TOP = WORDS_BOT + WORDS_H

        # ─ Header ─
        TITLE_Y  = PH - MT - 4.5 * mm      # baseline of PURCHASE ORDER text
        HDR_TOP  = TITLE_Y - 5.5 * mm
        HDR_H    = 68.0 * mm
        HDR_BOT  = HDR_TOP - HDR_H

        # ─ Items table ─
        ITEM_TOP = HDR_BOT
        ITEM_BOT = WORDS_TOP
        ITEM_H   = ITEM_TOP - ITEM_BOT

        # ─ Header left / right split ─
        LEFT_W  = CW * 0.48
        RIGHT_W = CW * 0.52
        SPX     = ML + LEFT_W              # x of the vertical divider in header

        # ─ Items column layout (proportions must add to 1) ─
        col_pct = [0.05, 0.56, 0.11, 0.10, 0.05, 0.13]
        col_w   = [CW * p for p in col_pct]
        col_x   = []
        _cx = ML
        for _w in col_w:
            col_x.append(_cx)
            _cx += _w

        ITEM_HDR_H = 8.5 * mm
        ITEM_ROW_H = 6.5 * mm

        # ═══════════════════════════════════════════════════════════════════════
        # 1. TITLE
        # ═══════════════════════════════════════════════════════════════════════
        dc(PW / 2, TITLE_Y, 'PURCHASE ORDER', 'Helvetica-Bold', 18)

        # ═══════════════════════════════════════════════════════════════════════
        # 2. HEADER BOX
        # ═══════════════════════════════════════════════════════════════════════
        box(ML, HDR_BOT, CW, HDR_H, lw=0.75)     # outer border
        vline(SPX, HDR_BOT, HDR_TOP)              # left | right divider

        # ── LEFT column ──────────────────────────────────────────────────────
        LX      = ML + P
        LOGO_SIZE = 16.0 * mm
        logo_x  = ML + P
        logo_y  = HDR_TOP - P - LOGO_SIZE

        # Draw PNG Logo image if exists, otherwise fallback to stylized logo
        logo_drawn = False
        logo_path = r"C:\Users\User\OneDrive\Desktop\ERP Furniture\Frontend\src\assets\Pinkcity_Logo.png"
        if os.path.exists(logo_path):
            try:
                c.drawImage(logo_path, logo_x, logo_y, width=LOGO_SIZE, height=LOGO_SIZE, mask='auto')
                logo_drawn = True
            except Exception:
                pass

        if not logo_drawn:
            # Stylised 'pe' logo box fallback
            LOGO_W = 16.0 * mm
            LOGO_H = 14.0 * mm
            fallback_logo_y = HDR_TOP - P - LOGO_H
            c.setLineWidth(1.0)
            c.rect(logo_x, fallback_logo_y, LOGO_W, LOGO_H, stroke=1, fill=0)
            bar_y = fallback_logo_y + LOGO_H - 3.2*mm
            c.setLineWidth(1.2)
            c.line(logo_x + 6.5*mm, bar_y, logo_x + LOGO_W - 1.0*mm, bar_y)
            c.setLineWidth(0.5)
            ds(logo_x + 1.0*mm, fallback_logo_y + 3.5*mm,  'p', 'Helvetica-Bold', 11)
            ds(logo_x + 6.5*mm, fallback_logo_y + 3.5*mm,  'e', 'Helvetica-Bold', 11)

        # "Invoice To" label beside logo
        info_x = logo_x + LOGO_SIZE + 3.0 * mm
        ds(info_x, HDR_TOP - 3.0 * mm, 'Invoice To', 'Helvetica', 7)

        # Company details beside logo (Invoice To block)
        ds(info_x, HDR_TOP - 6.5 * mm, CNAME, 'Helvetica-Bold', 10)
        ds(info_x, HDR_TOP - 10.5 * mm, CADDR1, 'Helvetica', 8)
        ds(info_x, HDR_TOP - 14.0 * mm, CADDR2, 'Helvetica', 8)
        ds(info_x, HDR_TOP - 17.5 * mm, CIEC, 'Helvetica', 8)
        ds(info_x, HDR_TOP - 21.0 * mm, CGSTIN, 'Helvetica', 8)
        ds(info_x, HDR_TOP - 24.5 * mm, CSTATE, 'Helvetica', 8)

        # Horizontal divider between company and supplier (exactly halfway)
        div_y = HDR_TOP - 34.0 * mm
        hline(ML, div_y, SPX, lw=0.5)

        # Supplier block (bottom half)
        ds(LX, div_y - 3.0 * mm, 'Supplier (Bill from)', 'Helvetica', 7)
        ds(LX, div_y - 6.5 * mm, sup.name, 'Helvetica-Bold', 10)

        cy = div_y - 10.5 * mm
        sup_max_w = LEFT_W - 4.0 * mm
        lines_to_draw = []
        if sup.address:
            for addr_seg in sup.address.split('\n'):
                addr_seg = addr_seg.strip()
                if addr_seg:
                    lines_to_draw.extend(wrap_line(addr_seg, 'Helvetica', 8, sup_max_w))
        if sup.phone:
            for ph in sup.phone.split('\n'):
                ph = ph.strip()
                if ph:
                    prefix = "" if (ph.startswith("M.") or ph.startswith("M. ")) else "M."
                    lines_to_draw.append(f"{prefix}{ph}")
        if sup.gstin:
            lines_to_draw.append(f"GSTIN/UIN   : {sup.gstin}")
        if sup.state_name:
            lines_to_draw.append(f"State Name  : {sup.state_name}")

        spacing = 3.2 * mm if len(lines_to_draw) > 5 else 3.5 * mm
        for ln in lines_to_draw:
            if cy >= HDR_BOT + 1.5 * mm:
                ds(LX, cy, ln, 'Helvetica', 8)
                cy -= spacing

        # ── RIGHT column grid (4 equal rows) ──────────────────────────────────
        row_h = 17.0 * mm
        row1_bot = HDR_TOP - row_h
        row2_bot = HDR_TOP - 2 * row_h
        row3_bot = HDR_TOP - 3 * row_h
        row4_bot = HDR_TOP - 4 * row_h  # HDR_BOT

        # Draw horizontal lines for the right grid
        hline(SPX, row1_bot, SPX + RIGHT_W)
        hline(SPX, row2_bot, SPX + RIGHT_W)
        hline(SPX, row3_bot, SPX + RIGHT_W)

        # Draw vertical lines for specific split rows
        # Row 1 is split ~55/45
        vline(SPX + RIGHT_W * 0.55, row1_bot, HDR_TOP)
        # Row 3 is split ~75/25
        vline(SPX + RIGHT_W * 0.75, row3_bot, row2_bot)

        R_LABEL_H = 3.8 * mm
        R_VALUE_BOT = 2.0 * mm

        # Row 1: Purchase Order No. & Dated
        ds(SPX + P, HDR_TOP - R_LABEL_H, 'Purchase Order No.', 'Helvetica', 7)
        ds(SPX + P, row1_bot + R_VALUE_BOT, po.po_number, 'Helvetica-Bold', 10)
        ds(SPX + RIGHT_W * 0.55 + P, HDR_TOP - R_LABEL_H, 'Dated', 'Helvetica', 7)
        ds(SPX + RIGHT_W * 0.55 + P, row1_bot + R_VALUE_BOT, po_date_str, 'Helvetica-Bold', 10)

        # Row 2: Mode/Terms of Payment
        ds(SPX + P, row1_bot - R_LABEL_H, 'Mode/Terms of Payment', 'Helvetica', 7)
        ds(SPX + P, row2_bot + R_VALUE_BOT, po.mode_of_payment or '', 'Helvetica-Bold', 10)

        # Row 3: PO Due Date & Supervisor
        ds(SPX + P, row2_bot - R_LABEL_H, 'PO Due Date', 'Helvetica', 7)
        due_date_val = f"{due_date_str}, No Delay Please" if due_date_str else ""
        ds(SPX + P, row3_bot + R_VALUE_BOT, due_date_val, 'Helvetica-Bold', 9)
        ds(SPX + RIGHT_W * 0.75 + P, row2_bot - R_LABEL_H, 'Supervisor', 'Helvetica', 7)
        ds(SPX + RIGHT_W * 0.75 + P, row3_bot + R_VALUE_BOT, po.supervisor or '', 'Helvetica-Bold', 10)

        # Row 4: Terms of Delivery
        ds(SPX + P, row3_bot - R_LABEL_H, 'Terms of Delivery', 'Helvetica', 7)
        delivery_lines = []
        if po.terms_of_delivery:
            for ln in po.terms_of_delivery.split('\n'):
                ln = ln.strip()
                if ln:
                    delivery_lines.extend(wrap_line(ln, 'Helvetica-Bold', 10, RIGHT_W - 4.0 * mm))
        if po.nku_refs:
            for ln in po.nku_refs.split('\n'):
                ln = ln.strip()
                if ln:
                    delivery_lines.extend(wrap_line(ln, 'Helvetica-Bold', 10, RIGHT_W - 4.0 * mm))

        dy_deliv = row3_bot - 8.0 * mm
        for d_ln in delivery_lines[:2]:
            ds(SPX + P, dy_deliv, d_ln, 'Helvetica-Bold', 10)
            dy_deliv -= 4.0 * mm


        # ═══════════════════════════════════════════════════════════════════════
        # 3. ITEMS TABLE
        # ═══════════════════════════════════════════════════════════════════════
        box(ML, ITEM_BOT, CW, ITEM_H, lw=0.75)   # outer box

        # Vertical column dividers run the full height of the table
        for ci in range(1, len(col_x)):
            vline(col_x[ci], ITEM_BOT, ITEM_TOP)

        # Header row (bottom border is the separator)
        HDR_ROW_BOT = ITEM_TOP - ITEM_HDR_H
        hline(ML, HDR_ROW_BOT, ML + CW, lw=0.75)

        # Header labels
        ds(col_x[0] + 1.0*mm, HDR_ROW_BOT + 4.5*mm, 'Sl',  'Helvetica', 8)
        ds(col_x[0] + 1.0*mm, HDR_ROW_BOT + 1.5*mm, 'No.', 'Helvetica', 8)
        dc(col_x[1] + col_w[1]/2, HDR_ROW_BOT + 2.5*mm,
           'Description of Goods', 'Helvetica', 8)
        dc(col_x[2] + col_w[2]/2, HDR_ROW_BOT + 2.5*mm, 'Quantity', 'Helvetica', 8)
        dc(col_x[3] + col_w[3]/2, HDR_ROW_BOT + 2.5*mm, 'Rate',     'Helvetica', 8)
        dc(col_x[4] + col_w[4]/2, HDR_ROW_BOT + 2.5*mm, 'per',      'Helvetica', 8)
        dc(col_x[5] + col_w[5]/2, HDR_ROW_BOT + 2.5*mm, 'Amount',   'Helvetica', 8)

        # Data rows (only where actual items exist)
        IY = HDR_ROW_BOT
        for idx, item in enumerate(items_qs, 1):
            buyer_ref = f' [{item.buyer.name}]' if item.buyer else ''
            raw_desc  = str(item.description) + buyer_ref
            desc_lines = []
            for para in raw_desc.split('\n'):
                desc_lines.extend(
                    wrap_line(para.strip(), 'Helvetica-Bold', 9, col_w[1] - 2.5*mm)
                )
            if item.remark:
                for para in str(item.remark).split('\n'):
                    desc_lines.extend(
                        wrap_line(para.strip(), 'Helvetica-Bold', 9, col_w[1] - 2.5*mm)
                    )

            row_h   = max(ITEM_ROW_H, len(desc_lines) * 3.5*mm + 2.0*mm)
            row_bot = IY - row_h

            # Row bottom separator (thin, only between real rows)
            # NO horizontal lines between items according to reference
            # hline(ML, row_bot, ML + CW, lw=0.5)

            # SI number – vertically centred in row
            mid_y = IY - 4.0*mm
            dc(col_x[0] + col_w[0]/2, mid_y, str(idx), 'Helvetica', 9)

            # Description – top-aligned, multi-line
            dly = IY - 1.0*mm
            for dl in desc_lines:
                dly -= 4.0*mm
                ds(col_x[1] + 1.0*mm, dly, dl, 'Helvetica-Bold', 9)

            # Quantity
            qty_str = f'{float(item.quantity):.2f} {item.unit}'
            dr(col_x[2] + col_w[2] - 1.0*mm, mid_y, qty_str, 'Helvetica-Bold', 9)

            # Rate (right-aligned inside column)
            dr(col_x[3] + col_w[3] - 1.0*mm, mid_y,
               f'{float(item.rate):.2f}', 'Helvetica', 9)

            # Per
            dc(col_x[4] + col_w[4]/2, mid_y, str(item.unit), 'Helvetica', 9)

            # Amount (right-aligned, bold)
            amt = float(item.amount or 0)
            dr(ML + CW - 1.0*mm, mid_y, f'{amt:,.2f}', 'Helvetica-Bold', 9)

            IY = row_bot

        # Total row — single line at the bottom of the items box
        TOTAL_LINE_Y = ITEM_BOT + 6.0*mm
        hline(ML, TOTAL_LINE_Y, ML + CW, lw=0.75)
        dr(col_x[2] - 2.0*mm, ITEM_BOT + 2.0*mm, 'Total', 'Helvetica', 8)
        dr(ML + CW - 1.0*mm, ITEM_BOT + 1.5*mm,
           f'Rs. {float(total_amt):,.2f}', 'Helvetica-Bold', 11)

        # ═══════════════════════════════════════════════════════════════════════
        # 4. AMOUNT IN WORDS
        # ═══════════════════════════════════════════════════════════════════════
        box(ML, WORDS_BOT, CW, WORDS_H, lw=0.75)
        ds(ML + P, WORDS_TOP - 3.5*mm,
           'Amount Chargeable (in words)', 'Helvetica', 8)
        dr(ML + CW - P, WORDS_TOP - 3.5*mm, 'E. & O.E', 'Helvetica-Oblique', 8)
        ds(ML + P, WORDS_BOT + 4.5*mm, words_text, 'Helvetica-Bold', 9)

        # ═══════════════════════════════════════════════════════════════════════
        # 5. DECLARATION
        # ═══════════════════════════════════════════════════════════════════════
        box(ML, DECL_BOT, CW, DECL_H, lw=0.75)
        DECL_SPX = ML + CW * 0.62
        # Right aligned inner rectangle
        c.rect(DECL_SPX, DECL_BOT, CW - CW * 0.62, DECL_H/2.5, stroke=1, fill=0)

        # Left side: PAN + declaration text
        dy = DECL_TOP - 4.5*mm
        ds(ML + P, dy, f"Company's PAN", 'Helvetica', 8)
        ds(ML + 25*mm, dy, f":   {CPAN}", 'Helvetica-Bold', 9)
        dy -= 4.0*mm
        ds(ML + P, dy, 'Declaration', 'Helvetica', 8)
        dy -= 3.5*mm

        decl_max_w = CW * 0.62 - 3.0*mm
        decl_body  = (
            'Please Write the PO and Item Number in Delivery Challan as '
            'well as Invoice. Penalty will be apply for late delivery if '
            'your material recieved after due date to :'
        )
        for wl in wrap_line(decl_body, 'Helvetica', 8, decl_max_w):
            ds(ML + P, dy, wl, 'Helvetica', 8)
            dy -= 3.5*mm

        for penalty in [
            '1. One Week @ 10%.',
            '2. Two Week @ 25%.',
            '3. Three Week@ 50% deduct.',
            'Note:- Order poora hone par hi bhugtaan kiya jaavega !',
        ]:
            ds(ML + P, dy, penalty, 'Helvetica', 8)
            dy -= 3.5*mm

        # Right side: company name + Authorised Signatory
        dr(ML + CW - P, DECL_TOP - 4.0*mm,
           f'for {CNAME}', 'Helvetica-Bold', 9)
        dr(ML + CW - P, DECL_BOT + 2.5*mm,
           'Authorised Signatory', 'Helvetica', 9)

        # ═══════════════════════════════════════════════════════════════════════
        # 6. FOOTER
        # ═══════════════════════════════════════════════════════════════════════
        dc(PW / 2, FOOTER_Y,
           'This is a Computer Generated Document', 'Helvetica', 8)

        # Render page
        c.save()
        pdf_bytes = buf.getvalue()
        buf.close()

        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{po.po_number}.pdf"'
        return response

# ─── Sanding Workflow ViewSets ────────────────────────────────────────────────

# ─── Production & Stock Pipeline ViewSets ───────────────────────────────────

class ProductionJobViewSet(viewsets.ModelViewSet):
    """
    Manage stage production jobs (Sanding, Polishing, Packaging).
    - Supervisor assigns quantity from source stock -> Job created with status='assigned'.
    - Contractor marks work in progress or requests QC -> status='qc_requested'.
    - Supervisor performs QC -> passed_qty moves to destination Stock, rejected_qty stays for Rework.
    """
    serializer_class = ProductionJobSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        qs = ProductionJob.objects.select_related('stock_item', 'buyer_master', 'sample', 'buyer', 'contractor', 'assigned_by').prefetch_related('qc_logs', 'qc_logs__inspected_by').all()
        
        stage_param = self.request.query_params.get('stage')
        status_param = self.request.query_params.get('status')
        contractor_param = self.request.query_params.get('contractor')
        
        if stage_param:
            qs = qs.filter(stage=stage_param)
        if status_param:
            qs = qs.filter(status=status_param)
        if contractor_param:
            qs = qs.filter(contractor_id=contractor_param)
            
        if user.role == 'contractor':
            qs = qs.filter(contractor=user)
        elif user.role == 'supervisor':
            if user.production_unit:
                qs = qs.filter(Q(assigned_by=user) | Q(production_unit=user.production_unit) | Q(assigned_by__isnull=True))
            else:
                qs = qs.filter(Q(assigned_by=user) | Q(assigned_by__isnull=True))
            
        return qs

    def perform_create(self, serializer):
        user = self.request.user
        data = serializer.validated_data
        
        stock_item = data.get('stock_item')
        assigned_qty = Decimal(str(data.get('assigned_qty', 0)))
        
        if stock_item:
            if stock_item.quantity < assigned_qty:
                raise serializers.ValidationError({"assigned_qty": f"Insufficient stock. Available: {stock_item.quantity} {stock_item.unit}"})
            stock_item.quantity -= assigned_qty
            if stock_item.quantity <= 0:
                stock_item.status = 'Out of Stock'
            stock_item.save()
            
        serializer.save(assigned_by=user, status='assigned')

    @action(detail=True, methods=['post'], url_path='request-qc')
    def request_qc(self, request, pk=None):
        """Contractor submits job for Supervisor QC Inspection."""
        job = self.get_object()
        if request.user.role == 'contractor' and job.contractor != request.user:
            return Response({'detail': 'You can only request QC for your assigned jobs.'}, status=status.HTTP_403_FORBIDDEN)
            
        job.status = 'qc_requested'
        job.qc_requested_at = timezone.now()
        if 'contractor_notes' in request.data:
            job.contractor_notes = request.data['contractor_notes']
        job.save()
        return Response(ProductionJobSerializer(job).data)

    @action(detail=True, methods=['post'], url_path='perform-qc')
    def perform_qc(self, request, pk=None):
        """Supervisor inspects job: passes X pieces, rejects Y pieces."""
        job = self.get_object()
        user = request.user
        
        if user.role not in ('admin', 'supervisor'):
            return Response({'detail': 'Only supervisors or admins can perform QC inspection.'}, status=status.HTTP_403_FORBIDDEN)
            
        passed_qty = Decimal(str(request.data.get('passed_qty', 0)))
        rejected_qty = Decimal(str(request.data.get('rejected_qty', 0)))
        qc_notes = request.data.get('notes', '')

        if (passed_qty + rejected_qty) <= 0:
            return Response({'detail': 'Please enter valid passed or rejected quantities.'}, status=status.HTTP_400_BAD_REQUEST)

        new_passed_total = (job.passed_qty or Decimal(0)) + passed_qty
        if (new_passed_total + rejected_qty) > job.assigned_qty:
            return Response({
                'detail': f'Total passed ({new_passed_total}) + rejected ({rejected_qty}) cannot exceed assigned quantity ({job.assigned_qty}).'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Record QC log
        ProductionQCLog.objects.create(
            job=job,
            inspected_by=user,
            passed_qty=passed_qty,
            rejected_qty=rejected_qty,
            notes=qc_notes
        )

        job.passed_qty = new_passed_total
        job.qc_notes = qc_notes

        # If all assigned pieces have passed, complete job & clear rejected count
        if job.passed_qty >= job.assigned_qty:
            job.passed_qty = job.assigned_qty
            job.rejected_qty = Decimal('0')
            job.status = 'qc_completed'
            job.qc_completed_at = timezone.now()
        else:
            job.rejected_qty = rejected_qty
            if rejected_qty > 0:
                job.status = 'in_progress'
            else:
                job.status = 'qc_completed'
                job.qc_completed_at = timezone.now()

        job.save()

        # Add passed pieces to target stock stage
        if passed_qty > 0:
            target_stock_type = 'sanded'
            if job.stage == 'sanding':
                target_stock_type = 'sanded'
            elif job.stage == 'polishing':
                target_stock_type = 'polished'
            elif job.stage == 'packaging':
                target_stock_type = 'packaged'

            # Update or create target stock
            dest_stock, _ = StockItem.objects.get_or_create(
                stock_type=target_stock_type,
                style_no=job.style_no,
                defaults={
                    'item_name': job.item_name,
                    'quantity': Decimal('0'),
                    'unit': job.unit,
                    'buyer_master': job.buyer_master,
                    'sample': job.sample,
                    'buyer': job.buyer,
                    'status': 'In Stock'
                }
            )
            dest_stock.quantity += passed_qty
            dest_stock.status = 'In Stock'
            dest_stock.save()

        return Response(ProductionJobSerializer(job).data)


class ProductionQCLogViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ProductionQCLog.objects.select_related('job', 'inspected_by').all()
    serializer_class = ProductionQCLogSerializer
    permission_classes = [IsAuthenticated]


# ─── Number To Words Helper ───────────────────────────────────────────────────

def num2words(num, *args, **kwargs):
    if num is None:
        return ""
    try:
        val = float(num)
    except (ValueError, TypeError):
        return str(num)

    units = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten",
             "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

    def _convert_below_thousand(n):
        if n == 0:
            return ""
        elif n < 20:
            return units[n]
        elif n < 100:
            return tens[n // 10] + (" " + units[n % 10] if n % 10 != 0 else "")
        else:
            return units[n // 100] + " Hundred" + (" " + _convert_below_thousand(n % 100) if n % 100 != 0 else "")

    int_part = int(val)
    cents = int(round((val - int_part) * 100))

    if int_part == 0:
        words = "Zero"
    else:
        parts = []
        if int_part >= 1000000:
            millions = int_part // 1000000
            parts.append(_convert_below_thousand(millions) + " Million")
            int_part %= 1000000
        if int_part >= 1000:
            thousands = int_part // 1000
            parts.append(_convert_below_thousand(thousands) + " Thousand")
            int_part %= 1000
        if int_part > 0:
            parts.append(_convert_below_thousand(int_part))
        words = " ".join(parts)

    res = f"In Words : {words}"
    if cents > 0:
        res += f" and Cents {_convert_below_thousand(cents)}"
    res += " Only."
    return res


# ─── Performa Invoice ViewSet ──────────────────────────────────────────────────

class PerformaInvoiceViewSet(viewsets.ModelViewSet):
    queryset = PerformaInvoice.objects.select_related('buyer').prefetch_related('items').all()
    serializer_class = PerformaInvoiceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        buyer_id = self.request.query_params.get('buyer')
        if buyer_id:
            qs = qs.filter(buyer_id=buyer_id)
        return qs

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [IsAuthenticated(), IsAdmin()]
        return [IsAuthenticated()]

    @action(detail=True, methods=['get'], url_path='export-excel')
    def export_excel(self, request, pk=None):
        pi = self.get_object()
        buyer = pi.buyer

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"PI_{pi.pi_no}"
        ws.views.sheetView[0].showGridLines = True

        # Typography
        font_main = Font(name='Times New Roman', size=9)
        font_bold = Font(name='Times New Roman', size=9, bold=True)
        font_title = Font(name='Times New Roman', size=10, bold=True)
        font_sub = Font(name='Times New Roman', size=8)

        border_thin = Side(style='thin', color='000000')
        box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

        align_left_top = Alignment(horizontal='left', vertical='top', wrap_text=True)
        align_left_center = Alignment(horizontal='left', vertical='center', wrap_text=True)
        align_center_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_right_center = Alignment(horizontal='right', vertical='center')

        # 1. Exact Column Widths
        col_widths = {
            'A': 13, # Style No.
            'B': 18, # Image
            'C': 22, # Description 1
            'D': 22, # Description 2
            'E': 6,  # W / Spacer
            'F': 6,  # D
            'G': 6,  # H
            'H': 10, # Volume Per Pc
            'I': 6,  # Qty.
            'J': 10, # Total Volume
            'K': 11, # Rate US$
            'L': 15, # Amount US$
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # Helper to style a range of cells with values, font, alignment
        def style_range(start_r, start_c, end_r, end_c, value=None, font=font_main, alignment=align_left_center):
            if start_r != end_r or start_c != end_c:
                ws.merge_cells(start_row=start_r, start_column=start_c, end_row=end_r, end_column=end_c)
            top_left = ws.cell(row=start_r, column=start_c)
            if value is not None:
                top_left.value = value
            for r in range(start_r, end_r + 1):
                for c in range(start_c, end_c + 1):
                    cell = ws.cell(row=r, column=c)
                    if font: cell.font = font
                    if alignment: cell.alignment = alignment

        # Helper to apply full cell border to every cell in range
        def style_full_box(start_r, start_c, end_r, end_c, value=None, font=font_main, alignment=align_left_center):
            style_range(start_r, start_c, end_r, end_c, value=value, font=font, alignment=alignment)
            for r in range(start_r, end_r + 1):
                for c in range(start_c, end_c + 1):
                    ws.cell(row=r, column=c).border = box_border

        # Helper to apply outer box border only
        def apply_outer_border(start_r, start_c, end_r, end_c):
            for r in range(start_r, end_r + 1):
                for c in range(start_c, end_c + 1):
                    cell = ws.cell(row=r, column=c)
                    current_b = cell.border
                    cell.border = Border(
                        left=border_thin if c == start_c else current_b.left,
                        right=border_thin if c == end_c else current_b.right,
                        top=border_thin if r == start_r else current_b.top,
                        bottom=border_thin if r == end_r else current_b.bottom
                    )

        # Helper for explicit internal horizontal divider
        def add_horizontal_divider(row_idx, start_c, end_c):
            for c in range(start_c, end_c + 1):
                cell = ws.cell(row=row_idx, column=c)
                current_b = cell.border
                cell.border = Border(
                    left=current_b.left,
                    right=current_b.right,
                    top=current_b.top,
                    bottom=border_thin
                )

        # ─── 1. EXPORTER BOX (A1:D6) ──────────────────────────────────────────
        for r in range(1, 7):
            ws.row_dimensions[r].height = 18

        style_range(1, 1, 1, 4, "Exporter:", font=font_bold)
        style_range(2, 1, 2, 4, "Pinkcity Enterprises", font=font_title)
        style_range(3, 1, 3, 4, "G-78, EPIP, Sitapura Industrial Area, Tonk Road, Jaipur-302022 Rajasthan, India.")
        style_range(4, 1, 4, 4, "TEL: +91-141-2771144 / 2770033 | GSTIN: 08ABXPS4077R1Z8")
        style_range(5, 1, 5, 4, "IEC CODE: 1397002620 | State: Rajasthan, Code: 08")
        style_range(6, 1, 6, 4, "")
        apply_outer_border(1, 1, 6, 4)

        # ─── 2. INVOICE & BUYER REF BOX (Cols F to L, Rows 1 to 12) ───────────
        pi_date_str = pi.pi_date.strftime('%d %b, %Y') if pi.pi_date else ''
        order_date_str = pi.buyer_order_date.strftime('%d %b, %Y') if pi.buyer_order_date else ''

        style_range(1, 6, 1, 10, "Invoice No. & Date", font=font_bold)
        style_range(2, 6, 2, 10, f"{pi.pi_no} Dt. {pi_date_str}" if pi_date_str else (pi.pi_no or ''), font=font_bold)
        add_horizontal_divider(1, 6, 10)
        apply_outer_border(1, 6, 2, 10)

        style_range(1, 11, 1, 12, "Exporter's Ref.", font=font_bold)
        style_range(2, 11, 5, 12, pi.exporter_ref or '')
        add_horizontal_divider(1, 11, 12)
        apply_outer_border(1, 11, 5, 12)

        style_range(3, 6, 3, 10, "Buyer's Order No. & Date", font=font_bold)
        style_range(4, 6, 4, 10, f"{pi.buyer_order_no or ''} Dt. {order_date_str}" if order_date_str else (pi.buyer_order_no or ''), font=font_bold)
        add_horizontal_divider(3, 6, 10)
        apply_outer_border(3, 6, 4, 10)

        style_range(5, 6, 5, 10, "Other Reference(s)", font=font_bold)
        apply_outer_border(5, 6, 5, 10)

        style_range(6, 6, 6, 12, f"Buyer: {pi.buyer_name or buyer.name}", font=font_bold)
        apply_outer_border(6, 6, 6, 12)

        style_range(7, 6, 7, 12, "Buyer (if other than consignee)", font=font_bold)
        style_range(8, 6, 8, 12, f"Department # {pi.department_no or '69'}", font=font_bold)
        style_range(9, 6, 12, 12, "")
        add_horizontal_divider(7, 6, 12)
        add_horizontal_divider(8, 6, 12)
        apply_outer_border(7, 6, 12, 12)

        # ─── 3. CONSIGNEE BOX (A7:D12) ────────────────────────────────────────
        for r in range(7, 13):
            ws.row_dimensions[r].height = 18

        address_lines = (buyer.address or '').split('\n')
        line1 = address_lines[0] if len(address_lines) > 0 else ''
        line2 = address_lines[1] if len(address_lines) > 1 else ''

        style_range(7, 1, 7, 4, "Consignee:", font=font_bold)
        style_range(8, 1, 8, 4, buyer.name, font=font_title)
        style_range(9, 1, 9, 4, line1)
        style_range(10, 1, 10, 4, line2)
        style_range(11, 1, 11, 4, f"Tel: {buyer.phone}" if buyer.phone else "")
        style_range(12, 1, 12, 4, "VAT No. GB662563524, Reg No. 03094828")
        apply_outer_border(7, 1, 12, 4)

        # ─── 4. CARRIAGE & SHIPPING DETAILS (Rows 13 to 18) ───────────────────
        for r in range(13, 19):
            ws.row_dimensions[r].height = 18

        style_range(13, 1, 13, 1, "Pre Carriage by", font=font_bold)
        style_range(14, 1, 14, 1, pi.pre_carriage_by or 'Trailer')
        add_horizontal_divider(13, 1, 1)
        apply_outer_border(13, 1, 14, 1)

        style_range(13, 2, 13, 4, "Place of Receipt by Pre-carrier", font=font_bold)
        style_range(14, 2, 14, 4, pi.place_of_receipt or 'Jaipur')
        add_horizontal_divider(13, 2, 4)
        apply_outer_border(13, 2, 14, 4)

        style_range(13, 6, 13, 10, "Country of Origin of Goods", font=font_bold)
        style_range(14, 6, 14, 10, pi.country_of_origin or 'INDIA', font=font_bold)
        add_horizontal_divider(13, 6, 10)
        apply_outer_border(13, 6, 14, 10)

        style_range(13, 11, 13, 12, "Country of Final Destination", font=font_bold)
        style_range(14, 11, 14, 12, pi.country_final_destination or 'UK', font=font_bold)
        add_horizontal_divider(13, 11, 12)
        apply_outer_border(13, 11, 14, 12)

        style_range(15, 1, 15, 1, "Vessel/Flight No.", font=font_bold)
        style_range(16, 1, 16, 1, pi.vessel_flight_no or 'By Sea')
        add_horizontal_divider(15, 1, 1)
        apply_outer_border(15, 1, 16, 1)

        style_range(15, 2, 15, 4, "Port of Loading", font=font_bold)
        style_range(16, 2, 16, 4, pi.port_of_loading or 'Mundra')
        add_horizontal_divider(15, 2, 4)
        apply_outer_border(15, 2, 16, 4)

        style_range(15, 6, 15, 12, "Terms of Delivery and Payment", font=font_bold)
        style_range(16, 6, 16, 12, pi.terms_payment or 'Payment: T/T', font=font_bold)
        add_horizontal_divider(15, 6, 12)
        apply_outer_border(15, 6, 16, 12)

        style_range(17, 1, 17, 1, "Port of Discharge", font=font_bold)
        style_range(18, 1, 18, 1, pi.port_of_discharge or '')
        add_horizontal_divider(17, 1, 1)
        apply_outer_border(17, 1, 18, 1)

        style_range(17, 2, 17, 4, "Place of Delivery", font=font_bold)
        style_range(18, 2, 18, 4, pi.place_of_delivery or 'UNITED KINGDOM')
        add_horizontal_divider(17, 2, 4)
        apply_outer_border(17, 2, 18, 4)

        style_range(17, 6, 17, 12, pi.terms_delivery or '', font=font_bold)
        style_range(18, 6, 18, 12, "")
        add_horizontal_divider(17, 6, 12)
        apply_outer_border(17, 6, 18, 12)

        # ─── 5. TABLE HEADERS (Rows 19 & 20) ─────────────────────────────────
        ws.row_dimensions[19].height = 20
        ws.row_dimensions[20].height = 18

        style_full_box(19, 1, 20, 1, "Style No.", font=font_bold, alignment=align_center_center)
        style_full_box(19, 2, 20, 2, "Image", font=font_bold, alignment=align_center_center)
        style_full_box(19, 3, 20, 4, "Description of Goods", font=font_bold, alignment=align_center_center)

        style_full_box(19, 5, 19, 7, "Dimension (CM)", font=font_bold, alignment=align_center_center)
        style_full_box(20, 5, 20, 5, "W", font=font_bold, alignment=align_center_center)
        style_full_box(20, 6, 20, 6, "D", font=font_bold, alignment=align_center_center)
        style_full_box(20, 7, 20, 7, "H", font=font_bold, alignment=align_center_center)

        style_full_box(19, 8, 20, 8, "Volume\nPer Pc", font=font_bold, alignment=align_center_center)
        style_full_box(19, 9, 20, 9, "Qty.", font=font_bold, alignment=align_center_center)
        style_full_box(19, 10, 20, 10, "Total\nVolume", font=font_bold, alignment=align_center_center)
        style_full_box(19, 11, 20, 11, "Rate\nUS$", font=font_bold, alignment=align_center_center)
        style_full_box(19, 12, 20, 12, "Amount\nUS$", font=font_bold, alignment=align_center_center)

        # ─── 6. CATEGORY HEADER (Row 21) ──────────────────────────────────────
        ws.row_dimensions[21].height = 22
        cat_font = Font(name='Times New Roman', size=10, bold=True, underline='single')
        style_full_box(21, 1, 21, 12, pi.category_header or "Wooden Furniture Items", font=cat_font, alignment=align_center_center)

        # ─── 7. LINE ITEMS (Row 22 onwards) ──────────────────────────────────
        curr_row = 22
        tot_qty = 0
        tot_vol = 0.0
        tot_amt = 0.0
        temp_files = []

        items = list(pi.items.all())
        for item in items:
            ws.row_dimensions[curr_row].height = 75

            # Style No
            style_full_box(curr_row, 1, curr_row, 1, item.style_no, alignment=align_center_center)
            
            # Image Cell (Col B)
            style_full_box(curr_row, 2, curr_row, 2, "", alignment=align_center_center)

            # Description (C & D merged)
            style_full_box(curr_row, 3, curr_row, 4, item.description or '', alignment=align_center_center)

            # Dimensions
            style_full_box(curr_row, 5, curr_row, 5, float(item.dimension_w) if item.dimension_w else "", alignment=align_center_center)
            style_full_box(curr_row, 6, curr_row, 6, float(item.dimension_d) if item.dimension_d else "", alignment=align_center_center)
            style_full_box(curr_row, 7, curr_row, 7, float(item.dimension_h) if item.dimension_h else "", alignment=align_center_center)

            # Vol Per Pc
            vol_pc = float(item.volume_per_pc) if item.volume_per_pc else 0.0
            style_full_box(curr_row, 8, curr_row, 8, vol_pc, alignment=align_center_center)
            ws.cell(row=curr_row, column=8).number_format = '0.00'

            # Qty
            q = item.qty or 0
            tot_qty += q
            style_full_box(curr_row, 9, curr_row, 9, q, font=font_bold, alignment=align_center_center)

            # Total Volume
            v_tot = float(item.total_volume) if item.total_volume else (q * vol_pc)
            tot_vol += v_tot
            style_full_box(curr_row, 10, curr_row, 10, v_tot, alignment=align_right_center)
            ws.cell(row=curr_row, column=10).number_format = '0.00'

            # Rate US$
            r_usd = float(item.rate_usd) if item.rate_usd else 0.0
            style_full_box(curr_row, 11, curr_row, 11, r_usd, alignment=align_right_center)
            ws.cell(row=curr_row, column=11).number_format = '"$"#,##0.00'

            # Amount US$
            amt = float(item.amount_usd) if item.amount_usd else (q * r_usd)
            tot_amt += amt
            style_full_box(curr_row, 12, curr_row, 12, amt, alignment=align_right_center)
            ws.cell(row=curr_row, column=12).number_format = '"$"#,##0.00'

            # Embed Product Image in Cell B
            image_path = None
            if item.po and item.po.buyer_master and item.po.buyer_master.sample:
                sample_imgs = item.po.buyer_master.sample.images.all()
                if sample_imgs.exists():
                    image_path = sample_imgs.first().image.path

            if not image_path and item.image_url:
                rel_path = item.image_url.lstrip('/')
                abs_path = os.path.join(settings.MEDIA_ROOT, rel_path.replace('media/', ''))
                if os.path.exists(abs_path):
                    image_path = abs_path

            if image_path and os.path.exists(image_path):
                try:
                    pil_img = PILImage.open(image_path)
                    if pil_img.mode in ('RGBA', 'LA', 'P'):
                        pil_img = pil_img.convert('RGB')
                    pil_img.thumbnail((90, 60))

                    tmp_f = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False)
                    pil_img.save(tmp_f.name, format='JPEG', quality=85)
                    tmp_f.close()
                    temp_files.append(tmp_f.name)

                    xl_img = OpenpyxlImage(tmp_f.name)
                    add_centered_image(ws, f"B{curr_row}", xl_img)
                except Exception as e:
                    print(f"Error drawing image: {e}")

            curr_row += 1

        # ─── 8. AMOUNT CHARGEABLE SUMMARY ROW ────────────────────────────────
        ws.row_dimensions[curr_row].height = 22
        style_full_box(curr_row, 1, curr_row, 8, "Amount Chargeable", font=font_bold, alignment=align_left_center)
        style_full_box(curr_row, 9, curr_row, 9, tot_qty, font=font_bold, alignment=align_center_center)
        style_full_box(curr_row, 10, curr_row, 10, tot_vol, font=font_bold, alignment=align_right_center)
        ws.cell(row=curr_row, column=10).number_format = '0.00'

        style_full_box(curr_row, 11, curr_row, 11, "$", font=font_bold, alignment=align_right_center)

        style_full_box(curr_row, 12, curr_row, 12, tot_amt, font=font_bold, alignment=align_right_center)
        ws.cell(row=curr_row, column=12).number_format = '"$"#,##0.00'

        curr_row += 1

        # ─── 9. IN WORDS ROW ─────────────────────────────────────────────────
        ws.row_dimensions[curr_row].height = 20
        words_str = num2words(tot_amt)
        style_full_box(curr_row, 1, curr_row, 12, words_str, font=font_bold, alignment=align_left_center)

        curr_row += 1

        # ─── 10. DECLARATION SECTION ─────────────────────────────────────────
        ws.row_dimensions[curr_row].height = 18
        style_range(curr_row, 1, curr_row, 12, "Declaration:", font=font_bold, alignment=align_left_center)
        curr_row += 1

        ws.row_dimensions[curr_row].height = 55
        style_range(curr_row, 1, curr_row+2, 12, pi.declaration_text or '', font=font_sub, alignment=align_left_top)

        curr_row += 4

        # ─── 11. SIGNATURE & DATE ────────────────────────────────────────────
        ws.row_dimensions[curr_row].height = 30
        style_range(curr_row, 9, curr_row, 12, "Signature & Date", font=font_bold, alignment=align_right_center)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="PI_{pi.pi_no}.xlsx"'
        wb.save(response)

        for f in temp_files:
            try: os.remove(f)
            except: pass

        return response


# ─── Buyer Performa Invoice ViewSet (Pre-PO PI) ──────────────────────────────

class BuyerPIViewSet(viewsets.ModelViewSet):
    queryset = BuyerPI.objects.prefetch_related('items', 'items__buyer_master', 'items__buyer_master__sample').select_related('buyer').all()
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == 'list':
            return BuyerPIListSerializer
        return BuyerPISerializer

    def get_queryset(self):
        qs = super().get_queryset()
        buyer_id = self.request.query_params.get('buyer')
        if buyer_id:
            qs = qs.filter(buyer_id=buyer_id)
        return qs

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['request'] = self.request
        return context

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [IsAuthenticated(), IsAdmin()]
        return [IsAuthenticated()]

    @action(detail=True, methods=['get'], url_path='export-excel')
    def export_excel(self, request, pk=None):
        pi = self.get_object()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"PI_{pi.pi_no}"
        ws.views.sheetView[0].showGridLines = True

        thin_side = Side(style='thin', color='000000')
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        font_bold_lg = Font(name='Calibri', size=14, bold=True)
        font_bold_md = Font(name='Calibri', size=11, bold=True)
        font_bold_sm = Font(name='Calibri', size=10, bold=True)
        font_regular = Font(name='Calibri', size=9)
        fill_gray = PatternFill(start_color='EAEAEA', end_color='EAEAEA', fill_type='solid')

        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
        align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True, indent=1)

        def style_box(row_start, col_start, row_end, col_end, text="", font=font_regular, alignment=align_center, fill=None):
            if row_start != row_end or col_start != col_end:
                ws.merge_cells(start_row=row_start, start_column=col_start, end_row=row_end, end_column=col_end)
            
            top_left = ws.cell(row=row_start, column=col_start)
            top_left.value = text
            if not isinstance(text, CellRichText):
                top_left.font = font
            top_left.alignment = alignment
            
            for r in range(row_start, row_end + 1):
                for c in range(col_start, col_end + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = border_all
                    if fill:
                        cell.fill = fill

        ws.row_dimensions[1].height = 110

        # Block 1: Pinkcity Enterprises Header (A1:E1)
        font_company = InlineFont(rFont='Times New Roman', sz=20, b=True)
        font_block1_addr = InlineFont(rFont='Times New Roman', sz=10)
        company_text = CellRichText(
            TextBlock(font_company, "Pinkcity Enterprises\n"),
            TextBlock(font_block1_addr, "G 78 EPIP, Sitapura Industrial Area,\n"),
            TextBlock(font_block1_addr, "Jaipur 302022 -\n"),
            TextBlock(font_block1_addr, "Phone: +91 141 277 1144\n")  # Trailing newline to push title slightly higher
        )
        style_box(1, 1, 1, 5, company_text, alignment=align_center)

        # Block 2: Buyer Address (F1:J1)
        buyer_name = pi.buyer.name if pi.buyer else "BUYER"
        buyer_addr = pi.buyer.address if (pi.buyer and pi.buyer.address) else ""
        font_buyer_title = InlineFont(rFont='Calibri', sz=20, b=True)
        font_buyer_addr = InlineFont(rFont='Calibri', sz=10)
        font_spacer_calibri = InlineFont(rFont='Calibri', sz=5)
        
        buyer_blocks = [
            TextBlock(font_spacer_calibri, "\n"),  # Top padding
            TextBlock(font_buyer_title, f"{buyer_name.upper()}\n"),
            TextBlock(font_spacer_calibri, "\n")   # Spacing between title and address
        ]
        if buyer_addr:
            buyer_blocks.append(TextBlock(font_buyer_addr, buyer_addr))
        style_box(1, 6, 1, 10, CellRichText(*buyer_blocks), alignment=align_top_left, fill=fill_gray)

        # Block 3: Delivered To (K1:N1)
        delivered_contact = pi.delivered_to_name or ""
        delivered_comp = pi.delivered_to_company or (pi.buyer.name if pi.buyer else "")
        delivered_addr = pi.delivered_to_address or ""
        font_delim = InlineFont(rFont='Calibri', sz=9)
        font_comp = InlineFont(rFont='Calibri', sz=20, b=True)
        font_addr = InlineFont(rFont='Calibri', sz=10)
        
        deliv_blocks = [
            TextBlock(font_spacer_calibri, "\n")  # Top padding
        ]
        if delivered_contact:
            deliv_blocks.append(TextBlock(font_delim, f"DELIVERED TO: {delivered_contact}\n"))
        else:
            deliv_blocks.append(TextBlock(font_delim, "DELIVERED TO:\n"))
        deliv_blocks.append(TextBlock(font_comp, f"{delivered_comp}\n"))
        deliv_blocks.append(TextBlock(font_spacer_calibri, "\n"))  # Spacing
        if delivered_addr:
            deliv_blocks.append(TextBlock(font_addr, delivered_addr))
        style_box(1, 11, 1, 14, CellRichText(*deliv_blocks), alignment=align_top_left, fill=fill_gray)

        # Block 4: PI Summary Info (O1:Q1)
        formatted_date = pi.pi_date.strftime('%d/%m/%Y') if pi.pi_date else ''
        formatted_ex_fac = pi.ex_factory_date.strftime('%d %B, %Y') if pi.ex_factory_date else ''
        font_pi_bold = InlineFont(rFont='Times New Roman', sz=12, b=True)
        font_spacer_times = InlineFont(rFont='Times New Roman', sz=6)
        
        pi_summary_text = CellRichText(
            TextBlock(font_spacer_times, "\n"),  # Top padding
            TextBlock(font_pi_bold, f"PI of PO # {pi.pi_no}\n"),
            TextBlock(font_pi_bold, f"Date : {formatted_date}\n"),
            TextBlock(font_pi_bold, f"Ex-Factory : {formatted_ex_fac}\n"),
            TextBlock(font_pi_bold, f"Payment: {pi.payment_terms or '100% TT 30 Days from BL'}")
        )
        style_box(1, 15, 1, 17, pi_summary_text, alignment=align_top_left, fill=fill_gray)

        # ─── Table Headers (Row 2) ───────────────────────────────────────────
        ws.row_dimensions[2].height = 28
        headers_first_part = [
            ("S. No.", 1, 1),
            ("Barcode", 2, 2),
            ("Buyer #", 3, 3),
            ("Style No.", 4, 4),
            ("Picture", 5, 5),
            ("Name", 6, 6),
            ("Size CMs", 7, 9),
        ]
        headers_second_part = [
            ("Material", 10),
            ("Finish", 11),
            ("CBM", 12),
            ("Price USD", 13),
            ("Units", 14),
            ("Total CBM", 15),
            ("Total Amount", 16),
            ("Remarks", 17),
        ]

        for title, start_col, end_col in headers_first_part:
            style_box(2, start_col, 2, end_col, title, font=font_bold_sm, alignment=align_center)
        for title, col_idx in headers_second_part:
            style_box(2, col_idx, 2, col_idx, title, font=font_bold_sm, alignment=align_center)

        # Apply custom borders to the header blocks and headers row (Rows 1 & 2)
        blue_medium = Side(style='medium', color='1F4E78')
        thin_black = Side(style='thin', color='000000')

        for r in range(1, 3):
            for c in range(1, 18):
                cell = ws.cell(row=r, column=c)
                
                # Determine individual sides
                left_side = blue_medium if c == 1 else (thin_black if c in (6, 11, 15) else Side(style=None))
                right_side = blue_medium if c == 17 else (thin_black if c in (5, 10, 14) else Side(style=None))
                top_side = blue_medium if r == 1 else thin_black
                bottom_side = thin_black
                
                # For Row 2, every vertical divider should be thin black except the outer edges
                if r == 2:
                    left_side = blue_medium if c == 1 else thin_black
                    right_side = blue_medium if c == 17 else thin_black
                
                cell.border = Border(left=left_side, right=right_side, top=top_side, bottom=bottom_side)

        # ─── Line Items (Row 3 onwards) ──────────────────────────────────────
        curr_row = 3
        tot_units = 0
        tot_cbm = 0.0
        tot_amt = 0.0
        temp_files = []

        # Set column widths upfront so add_centered_image knows exact cell width
        col_widths = {
            1: 5, 2: 12, 3: 12, 4: 14, 5: 20,
            6: 30, 7: 7, 8: 7, 9: 7, 10: 12,
            11: 14, 12: 9, 13: 12, 14: 9, 15: 12, 16: 14, 17: 22
        }
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
            if col_idx == 9:
                ws.column_dimensions[col_letter].hidden = True

        items = list(pi.items.all())
        for idx, item in enumerate(items, 1):
            ws.row_dimensions[curr_row].height = 80

            # S. No. (Col 1)
            c = ws.cell(row=curr_row, column=1, value=idx)
            c.font = font_regular
            c.alignment = align_center
            c.border = border_all

            # Barcode (Col 2)
            c = ws.cell(row=curr_row, column=2, value=item.barcode or "")
            c.font = font_regular
            c.alignment = align_center
            c.border = border_all

            # Buyer # (Col 3)
            c = ws.cell(row=curr_row, column=3, value=item.buyer_no or "")
            c.font = font_regular
            c.alignment = align_center
            c.border = border_all

            # Style No (Col 4)
            c = ws.cell(row=curr_row, column=4, value=item.style_no or "")
            c.font = font_bold_sm
            c.alignment = align_center
            c.border = border_all

            # Picture (Col 5)
            c = ws.cell(row=curr_row, column=5, value="")
            c.border = border_all

            # Name (Col 6)
            c = ws.cell(row=curr_row, column=6, value=item.product_name or "")
            c.font = font_regular
            c.alignment = align_left
            c.border = border_all

            # Size CMs (L, W, H in Cols 7, 8, 9)
            l = float(item.size_length) if item.size_length else ""
            b_dim = float(item.size_breadth) if item.size_breadth else ""
            h = float(item.size_height) if item.size_height else ""
            
            c_l = ws.cell(row=curr_row, column=7, value=l)
            c_l.font = font_regular
            c_l.alignment = align_center
            c_l.border = border_all
            
            c_b = ws.cell(row=curr_row, column=8, value=b_dim)
            c_b.font = font_regular
            c_b.alignment = align_center
            c_b.border = border_all
            
            c_h = ws.cell(row=curr_row, column=9, value=h)
            c_h.font = font_regular
            c_h.alignment = align_center
            c_h.border = border_all

            # Material (Col 10)
            c = ws.cell(row=curr_row, column=10, value=item.material or "")
            c.font = font_regular
            c.alignment = align_center
            c.border = border_all

            # Finish (Col 11)
            c = ws.cell(row=curr_row, column=11, value=item.finish_color or "")
            c.font = font_regular
            c.alignment = align_center
            c.border = border_all

            # CBM (Col 12)
            cbm_val = float(item.cbm) if item.cbm else 0.0
            c = ws.cell(row=curr_row, column=12, value=cbm_val)
            c.font = font_regular
            c.alignment = align_right
            c.number_format = '0.0000'
            c.border = border_all

            # Price USD (Col 13)
            price_val = float(item.price_usd) if item.price_usd else 0.0
            c = ws.cell(row=curr_row, column=13, value=price_val)
            c.font = font_regular
            c.alignment = align_right
            c.number_format = '"$"#,##0.00'
            c.border = border_all

            # Units (Col 14)
            u_val = item.units or 0
            tot_units += u_val
            c = ws.cell(row=curr_row, column=14, value=u_val)
            c.font = font_bold_sm
            c.alignment = align_center
            c.border = border_all

            # Total CBM (Col 15)
            tcbm_val = float(item.total_cbm) if item.total_cbm else (u_val * cbm_val)
            tot_cbm += tcbm_val
            c = ws.cell(row=curr_row, column=15, value=tcbm_val)
            c.font = font_regular
            c.alignment = align_right
            c.number_format = '0.0000'
            c.border = border_all

            # Total Amount (Col 16)
            tamt_val = float(item.total_amount) if item.total_amount else (u_val * price_val)
            tot_amt += tamt_val
            c = ws.cell(row=curr_row, column=16, value=tamt_val)
            c.font = font_bold_sm
            c.alignment = align_right
            c.number_format = '"$"#,##0.00'
            c.border = border_all

            # Remarks (Col 17)
            c = ws.cell(row=curr_row, column=17, value=item.remarks or "")
            c.font = font_regular
            c.alignment = align_left
            c.border = border_all

            # Product Image embedding in Picture Column (Col 5)
            img_path = None
            if item.buyer_master and item.buyer_master.sample:
                imgs = item.buyer_master.sample.images.all()
                if imgs.exists() and imgs.first().image:
                    img_path = imgs.first().image.path

            if img_path and os.path.exists(img_path):
                try:
                    pil_img = PILImage.open(img_path)
                    if pil_img.mode in ('RGBA', 'LA', 'P'):
                        pil_img = pil_img.convert('RGB')
                    pil_img.thumbnail((75, 50))
                    tmp_f = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False)
                    pil_img.save(tmp_f.name, format='JPEG', quality=85)
                    tmp_f.close()
                    temp_files.append(tmp_f.name)

                    xl_img = OpenpyxlImage(tmp_f.name)
                    add_centered_image(ws, f"E{curr_row}", xl_img)
                except Exception as e:
                    print(f"Failed to embed image: {e}")

            curr_row += 1

        # Totals Row
        ws.row_dimensions[curr_row].height = 30
        style_box(curr_row, 1, curr_row, 13, "TOTAL", font=font_bold_md, alignment=align_right)
        
        c_u = ws.cell(row=curr_row, column=14, value=tot_units)
        c_u.font = font_bold_md
        c_u.alignment = align_center
        c_u.border = border_all

        c_cbm = ws.cell(row=curr_row, column=15, value=tot_cbm)
        c_cbm.font = font_bold_md
        c_cbm.alignment = align_right
        c_cbm.number_format = '0.0000'
        c_cbm.border = border_all

        c_amt = ws.cell(row=curr_row, column=16, value=tot_amt)
        c_amt.font = font_bold_md
        c_amt.alignment = align_right
        c_amt.number_format = '"$"#,##0.00'
        c_amt.border = border_all

        c_rem = ws.cell(row=curr_row, column=17, value="")
        c_rem.border = border_all

        # Adjust column widths
        col_widths = {
            1: 5, 2: 12, 3: 12, 4: 14, 5: 20,
            6: 30, 7: 7, 8: 7, 9: 7, 10: 12,
            11: 14, 12: 9, 13: 12, 14: 9, 15: 12, 16: 14, 17: 22
        }
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
            if col_idx == 9:
                ws.column_dimensions[col_letter].hidden = True

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="PI_{pi.pi_no}.xlsx"'
        wb.save(response)

        for f in temp_files:
            try:
                os.remove(f)
            except Exception:
                pass

        return response


class SupplierPOItemDefectViewSet(viewsets.ModelViewSet):
    queryset = SupplierPOItemDefect.objects.all()
    serializer_class = SupplierPOItemDefectSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        po_item_id = self.request.query_params.get('po_item')
        if po_item_id:
            qs = qs.filter(po_item_id=po_item_id)
        return qs

    def get_permissions(self):
        # We allow supervisors and admins to create/update
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [IsAuthenticated(), IsAdminOrSupervisor()]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        defect = serializer.save()
        
        # Save multiple images if uploaded
        images = self.request.FILES.getlist('images')
        for img in images:
            SupplierPOItemDefectImage.objects.create(defect=defect, image=img)
            
        # Notify admins
        admins = User.objects.filter(role='admin')
        for admin in admins:
            if admin != self.request.user:
                Notification.objects.create(
                    user=admin,
                    message=f"New defect reported by {self.request.user.username} on PO Item.",
                    link=f'/gate-entry/{defect.po_item.supplier_po.id}'
                )

    def perform_update(self, serializer):
        old_defect = self.get_object()
        new_defect = serializer.save()
        
        # If admin_reply changed (was empty and is now filled, or just changed by admin)
        if new_defect.admin_reply and new_defect.admin_reply != old_defect.admin_reply:
            if new_defect.reported_by and new_defect.reported_by != self.request.user:
                Notification.objects.create(
                    user=new_defect.reported_by,
                    message=f"Admin replied to your defect report on PO Item.",
                    link=f'/gate-entry/{new_defect.po_item.supplier_po.id}'
                )

class SupplierPOItemViewSet(viewsets.ModelViewSet):
    queryset = SupplierPOItem.objects.all()
    serializer_class = SupplierPOItemSerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy', 'receive_qc'):
            return [IsAuthenticated(), IsAdminOrSupervisor()]
        return [IsAuthenticated()]

    def perform_update(self, serializer):
        item = serializer.save()
        po = item.supplier_po
        if po:
            all_received = True
            for it in po.items.all():
                passed_tot = it.passed_quantity or Decimal(0)
                if passed_tot < it.quantity:
                    all_received = False
                    break
            if all_received:
                po.status = 'Received'
                po.save()

    @action(detail=True, methods=['post'], url_path='receive-qc')
    def receive_qc(self, request, pk=None):
        """Supervisor inspects gate entry items: passed pcs enter Raw Stock, rejected pcs log defect & generate GRN & Debit Notes."""
        po_item = self.get_object()
        po = po_item.supplier_po
        passed_qty = Decimal(str(request.data.get('passed_qty', 0)))
        rejected_qty = Decimal(str(request.data.get('rejected_qty', 0)))
        challan_no = request.data.get('challan_no', '') or request.data.get('supplier_invoice_no', '')
        
        if passed_qty < 0 or rejected_qty < 0:
            return Response({'detail': 'Quantities cannot be negative.'}, status=status.HTTP_400_BAD_REQUEST)
        if passed_qty == 0 and rejected_qty == 0:
            return Response({'detail': 'Please enter a valid passed or rejected quantity.'}, status=status.HTTP_400_BAD_REQUEST)
            
        po_item.passed_quantity = (po_item.passed_quantity or Decimal(0)) + passed_qty
        po_item.save()
        
        # Determine Delivery Round & GRN Number
        existing_grns = list(GateInwardReceipt.objects.filter(supplier_po=po).values_list('grn_number', flat=True).distinct())
        existing_grns = [g for g in existing_grns if g]
        round_number = len(existing_grns) + 1
        grn_number = f"GRN-{po.po_number}-R{round_number}"

        # Create Inward Receipt (GRN Record)
        receipt = GateInwardReceipt.objects.create(
            grn_number=grn_number,
            round_number=round_number,
            supplier_po=po,
            po_item=po_item,
            receipt_date=request.data.get('receipt_date') or timezone.now().date(),
            challan_no=challan_no,
            supplier_invoice_no=request.data.get('supplier_invoice_no', '') or challan_no,
            supplier_invoice_date=request.data.get('supplier_invoice_date') or timezone.now().date(),
            supplier_invoice_amount=Decimal(str(request.data.get('supplier_invoice_amount', 0))) if request.data.get('supplier_invoice_amount') else None,
            vehicle_no=request.data.get('vehicle_no', ''),
            driver_contact=request.data.get('driver_contact', ''),
            received_qty=passed_qty + rejected_qty,
            passed_qty=passed_qty,
            rejected_qty=rejected_qty,
            notes=request.data.get('remark', '') or request.data.get('notes', ''),
            inspected_by=request.user if request.user.is_authenticated else None
        )

        # Automatically add passed_qty to Raw Stock
        if passed_qty > 0:
            words = po_item.description.split()
            style = words[0] if words else "RAW-ITEM"
            item_name = po_item.description[:100] if po_item.description else "Raw Furniture Item"
            
            raw_stock, _ = StockItem.objects.get_or_create(
                stock_type='raw',
                po_item=po_item,
                defaults={
                    'style_no': style,
                    'item_name': item_name,
                    'quantity': Decimal(0),
                    'unit': po_item.unit or 'pcs',
                    'buyer': po_item.buyer,
                    'status': 'In Stock'
                }
            )
            raw_stock.quantity += passed_qty
            raw_stock.status = 'In Stock'
            raw_stock.save()
            
        # Log defect & Debit Note if any pieces rejected
        generated_debit_notes = []
        if rejected_qty > 0:
            remark = request.data.get('remark', 'Gate inspection rejected pieces')
            SupplierPOItemDefect.objects.create(
                po_item=po_item,
                reported_by=request.user,
                quantity=rejected_qty,
                remark=remark
            )

            unit_rate = po_item.rate or Decimal('0')
            total_rejection_val = rejected_qty * unit_rate
            max_limit = Decimal('200000.00')

            if total_rejection_val <= max_limit:
                batch_pcs = [rejected_qty]
            else:
                max_pcs_per_note = int(max_limit // unit_rate) if unit_rate > 0 else int(rejected_qty)
                full_notes_count = int(rejected_qty // max_pcs_per_note)
                rem_pcs = rejected_qty % max_pcs_per_note
                batch_pcs = [Decimal(str(max_pcs_per_note))] * full_notes_count
                if rem_pcs > 0:
                    batch_pcs.append(rem_pcs)

            for idx, batch_q in enumerate(batch_pcs, 1):
                dn_suffix = f"-{idx}" if len(batch_pcs) > 1 else ""
                dn_number = f"DN-{po.po_number}-{receipt.id.hex[:4].upper()}{dn_suffix}"

                subtotal = batch_q * unit_rate
                cartage_gst = round(subtotal * Decimal('0.18') * Decimal('0.03'), 2)
                cgst = round(subtotal * Decimal('0.09'), 2)
                sgst = round(subtotal * Decimal('0.09'), 2)
                
                raw_total = subtotal + cartage_gst + cgst + sgst
                final_amount = Decimal(str(round(raw_total)))
                round_off = final_amount - raw_total
                words_str = num2words(float(final_amount), lang='en_IN').title() + " Rupees Only"

                dn = SupplierDebitNote.objects.create(
                    vch_no=dn_number,
                    vch_date=request.data.get('receipt_date') or timezone.now().date(),
                    original_inv_no=receipt.supplier_invoice_no or challan_no or po.po_number,
                    original_inv_date=request.data.get('supplier_invoice_date') or timezone.now().date(),
                    supplier=po.supplier,
                    supplier_po=po,
                    po_item=po_item,
                    status='Issued',
                    item_description=f"{po_item.description} (GRN #{grn_number} - Round #{round_number})",
                    rejected_qty=batch_q,
                    unit=po_item.unit or 'pcs',
                    rate=unit_rate,
                    subtotal_amount=subtotal,
                    cartage_gst_rate=Decimal('18.0'),
                    cartage_gst_amount=cartage_gst,
                    cgst_rate=Decimal('9.0'),
                    cgst_amount=cgst,
                    sgst_rate=Decimal('9.0'),
                    sgst_amount=sgst,
                    round_off=round_off,
                    total_amount=final_amount,
                    amount_in_words=words_str,
                    remarks=f"BEING AMOUNT DEBITED FOR REJECTED GOODS IN GRN #{grn_number} (ROUND #{round_number})",
                    company_pan='ABXPS4077R'
                )
                generated_debit_notes.append(SupplierDebitNoteSerializer(dn).data)

        # Check if parent PO status should update
        if po:
            all_items = po.items.all()
            total_ordered = sum(it.quantity or Decimal('0') for it in all_items)
            total_passed = sum(it.passed_quantity or Decimal('0') for it in all_items)

            if total_passed >= total_ordered and total_ordered > 0:
                po.status = 'Received'
            elif total_passed > 0:
                po.status = 'Partial Received'
            po.save()
            
        return Response({
            'detail': f'Gate QC recorded for Round #{round_number} (GRN #{grn_number}).',
            'passed_quantity': float(po_item.passed_quantity),
            'po_status': po.status if po else 'Pending',
            'receipt': GateInwardReceiptSerializer(receipt).data,
            'debit_notes': generated_debit_notes
        })


class NotificationViewSet(viewsets.ModelViewSet):
    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user).order_by('-created_at')

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        return Response({'status': 'ok'})

    @action(detail=True, methods=['patch'])
    def mark_read(self, request, pk=None):
        notification = self.get_object()
        notification.is_read = True
        notification.save()
        return Response({'status': 'ok'})


class StockItemViewSet(viewsets.ModelViewSet):
    queryset = StockItem.objects.select_related('po_item', 'sample', 'buyer', 'buyer_master').all()
    serializer_class = StockItemSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        stock_type_param = self.request.query_params.get('stock_type')
        status_param = self.request.query_params.get('status')
        buyer_param = self.request.query_params.get('buyer')
        search_param = self.request.query_params.get('search')

        if stock_type_param:
            qs = qs.filter(stock_type=stock_type_param)
        if status_param:
            qs = qs.filter(status=status_param)
        if buyer_param:
            qs = qs.filter(buyer_id=buyer_param)
        if search_param:
            qs = qs.filter(
                Q(style_no__icontains=search_param) |
                Q(item_name__icontains=search_param) |
                Q(location__icontains=search_param) |
                Q(remarks__icontains=search_param)
            )

        ordering = self.request.query_params.get('ordering', '-created_at')
        if ordering:
            qs = qs.order_by(ordering)
        return qs

    def get_permissions(self):
        if self.action in ('create', 'update', 'partial_update', 'destroy'):
            return [IsAuthenticated(), IsAdminOrSupervisor()]
        return [IsAuthenticated()]

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory_Stock"
        ws.views.sheetView[0].showGridLines = True

        headers = ['Style No', 'Item Name', 'Quantity', 'Unit', 'Unit Price', 'Location', 'Status', 'Buyer', 'PO Ref', 'Created At']
        ws.append(headers)

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for item in queryset:
            ws.append([
                item.style_no,
                item.item_name,
                float(item.quantity) if item.quantity else 0,
                item.unit,
                float(item.unit_price) if item.unit_price else '',
                item.location or '',
                item.status,
                item.buyer.name if item.buyer else '',
                item.po_item.supplier_po.po_number if (item.po_item and item.po_item.supplier_po) else '',
                item.created_at.strftime('%Y-%m-%d %H:%M') if item.created_at else ''
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Inventory_Stock.xlsx"'
        wb.save(response)
        return response


class GeneratePresentationView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        presentation_type = request.data.get('presentation_type', 'buyer_sample')

        if presentation_type == 'vendor_inspection':
            cover_info_raw = request.data.get('cover_info', '{}')
            if isinstance(cover_info_raw, str):
                try:
                    cover_info = json.loads(cover_info_raw)
                except Exception:
                    cover_info = {}
            elif isinstance(cover_info_raw, dict):
                cover_info = cover_info_raw
            else:
                cover_info = {}

            slides_meta_raw = request.data.get('slides_meta', '[]')
            if isinstance(slides_meta_raw, str):
                try:
                    slides_meta = json.loads(slides_meta_raw)
                except Exception:
                    slides_meta = []
            elif isinstance(slides_meta_raw, list):
                slides_meta = slides_meta_raw
            else:
                slides_meta = []

            slides = []
            for idx, s_meta in enumerate(slides_meta):
                s_title = s_meta.get('title', f'Section #{idx+1}')
                s_images = []

                image_keys = s_meta.get('image_keys', [])
                for key in image_keys:
                    if key in request.FILES:
                        s_images.append(request.FILES[key])

                slides.append({
                    'title': s_title,
                    'images': s_images
                })

            if 'dqa_report_image' in request.FILES:
                cover_info['dqa_report_image'] = request.FILES['dqa_report_image']


            pptx_bytes = generate_vendor_inspection_pptx(
                cover_info=cover_info,
                slides_data=slides
            )


            po_str = str(cover_info.get('po_number', 'Report')).replace('/', '_').replace(' ', '_')
            response = HttpResponse(pptx_bytes, content_type='application/vnd.openxmlformats-officedocument.presentationml.presentation')
            response['Content-Disposition'] = f'attachment; filename="Vendor_Internal_Inspection_Report_{po_str}.pptx"'
            return response

        elif presentation_type == 'brand':

            buyer_name = request.data.get('buyer_name', '')
            buyer_id = request.data.get('buyer_id')
            if not buyer_name and buyer_id:
                try:
                    b = Buyer.objects.get(id=buyer_id)
                    buyer_name = b.name
                except Exception:
                    pass

            buyer_po_numbers = request.data.get('buyer_po_numbers', '')
            title = request.data.get('title', 'BRAND PRESENTATION')

            slides = []
            slides_meta_raw = request.data.get('slides_meta', '[]')
            if isinstance(slides_meta_raw, str):
                try:
                    slides_meta = json.loads(slides_meta_raw)
                except Exception:
                    slides_meta = []
            elif isinstance(slides_meta_raw, list):
                slides_meta = slides_meta_raw
            else:
                slides_meta = []

            for idx, s_meta in enumerate(slides_meta):
                s_title = s_meta.get('title', f'Product #{idx+1}')
                s_images = []

                # 1. Gather files from request.FILES
                image_keys = s_meta.get('image_keys', [])
                for key in image_keys:
                    if key in request.FILES:
                        s_images.append(request.FILES[key])

                # 2. Check ERP item if associated
                sample_id = s_meta.get('sample_id')
                buyer_master_id = s_meta.get('buyer_master_id')

                item_obj = None
                if sample_id:
                    item_obj = Sample.objects.filter(id=sample_id).first()
                elif buyer_master_id:
                    item_obj = BuyerMaster.objects.filter(id=buyer_master_id).first()

                if item_obj:
                    # Gather existing images
                    main_img_path = find_image_path(item_obj)
                    if main_img_path and main_img_path not in s_images:
                        s_images.insert(0, main_img_path)

                    target_sample = item_obj if isinstance(item_obj, Sample) else getattr(item_obj, 'sample', None)
                    if target_sample and hasattr(target_sample, 'images'):
                        for simg in target_sample.images.all():
                            if simg.image and hasattr(simg.image, 'path') and os.path.exists(simg.image.path):
                                if simg.image.path not in s_images:
                                    s_images.append(simg.image.path)

                    if hasattr(item_obj, 'finishing_images'):
                        for fimg in item_obj.finishing_images.all():
                            if fimg.image and hasattr(fimg.image, 'path') and os.path.exists(fimg.image.path):
                                if fimg.image.path not in s_images:
                                    s_images.append(fimg.image.path)

                if s_images:
                    slides.append({
                        'title': s_title,
                        'images': s_images
                    })

            if not slides:
                return Response({'error': 'Please add at least one product slide with images for Brand PPT.'}, status=status.HTTP_400_BAD_REQUEST)

            pptx_bytes = generate_brand_pptx_presentation(
                buyer_name=buyer_name,
                buyer_po_numbers=buyer_po_numbers,
                title=title,
                company_name="PINKCITY ENTERPRISES",
                slides_data=slides
            )

            safe_name = (buyer_name or 'Brand').replace(' ', '_')
            response = HttpResponse(pptx_bytes, content_type='application/vnd.openxmlformats-officedocument.presentationml.presentation')
            response['Content-Disposition'] = f'attachment; filename="Brand_Presentation_{safe_name}.pptx"'
            return response

        # Default / Buyer Sample Presentation
        buyer_id = request.data.get('buyer_id')
        sample_ids = request.data.get('sample_ids', [])
        buyer_master_ids = request.data.get('buyer_master_ids', [])
        
        items_per_slide = int(request.data.get('items_per_slide', 2))
        include_price = request.data.get('include_price', True) in (True, 'true', '1', 1)
        include_specs = request.data.get('include_specs', True) in (True, 'true', '1', 1)

        buyer = None
        if buyer_id:
            try:
                buyer = Buyer.objects.get(id=buyer_id)
            except Buyer.DoesNotExist:
                pass

        items = []
        if sample_ids:
            items = list(Sample.objects.filter(id__in=sample_ids))
        elif buyer_master_ids:
            items = list(BuyerMaster.objects.filter(id__in=buyer_master_ids))

        if not items:
            return Response({'error': 'Please select at least one sample or item for presentation.'}, status=status.HTTP_400_BAD_REQUEST)

        buyer_code_str = buyer.code if buyer else 'Catalog'

        pptx_bytes = generate_pptx_presentation(
            buyer=buyer,
            items=items,
            items_per_slide=items_per_slide,
            include_price=include_price,
            include_specs=include_specs
        )
        response = HttpResponse(pptx_bytes, content_type='application/vnd.openxmlformats-officedocument.presentationml.presentation')
        response['Content-Disposition'] = f'attachment; filename="Presentation_{buyer_code_str}.pptx"'
        return response


class ScanLookupView(APIView):
    """
    Scans a QR code payload or barcode string from an invoice/PO
    and returns matched PO / PI records with item details and validation flags.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        code_raw = str(request.data.get('code', '')).strip()
        if not code_raw:
            return Response({'error': 'No QR code or barcode payload provided.'}, status=status.HTTP_400_BAD_REQUEST)

        po_no = None
        pi_no = None
        scanned_meta = {}

        if code_raw.startswith('{') and code_raw.endswith('}'):
            try:
                parsed = json.loads(code_raw)
                po_no = parsed.get('po_number') or parsed.get('po_no')
                pi_no = parsed.get('pi_no') or parsed.get('invoice_no')
                scanned_meta = parsed
            except Exception:
                pass

        if not po_no and not pi_no:
            po_no = code_raw
            pi_no = code_raw

        matched_po = None
        matched_pi = None

        if po_no:
            matched_po = SupplierPO.objects.filter(Q(po_number__iexact=po_no) | Q(po_number__icontains=po_no)).first()

        if not matched_po and pi_no:
            matched_pi = PerformaInvoice.objects.filter(Q(pi_no__iexact=pi_no) | Q(pi_no__icontains=pi_no)).first()
            if matched_pi:
                matched_po = SupplierPO.objects.filter(items__buyer_pi__id=matched_pi.id).first()

        if not matched_po and not matched_pi:
            return Response({
                'found': False,
                'scanned_code': code_raw,
                'message': f"No matching Supplier PO or Invoice found in ERP database for code '{code_raw}'."
            }, status=status.HTTP_200_OK)

        items_data = []
        if matched_po:
            for item in matched_po.items.all():
                items_data.append({
                    'id': str(item.id),
                    'description': item.description,
                    'quantity': float(item.quantity or 0),
                    'passed_quantity': float(item.passed_quantity or 0),
                    'remaining_qty': float((item.quantity or 0) - (item.passed_quantity or 0)),
                    'unit': item.unit,
                    'rate': float(item.rate or 0),
                    'amount': float(item.amount or 0),
                    'buyer_pi_no': item.buyer_pi.pi_no if item.buyer_pi else None,
                })

        return Response({
            'found': True,
            'match_type': 'SupplierPO' if matched_po else 'PerformaInvoice',
            'scanned_code': code_raw,
            'po_details': {
                'id': str(matched_po.id) if matched_po else None,
                'po_number': matched_po.po_number if matched_po else (matched_pi.buyer_order_no if matched_pi else 'INV-MATCH'),
                'po_date': str(matched_po.po_date) if matched_po and matched_po.po_date else (str(matched_pi.pi_date) if matched_pi and matched_pi.pi_date else None),
                'supplier_name': matched_po.supplier.name if matched_po and matched_po.supplier else (matched_pi.buyer.name if matched_pi else 'Pinkcity Supplier'),
                'status': matched_po.status if matched_po else 'Valid Invoice',
                'supervisor': (matched_po.supervisor.get_full_name() or matched_po.supervisor.username) if (matched_po and matched_po.supervisor) else None,
                'items_count': len(items_data),
                'total_amount': float(matched_po.total_amount) if matched_po else 0.0,
            },
            'items': items_data,
            'scanned_meta': scanned_meta
        })


class GateInwardReceiptViewSet(viewsets.ModelViewSet):
    """
    Handles partial gate receipts for Supplier POs.
    Automatically updates PO item passed/rejected quantities and generates
    Tally Debit Notes (auto-split if rejection exceeds Rs 2 Lakhs E-Way bill threshold).
    """
    queryset = GateInwardReceipt.objects.select_related('supplier_po', 'po_item', 'supplier_po__supplier').all()
    serializer_class = GateInwardReceiptSerializer
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        data = request.data
        po_item_id = data.get('po_item')
        passed_qty = Decimal(str(data.get('passed_qty', 0)))
        rejected_qty = Decimal(str(data.get('rejected_qty', 0)))
        challan_no = data.get('challan_no', '')

        po_item = SupplierPOItem.objects.select_related('supplier_po', 'supplier_po__supplier').filter(id=po_item_id).first()
        if not po_item:
            return Response({'error': 'Supplier PO Item not found'}, status=status.HTTP_400_BAD_REQUEST)

        # Determine Delivery Round & GRN Number
        po = po_item.supplier_po
        existing_grns = list(GateInwardReceipt.objects.filter(supplier_po=po).values_list('grn_number', flat=True).distinct())
        existing_grns = [g for g in existing_grns if g]

        requested_grn = data.get('grn_number')
        requested_round = data.get('round_number')

        if requested_grn:
            grn_number = requested_grn
            round_number = int(requested_round or 1)
        else:
            round_number = len(existing_grns) + 1
            grn_number = f"GRN-{po.po_number}-R{round_number}"

        # Create Inward Receipt
        receipt = GateInwardReceipt.objects.create(
            grn_number=grn_number,
            round_number=round_number,
            supplier_po=po,
            po_item=po_item,
            receipt_date=data.get('receipt_date') or timezone.now().date(),
            challan_no=challan_no,
            supplier_invoice_no=data.get('supplier_invoice_no', '') or challan_no,
            supplier_invoice_date=data.get('supplier_invoice_date') or data.get('receipt_date') or timezone.now().date(),
            supplier_invoice_amount=Decimal(str(data.get('supplier_invoice_amount', 0))) if data.get('supplier_invoice_amount') else None,
            vehicle_no=data.get('vehicle_no', ''),
            driver_contact=data.get('driver_contact', ''),
            received_qty=passed_qty + rejected_qty,
            passed_qty=passed_qty,
            rejected_qty=rejected_qty,
            notes=data.get('notes', ''),
            inspected_by=request.user if request.user.is_authenticated else None
        )

        # Update PO Item passed quantity
        po_item.passed_quantity = (po_item.passed_quantity or Decimal('0')) + passed_qty
        po_item.save()

        # Update Overall Supplier PO Status
        all_items = po.items.all()
        total_ordered = sum(it.quantity or Decimal('0') for it in all_items)
        total_passed = sum(it.passed_quantity or Decimal('0') for it in all_items)

        if total_passed >= total_ordered and total_ordered > 0:
            po.status = 'Received'
        elif total_passed > 0:
            po.status = 'Partial Received'
        po.save()

        # Add passed quantity to Raw Stock
        if passed_qty > 0:
            StockItem.objects.create(
                stock_type='raw',
                po_item=po_item,
                buyer=po_item.buyer,
                style_no=po_item.buyer.code if po_item.buyer else 'RAW-PO',
                item_name=po_item.description[:250],
                quantity=passed_qty,
                unit=po_item.unit or 'pcs',
                unit_price=po_item.rate,
                location='Main Gate Raw Store',
                status='In Stock',
                remarks=f"GRN #{grn_number} (Round #{round_number}) via Inv #{receipt.supplier_invoice_no or challan_no or 'N/A'} for PO #{po.po_number}"
            )

        # Handle Rejections & Automated Debit Notes (E-Way Bill Limit <= Rs 2,00,000)
        generated_debit_notes = []
        if rejected_qty > 0:
            unit_rate = po_item.rate or Decimal('0')
            total_rejection_val = rejected_qty * unit_rate
            max_limit = Decimal('200000.00')

            # Determine split batches under Rs 2 Lakhs
            if total_rejection_val <= max_limit:
                batch_pcs = [rejected_qty]
            else:
                max_pcs_per_note = int(max_limit // unit_rate) if unit_rate > 0 else int(rejected_qty)
                if max_pcs_per_note < 1:
                    max_pcs_per_note = 1
                
                batch_pcs = []
                rem = int(rejected_qty)
                while rem > 0:
                    take = min(rem, max_pcs_per_note)
                    batch_pcs.append(Decimal(str(take)))
                    rem -= take

            # Create Debit Notes
            for idx, batch_q in enumerate(batch_pcs):
                suffix = string.ascii_uppercase[idx] if len(batch_pcs) > 1 else ""
                vch_num = f"DN/{timezone.now().strftime('%y-%m')}/{random.randint(100,999)}{suffix}"

                subtotal = batch_q * unit_rate
                cartage_gst = round(subtotal * Decimal('0.18'), 2)
                cgst = round(subtotal * Decimal('0.09'), 2)
                sgst = round(subtotal * Decimal('0.09'), 2)
                tot = subtotal + cartage_gst + cgst + sgst
                round_off = round(tot - int(tot), 2)
                final_amount = round(tot, 2)

                words_str = f"INR {num2words(float(final_amount)).title()} Only" if 'num2words' in globals() else f"INR {final_amount} Only"

                dn = SupplierDebitNote.objects.create(
                    vch_type='Debit Note',
                    vch_no=vch_num,
                    vch_date=timezone.now().date(),
                    original_inv_no=challan_no or po.po_number,
                    original_inv_date=po.po_date,
                    supplier=po.supplier,
                    supplier_po=po,
                    po_item=po_item,
                    hsn_sac='70099200',
                    item_description=f"{po_item.description[:200]} — Rejected Returns",
                    rejected_qty=batch_q,
                    unit=po_item.unit or 'No.',
                    rate=unit_rate,
                    subtotal_amount=subtotal,
                    cartage_gst_rate=Decimal('18.0'),
                    cartage_gst_amount=cartage_gst,
                    cgst_rate=Decimal('9.0'),
                    cgst_amount=cgst,
                    sgst_rate=Decimal('9.0'),
                    sgst_amount=sgst,
                    round_off=round_off,
                    total_amount=final_amount,
                    amount_in_words=words_str,
                    remarks=f"BEING AMOUNT DEBITED GOODS RETURN FURNITURE ITEM AS PER BILL NO. {challan_no or po.po_number}",
                    company_pan='ABXPS4077R'
                )
                generated_debit_notes.append(SupplierDebitNoteSerializer(dn).data)

        return Response({
            'receipt': GateInwardReceiptSerializer(receipt).data,
            'debit_notes': generated_debit_notes,
            'message': f"Gate receipt saved successfully. Created {len(generated_debit_notes)} Debit Note(s)."
        }, status=status.HTTP_201_CREATED)


class SupplierTaxInvoiceViewSet(viewsets.ModelViewSet):
    queryset = SupplierTaxInvoice.objects.select_related('supplier').prefetch_related('items__supplier_po').all()
    serializer_class = SupplierTaxInvoiceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        supplier_id = self.request.query_params.get('supplier')
        if supplier_id:
            qs = qs.filter(supplier_id=supplier_id)
        return qs


def generate_debit_note_pdf(debit_note):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36
    )

    styles = getSampleStyleSheet()
    normal_style = styles['Normal']

    title_style = ParagraphStyle(
        'DNTitle',
        parent=normal_style,
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        textColor=colors.HexColor('#8b5a2b'),
        alignment=1
    )

    subtitle_style = ParagraphStyle(
        'DNSubtitle',
        parent=normal_style,
        fontName='Helvetica-Bold',
        fontSize=10,
        leading=13,
        textColor=colors.HexColor('#334155'),
        alignment=1
    )

    header_style = ParagraphStyle(
        'DNHeader',
        parent=normal_style,
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#1e293b')
    )

    body_style = ParagraphStyle(
        'DNBody',
        parent=normal_style,
        fontName='Helvetica',
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#334155')
    )

    elements = []

    # Title Banner
    elements.append(Paragraph("PINKCITY ENTERPRISES (P) LTD", title_style))
    elements.append(Paragraph("DEBIT NOTE VOUCHER (GOODS RETURN / REJECTION)", subtitle_style))
    elements.append(Paragraph("G-21 Sitapura Industrial Area, Jaipur, Rajasthan 302022 | GSTIN: 08AADCP4381R2ZO", body_style))
    elements.append(Spacer(1, 8))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#8b5a2b'), spaceAfter=12))

    # Details Grid Table
    details_data = [
        [
            Paragraph(f"<b>Debit Note No:</b> {debit_note.vch_no}", body_style),
            Paragraph(f"<b>Dated:</b> {debit_note.vch_date}", body_style)
        ],
        [
            Paragraph(f"<b>Supplier:</b> {debit_note.supplier.name}", body_style),
            Paragraph(f"<b>Supplier GSTIN:</b> {debit_note.supplier.gstin or 'N/A'}", body_style)
        ],
        [
            Paragraph(f"<b>Address:</b> {debit_note.supplier.address or 'N/A'}", body_style),
            Paragraph(f"<b>Original Inv Ref:</b> {debit_note.original_inv_no or 'N/A'}", body_style)
        ],
        [
            Paragraph(f"<b>Status:</b> {debit_note.status}", body_style),
            Paragraph(f"<b>PO Ref:</b> {debit_note.supplier_po.po_number if debit_note.supplier_po else 'N/A'}", body_style)
        ]
    ]

    details_table = Table(details_data, colWidths=[260, 260])
    details_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#fcfaf6')),
        ('PADDING', (0,0), (-1,-1), 5),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#e2e8f0')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(details_table)
    elements.append(Spacer(1, 12))

    # Line Items Table
    items_data = [
        [
            Paragraph("<b>#</b>", header_style),
            Paragraph("<b>Description of Goods</b>", header_style),
            Paragraph("<b>HSN/SAC</b>", header_style),
            Paragraph("<b>Rejected Qty</b>", header_style),
            Paragraph("<b>Rate (₹)</b>", header_style),
            Paragraph("<b>Amount (₹)</b>", header_style),
        ]
    ]

    note_items = debit_note.items.all()
    if note_items.exists():
        for idx, item in enumerate(note_items, 1):
            items_data.append([
                Paragraph(str(idx), body_style),
                Paragraph(item.description, body_style),
                Paragraph(item.hsn_sac, body_style),
                Paragraph(f"{item.rejected_qty} {item.unit}", body_style),
                Paragraph(f"₹{float(item.rate):,.2f}", body_style),
                Paragraph(f"₹{float(item.amount):,.2f}", body_style),
            ])
    else:
        items_data.append([
            Paragraph("1", body_style),
            Paragraph(debit_note.item_description or "Rejected Furniture Components", body_style),
            Paragraph(debit_note.hsn_sac or "9403", body_style),
            Paragraph(f"{debit_note.rejected_qty} {debit_note.unit}", body_style),
            Paragraph(f"₹{float(debit_note.rate):,.2f}", body_style),
            Paragraph(f"₹{float(debit_note.total_amount):,.2f}", body_style),
        ])

    table = Table(items_data, colWidths=[25, 225, 70, 75, 60, 65])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('PADDING', (0,0), (-1,-1), 5),
        ('ALIGN', (3,0), (-1,-1), 'RIGHT'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 12))

    # Summary box
    tot = float(debit_note.total_amount or 0)
    summary_data = [
        [Paragraph("<b>Subtotal Amount:</b>", body_style), Paragraph(f"₹{tot:,.2f}", body_style)],
        [Paragraph("<b>Total Debit Note Amount:</b>", header_style), Paragraph(f"₹{tot:,.2f} INR", header_style)],
    ]
    summary_table = Table(summary_data, colWidths=[360, 160])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 25))

    # Signature Block
    sig_data = [
        [Paragraph("<b>Prepared By</b>", body_style), Paragraph("<b>Authorized Signatory</b>", body_style)]
    ]
    sig_table = Table(sig_data, colWidths=[260, 260])
    sig_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (0,0), 'LEFT'),
        ('ALIGN', (1,0), (1,0), 'RIGHT'),
    ]))
    elements.append(sig_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


class SupplierDebitNoteViewSet(viewsets.ModelViewSet):
    queryset = SupplierDebitNote.objects.select_related('supplier', 'supplier_po', 'po_item').all()
    serializer_class = SupplierDebitNoteSerializer
    permission_classes = [IsAuthenticated]

    @action(detail=True, methods=['get'], url_path='pdf')
    def download_pdf(self, request, pk=None):
        dn = self.get_object()
        pdf_bytes = generate_debit_note_pdf(dn)
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="DebitNote_{dn.vch_no}.pdf"'
        return response

    @action(detail=True, methods=['post'], url_path='resolve-repaired')
    def resolve_repaired(self, request, pk=None):
        dn = self.get_object()
        dn.status = 'Resolved (Repaired)'
        dn.save()

        # Increment passed qty on matching PO item if linked
        if dn.po_item:
            dn.po_item.passed_quantity = (dn.po_item.passed_quantity or Decimal('0')) + dn.rejected_qty
            dn.po_item.save()

        return Response({'status': 'Debit note resolved as Repaired & Accepted without financial penalty.', 'debit_note': SupplierDebitNoteSerializer(dn).data})

    @action(detail=True, methods=['post'], url_path='confirm-issue')
    def confirm_issue(self, request, pk=None):
        dn = self.get_object()
        dn.status = 'Issued'
        dn.save()
        return Response({'status': 'Debit Note issued successfully.', 'debit_note': SupplierDebitNoteSerializer(dn).data})


class StockOriginBreakdownView(APIView):
    """
    Returns breakdown of stock grouped by origin PO number and supplier
    for any of the 4 stock stages (raw, sanded, polished, packaged),
    including handling supervisor, clearance date, buyer details, and item breakdowns.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        stock_type = request.query_params.get('stock_type', 'raw').lower()
        unit_id = request.query_params.get('unit_id')

        items = StockItem.objects.filter(stock_type=stock_type).select_related(
            'po_item', 'po_item__supplier_po', 'po_item__supplier_po__supplier', 'buyer'
        )

        if unit_id and unit_id != 'all':
            items = items.filter(production_unit_id=unit_id)

        po_breakdown = {}
        for item in items:
            po = item.po_item.supplier_po if (item.po_item and item.po_item.supplier_po) else None
            po_key = po.po_number if po else "Unassigned Batch"
            supplier_name = po.supplier.name if (po and po.supplier) else "Pinkcity Internal"
            po_date = str(po.po_date) if (po and po.po_date) else str(item.created_at.date())
            
            # Supervisor info
            supervisor_info = (po.supervisor.get_full_name() or po.supervisor.username) if (po and po.supervisor) else "General Supervisor"
            
            # Clearance date (receipt date from GateInwardReceipt or created_at)
            clearance_date = item.created_at.strftime('%Y-%m-%d')
            if po and hasattr(po, 'gate_receipts') and po.gate_receipts.exists():
                latest_receipt = po.gate_receipts.order_by('-receipt_date').first()
                if latest_receipt and latest_receipt.receipt_date:
                    clearance_date = str(latest_receipt.receipt_date)

            if po_key not in po_breakdown:
                po_breakdown[po_key] = {
                    'po_number': po_key,
                    'supplier_name': supplier_name,
                    'po_date': po_date,
                    'clearance_date': clearance_date,
                    'supervisor': supervisor_info,
                    'total_qty': 0.0,
                    'unit': item.unit or 'pcs',
                    'unit_price': float(item.unit_price or 0),
                    'total_amount_inr': 0.0,
                    'items_list': []
                }

            item_qty = float(item.quantity or 0)
            item_price = float(item.unit_price or 0)
            po_breakdown[po_key]['total_qty'] += item_qty
            po_breakdown[po_key]['total_amount_inr'] += (item_qty * item_price)
            
            po_breakdown[po_key]['items_list'].append({
                'id': str(item.id),
                'style_no': item.style_no,
                'item_name': item.item_name,
                'quantity': item_qty,
                'unit_price': item_price,
                'location': item.location or 'Main Warehouse',
                'buyer_name': item.buyer.name if item.buyer else 'General Stock',
                'created_at': item.created_at.strftime('%Y-%m-%d %H:%M')
            })

        total_qty_sum = sum(b['total_qty'] for b in po_breakdown.values())

        return Response({
            'stock_type': stock_type,
            'total_stock_count': total_qty_sum,
            'total_po_count': len(po_breakdown),
            'po_breakdown': list(po_breakdown.values())
        }, status=status.HTTP_200_OK)


class DashboardStatsView(APIView):
    """
    GET /api/dashboard/stats/
    Returns real-time dynamic aggregate statistics, actual monthly revenue datasets,
    and manufacturing pipeline metrics from the database.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        sample_count = Sample.objects.count()
        buyer_count = Buyer.objects.filter(is_deleted=False).count()
        bm_count = BuyerMaster.objects.filter(buyer__is_deleted=False).count()
        po_count = SupplierPO.objects.count()
        pending_qc_count = SupplierPO.objects.filter(status='Pending').count()
        stock_count = StockItem.objects.count()

        performa_invoices = PerformaInvoice.objects.filter(buyer__is_deleted=False).prefetch_related('items')
        buyer_pis = BuyerPI.objects.filter(buyer__is_deleted=False).prefetch_related('items')

        pi_total_usd = 0.0
        monthly_map = {}

        for pi in performa_invoices:
            pi_sum = sum(float(item.amount_usd or 0) for item in pi.items.all())
            pi_total_usd += pi_sum
            dt = pi.pi_date or pi.created_at.date()
            year_str = str(dt.year)
            month_str = dt.strftime('%b')

            if year_str not in monthly_map:
                monthly_map[year_str] = {}
            if month_str not in monthly_map[year_str]:
                monthly_map[year_str][month_str] = {'revenue': 0.0, 'orders': 0}
            monthly_map[year_str][month_str]['revenue'] += pi_sum
            monthly_map[year_str][month_str]['orders'] += 1

        for bpi in buyer_pis:
            bpi_sum = sum(float(item.total_amount or 0) for item in bpi.items.all())
            pi_total_usd += bpi_sum
            dt = bpi.pi_date or bpi.created_at.date()
            year_str = str(dt.year)
            month_str = dt.strftime('%b')

            if year_str not in monthly_map:
                monthly_map[year_str] = {}
            if month_str not in monthly_map[year_str]:
                monthly_map[year_str][month_str] = {'revenue': 0.0, 'orders': 0}
            monthly_map[year_str][month_str]['revenue'] += bpi_sum
            monthly_map[year_str][month_str]['orders'] += 1

        months_list = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        current_year = str(timezone.now().year)

        year_data = []
        prev_rev = 0
        for m in months_list:
            m_info = monthly_map.get(current_year, {}).get(m, {'revenue': 0.0, 'orders': 0})
            rev = round(m_info['revenue'], 2)
            orders = m_info['orders']

            growth_str = ""
            if prev_rev > 0:
                pct = ((rev - prev_rev) / prev_rev) * 100
                growth_str = f"+{pct:.1f}%" if pct >= 0 else f"{pct:.1f}%"
            elif rev > 0:
                growth_str = "+100%"

            if rev > 0 or orders > 0:
                prev_rev = rev

            year_data.append({
                'month': m,
                'revenue': rev,
                'orders': orders,
                'growth': growth_str
            })

        # Pipeline Metrics
        tot_pos = max(1, po_count)
        gate_entry_pct = round(((tot_pos - pending_qc_count) / tot_pos) * 100) if po_count > 0 else 0

        total_jobs = ProductionJob.objects.count()
        sanding_jobs = ProductionJob.objects.filter(stage__iexact='sanding').count()
        polishing_jobs = ProductionJob.objects.filter(stage__iexact='polishing').count()

        sanding_pct = round((sanding_jobs / max(1, total_jobs)) * 100) if total_jobs > 0 else (round(gate_entry_pct * 0.85) if gate_entry_pct > 0 else 0)
        polishing_pct = round((polishing_jobs / max(1, total_jobs)) * 100) if total_jobs > 0 else (round(sanding_pct * 0.88) if sanding_pct > 0 else 0)
        packaging_pct = round((stock_count / max(1, sample_count)) * 100) if sample_count > 0 else 0

        qc_agg = ProductionQCLog.objects.aggregate(p=Sum('passed_qty'), r=Sum('rejected_qty'))
        passed_sum = float(qc_agg['p'] or 0)
        rejected_sum = float(qc_agg['r'] or 0)
        tot_inspected = passed_sum + rejected_sum

        if tot_inspected > 0 or ProductionQCLog.objects.exists():
            qc_pass_rate = round((passed_sum / tot_inspected) * 100, 1) if tot_inspected > 0 else 100.0
        else:
            po_agg = SupplierPOItem.objects.aggregate(tot=Sum('quantity'), passed=Sum('passed_quantity'))
            tot_qty = float(po_agg['tot'] or 0)
            pass_qty = float(po_agg['passed'] or 0)
            if tot_qty > 0:
                qc_pass_rate = round((pass_qty / tot_qty) * 100, 1)
            else:
                qc_pass_rate = 100.0 if po_count > 0 else 0.0

        recent_pos_qs = SupplierPO.objects.select_related(
            'supplier', 'supervisor', 'buyer_pi'
        ).prefetch_related(
            'items__buyer',
            'extension_logs',
            'supplier_history',
            'supplier_history__previous_supplier',
            'supplier_history__new_supplier',
            'supplier_history__changed_by'
        ).all()[:5]
        recent_pos = SupplierPOSerializer(recent_pos_qs, many=True).data

        recent_pis_qs = PerformaInvoice.objects.select_related('buyer').prefetch_related('items').all()[:5]
        recent_pis = PerformaInvoiceSerializer(recent_pis_qs, many=True).data

        total_pis_count = len(performa_invoices) + len(buyer_pis)

        return Response({
            'totalSamples': sample_count,
            'totalBuyers': buyer_count,
            'totalBuyerMasters': bm_count,
            'totalPOs': po_count,
            'totalPIs': total_pis_count,
            'totalStockItems': stock_count,
            'pendingQcCount': pending_qc_count,
            'totalRevenueUSD': round(pi_total_usd, 2),
            'recentPOs': recent_pos,
            'recentPIs': recent_pis,
            'revenueDatasets': {
                current_year: year_data,
                'last6': year_data[-6:],
                'ytd': year_data[:timezone.now().month],
            },
            'pipelineMetrics': {
                'gateEntry': min(100, gate_entry_pct),
                'sanding': min(100, sanding_pct),
                'polishing': min(100, polishing_pct),
                'packaging': min(100, packaging_pct),
                'passRate': min(100.0, qc_pass_rate),
            }
        }, status=status.HTTP_200_OK)


# ─── Production Unit & Dynamic Workload Allocation Views ────────────────────

class ProductionUnitViewSet(viewsets.ModelViewSet):
    queryset = ProductionUnit.objects.all()
    serializer_class = ProductionUnitSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        sup_sub = Subquery(
            User.objects.filter(production_unit=OuterRef('pk'), role='supervisor', is_active=True)
            .values('production_unit')
            .annotate(c=Count('id'))
            .values('c')[:1],
            output_field=IntegerField()
        )
        con_sub = Subquery(
            User.objects.filter(production_unit=OuterRef('pk'), role='contractor', is_active=True)
            .values('production_unit')
            .annotate(c=Count('id'))
            .values('c')[:1],
            output_field=IntegerField()
        )
        stk_sub = Subquery(
            StockItem.objects.filter(production_unit=OuterRef('pk'))
            .values('production_unit')
            .annotate(total=Sum('quantity'))
            .values('total')[:1],
            output_field=DecimalField(max_digits=12, decimal_places=2)
        )
        return super().get_queryset().annotate(
            annotated_supervisor_count=Coalesce(sup_sub, Value(0)),
            annotated_contractor_count=Coalesce(con_sub, Value(0)),
            annotated_stock_count=Coalesce(stk_sub, Value(Decimal('0.00'), output_field=DecimalField(max_digits=12, decimal_places=2)))
        )


class BuyerUnitAllocationViewSet(viewsets.ModelViewSet):
    queryset = BuyerUnitAllocation.objects.all()
    serializer_class = BuyerUnitAllocationSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        qs = super().get_queryset()
        buyer_id = self.request.query_params.get('buyer_id')
        unit_id = self.request.query_params.get('unit_id')
        if buyer_id:
            qs = qs.filter(buyer_id=buyer_id)
        if unit_id:
            qs = qs.filter(production_unit_id=unit_id)
        return qs


class UnitWorkReallocationViewSet(viewsets.ModelViewSet):
    queryset = UnitWorkReallocation.objects.all()
    serializer_class = UnitWorkReallocationSerializer
    permission_classes = [AllowAny]


class WorkloadReallocationView(APIView):
    """
    Dynamic Workload Re-allocation API.
    Allows Admins to shift active Buyer POs, Stock Items, and Contractor Jobs
    from one Production Unit to another on the fly.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        buyer_id = request.data.get('buyer_id')
        po_id = request.data.get('po_id')
        from_unit_id = request.data.get('from_unit_id')
        to_unit_id = request.data.get('to_unit_id')
        reason = request.data.get('reason', 'Dynamic Workload Re-allocation')

        if not to_unit_id:
            return Response({'error': 'Target production unit (to_unit_id) is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            to_unit = ProductionUnit.objects.get(id=to_unit_id)
        except ProductionUnit.DoesNotExist:
            return Response({'error': 'Target Production Unit does not exist.'}, status=status.HTTP_404_NOT_FOUND)

        from_unit = ProductionUnit.objects.filter(id=from_unit_id).first() if from_unit_id else None
        buyer = Buyer.objects.filter(id=buyer_id).first() if buyer_id else None
        po = SupplierPO.objects.filter(id=po_id).first() if po_id else None

        # 1. Update SupplierPOs
        po_qs = SupplierPO.objects.all()
        if po:
            po_qs = po_qs.filter(id=po.id)
        elif buyer:
            po_qs = po_qs.filter(items__buyer=buyer)
        elif from_unit:
            po_qs = po_qs.filter(production_unit=from_unit)

        pos_updated = po_qs.update(production_unit=to_unit)

        # 2. Update Stock Items
        stock_qs = StockItem.objects.all()
        if buyer:
            stock_qs = stock_qs.filter(buyer=buyer)
        if from_unit:
            stock_qs = stock_qs.filter(production_unit=from_unit)
        stock_updated = stock_qs.update(production_unit=to_unit)

        # 3. Update Production Jobs
        jobs_qs = ProductionJob.objects.all()
        if buyer:
            jobs_qs = jobs_qs.filter(buyer=buyer)
        if from_unit:
            jobs_qs = jobs_qs.filter(production_unit=from_unit)
        jobs_updated = jobs_qs.update(production_unit=to_unit)

        # 4. Log Audit Trail
        realloc_log = UnitWorkReallocation.objects.create(
            buyer=buyer,
            po=po,
            from_unit=from_unit,
            to_unit=to_unit,
            reallocated_by=request.user if request.user.is_authenticated else None,
            reason=f"{reason} (POs updated: {pos_updated}, Stock items updated: {stock_updated}, Jobs updated: {jobs_updated})"
        )

        return Response({
            'message': f'Work successfully re-allocated to {to_unit.name}',
            'pos_updated': pos_updated,
            'stock_items_updated': stock_updated,
            'jobs_updated': jobs_updated,
            'audit_log_id': str(realloc_log.id)
        }, status=status.HTTP_200_OK)


# ─── Samples & Finishes Custom Excel Export / Import Views ───────────────────

class SampleExcelExportView(APIView):
    """
    Exports selected (or filtered) samples to Excel with centered embedded images.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        sample_ids = request.data.get('sample_ids', [])
        qs = Sample.objects.select_related('buyer', 'finish').prefetch_related('images').all()

        if sample_ids:
            qs = qs.filter(id__in=sample_ids)
        else:
            q = request.data.get('q', '').strip()
            buyer_id = request.data.get('buyer_id')
            if q:
                qs = qs.filter(
                    Q(sample_id__icontains=q) |
                    Q(style_no__icontains=q) |
                    Q(product_name__icontains=q)
                )
            if buyer_id:
                qs = qs.filter(buyer_id=buyer_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Samples_Catalog"
        ws.views.sheetView[0].showGridLines = True

        headers = [
            "S.No.", "Picture", "Sample ID", "Style No.", "Product Name",
            "Buyer", "Material", "Finish / Color", "CBM", "Price (USD)",
            "Vendor Name", "Size (cm)", "Size (in)", "Remark"
        ]

        ws.row_dimensions[1].height = 28
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='0284c7', end_color='0284c7', fill_type='solid')
        border_thin = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)

        col_widths = {
            1: 8, 2: 18, 3: 16, 4: 16, 5: 26,
            6: 18, 7: 18, 8: 18, 9: 10, 10: 12,
            11: 18, 12: 16, 13: 16, 14: 24
        }
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        for col_idx, h_text in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_thin

        temp_files = []
        curr_row = 2

        for idx, sample in enumerate(qs, start=1):
            ws.row_dimensions[curr_row].height = 65

            size_cm = f"{sample.size_length or 0} × {sample.size_breadth or 0} × {sample.size_height or 0}"
            size_in = f"{sample.size_length_inch or 0} × {sample.size_breadth_inch or 0} × {sample.size_height_inch or 0}"

            row_data = [
                idx,
                "", # Image placeholder in col 2
                sample.sample_id,
                sample.style_no or "",
                sample.product_name,
                sample.buyer.name if sample.buyer else "",
                sample.material or "",
                sample.finish.name if sample.finish else (sample.finish_color or ""),
                float(sample.cbm) if sample.cbm else "",
                float(sample.usd) if sample.usd else "",
                sample.vendor_name or "",
                size_cm,
                size_in,
                sample.remark or ""
            ]

            for c_idx, val in enumerate(row_data, start=1):
                c = ws.cell(row=curr_row, column=c_idx, value=val)
                c.border = border_thin
                if c_idx in (1, 3, 4, 12, 13):
                    c.alignment = align_center
                elif c_idx in (9, 10):
                    c.alignment = align_right
                    if c_idx == 10 and val != "":
                        c.number_format = '"$"#,##0.00'
                else:
                    c.alignment = align_left

            # Image embedding in Col 2 (B)
            img_obj = sample.images.first() if sample.images.exists() else None
            img_file = img_obj.image if (img_obj and img_obj.image) else sample.image

            if img_file and hasattr(img_file, 'path') and os.path.exists(img_file.path):
                try:
                    pil_img = PILImage.open(img_file.path)
                    if pil_img.mode in ('RGBA', 'LA', 'P'):
                        pil_img = pil_img.convert('RGB')
                    pil_img.thumbnail((80, 60))
                    tmp_f = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False)
                    pil_img.save(tmp_f.name, format='JPEG', quality=85)
                    tmp_f.close()
                    temp_files.append(tmp_f.name)

                    xl_img = OpenpyxlImage(tmp_f.name)
                    add_centered_image(ws, f"B{curr_row}", xl_img)
                except Exception as e:
                    print(f"Failed to embed sample image: {e}")

            curr_row += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Samples_Catalog.xlsx"'
        wb.save(response)

        for f in temp_files:
            try:
                os.remove(f)
            except Exception:
                pass

        return response


class FinishExcelExportView(APIView):
    """
    Exports selected (or filtered) finishes to Excel with centered embedded images.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        finish_ids = request.data.get('finish_ids', [])
        qs = Finish.objects.all()

        if finish_ids:
            qs = qs.filter(id__in=finish_ids)
        else:
            q = request.data.get('q', '').strip()
            if q:
                qs = qs.filter(
                    Q(name__icontains=q) |
                    Q(finish_code__icontains=q) |
                    Q(color__icontains=q) |
                    Q(wood_type__icontains=q)
                )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Finishing_Catalog"
        ws.views.sheetView[0].showGridLines = True

        headers = ["S.No.", "Picture", "Finish Code", "Finish Name", "Color", "Wood Type", "Created Date"]

        ws.row_dimensions[1].height = 28
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='b45309', end_color='b45309', fill_type='solid')
        border_thin = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

        col_widths = {1: 8, 2: 18, 3: 16, 4: 24, 5: 18, 6: 20, 7: 16}
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        for col_idx, h_text in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_thin

        temp_files = []
        curr_row = 2

        for idx, finish in enumerate(qs, start=1):
            ws.row_dimensions[curr_row].height = 65

            row_data = [
                idx,
                "", # Image placeholder in col 2
                finish.finish_code or "",
                finish.name,
                finish.color or "",
                finish.wood_type or "",
                finish.created_at.strftime('%Y-%m-%d') if finish.created_at else ""
            ]

            for c_idx, val in enumerate(row_data, start=1):
                c = ws.cell(row=curr_row, column=c_idx, value=val)
                c.border = border_thin
                if c_idx in (1, 3, 7):
                    c.alignment = align_center
                else:
                    c.alignment = align_left

            # Image embedding in Col 2 (B)
            if finish.image and hasattr(finish.image, 'path') and os.path.exists(finish.image.path):
                try:
                    pil_img = PILImage.open(finish.image.path)
                    if pil_img.mode in ('RGBA', 'LA', 'P'):
                        pil_img = pil_img.convert('RGB')
                    pil_img.thumbnail((80, 60))
                    tmp_f = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False)
                    pil_img.save(tmp_f.name, format='JPEG', quality=85)
                    tmp_f.close()
                    temp_files.append(tmp_f.name)

                    xl_img = OpenpyxlImage(tmp_f.name)
                    add_centered_image(ws, f"B{curr_row}", xl_img)
                except Exception as e:
                    print(f"Failed to embed finish image: {e}")

            curr_row += 1

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Finishing_Catalog.xlsx"'
        wb.save(response)

        for f in temp_files:
            try:
                os.remove(f)
            except Exception:
                pass

        return response


class FinishExcelImportView(APIView):
    """
    Imports Finishes from uploaded .xlsx or .csv file into the database.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'error': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)

        file_name = file_obj.name.lower()
        imported_count = 0
        updated_count = 0
        images_extracted = 0

        try:
            if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                wb = openpyxl.load_workbook(file_obj, data_only=True)
                ws = wb.active

                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    return Response({'error': 'The uploaded Excel file is empty.'}, status=status.HTTP_400_BAD_REQUEST)

                header_row = [str(h or '').strip().lower() for h in rows[0]]

                def find_idx(candidates):
                    for idx, h in enumerate(header_row):
                        for c in candidates:
                            if c in h:
                                return idx
                    return -1

                code_col = find_idx(['finish code', 'code', 'finish_code'])
                name_col = find_idx(['finish name', 'name', 'finish_name', 'title'])
                color_col = find_idx(['color', 'finish color', 'shade'])
                wood_col = find_idx(['wood type', 'wood', 'material'])

                row_images = {}
                if hasattr(ws, '_images'):
                    for img in ws._images:
                        try:
                            img_row = img.anchor._from.row + 1
                            if img_row not in row_images:
                                row_images[img_row] = []
                            row_images[img_row].append(img)
                        except Exception:
                            pass

                for excel_row_num, r in enumerate(rows[1:], start=2):
                    if not r or not any(r):
                        continue
                    name_val = str(r[name_col] or '').strip() if (name_col != -1 and name_col < len(r)) else ''
                    code_val = str(r[code_col] or '').strip() if (code_col != -1 and code_col < len(r)) else ''
                    color_val = str(r[color_col] or '').strip() if (color_col != -1 and color_col < len(r)) else ''
                    wood_val = str(r[wood_col] or '').strip() if (wood_col != -1 and wood_col < len(r)) else ''

                    if not name_val and not code_val:
                        continue

                    if not name_val:
                        name_val = f"Finish {code_val}"

                    finish_obj = None
                    if code_val:
                        finish_obj = Finish.objects.filter(finish_code=code_val).first()
                    if not finish_obj and name_val:
                        finish_obj = Finish.objects.filter(name__iexact=name_val).first()

                    if finish_obj:
                        finish_obj.name = name_val
                        if code_val:
                            finish_obj.finish_code = code_val
                        if color_val:
                            finish_obj.color = color_val
                        if wood_val:
                            finish_obj.wood_type = wood_val
                        finish_obj.save()
                        updated_count += 1
                    else:
                        finish_obj = Finish.objects.create(
                            name=name_val,
                            finish_code=code_val or None,
                            color=color_val or None,
                            wood_type=wood_val or None
                        )
                        imported_count += 1

                    if excel_row_num in row_images and row_images[excel_row_num]:
                        try:
                            img_obj = row_images[excel_row_num][0]
                            image_bytes = img_obj._data()
                            ext = img_obj.format if hasattr(img_obj, 'format') and img_obj.format else 'png'
                            clean_name = (code_val or name_val).replace('/', '_').replace(' ', '_')
                            file_name_img = f"finish_{clean_name}.{ext}"
                            content_file = ContentFile(image_bytes, name=file_name_img)

                            finish_obj.image = content_file
                            finish_obj.save(update_fields=['image'])
                            images_extracted += 1
                        except Exception as img_err:
                            print(f"Error saving finish image for row {excel_row_num}: {img_err}")
            else:
                decoded_file = file_obj.read().decode('utf-8-sig')
                csv_reader = csv.DictReader(io.StringIO(decoded_file))

                for row in csv_reader:
                    name_val = row.get('Finish Name') or row.get('name') or row.get('Name') or ''
                    code_val = row.get('Finish Code') or row.get('code') or row.get('Code') or ''
                    color_val = row.get('Color') or row.get('color') or ''
                    wood_val = row.get('Wood Type') or row.get('wood_type') or ''

                    name_val = name_val.strip()
                    code_val = code_val.strip()

                    if not name_val and not code_val:
                        continue

                    if not name_val:
                        name_val = f"Finish {code_val}"

                    finish_obj = None
                    if code_val:
                        finish_obj = Finish.objects.filter(finish_code=code_val).first()
                    if not finish_obj and name_val:
                        finish_obj = Finish.objects.filter(name__iexact=name_val).first()

                    if finish_obj:
                        finish_obj.name = name_val
                        if code_val:
                            finish_obj.finish_code = code_val
                        if color_val:
                            finish_obj.color = color_val
                        if wood_val:
                            finish_obj.wood_type = wood_val
                        finish_obj.save()
                        updated_count += 1
                    else:
                        Finish.objects.create(
                            name=name_val,
                            finish_code=code_val or None,
                            color=color_val or None,
                            wood_type=wood_val or None
                        )
                        imported_count += 1

            msg = f"Successfully imported {imported_count} new finishes and updated {updated_count} finishes!"
            if images_extracted > 0:
                msg += f" {images_extracted} swatch image(s) extracted."

            return Response({
                'message': msg,
                'imported_count': imported_count,
                'updated_count': updated_count,
                'images_extracted': images_extracted
            }, status=status.HTTP_200_OK)


        except Exception as e:
            return Response({'error': f'Failed to process file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


# ── Bulk Delete Views ──────────────────────────────────────────────────────────

class SampleBulkDeleteView(APIView):
    """
    POST /api/samples/bulk-delete/
    Body: { "sample_ids": [1, 2, 3, ...] }
    Deletes all samples with the given IDs. Admin only.
    """
    permission_classes = [IsAuthenticated, IsAdmin]

    def post(self, request):
        sample_ids = request.data.get('sample_ids', [])
        if not sample_ids:
            return Response({'error': 'No sample IDs provided.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            qs = Sample.objects.filter(id__in=sample_ids)
            deleted_count = qs.count()
            qs.delete()
            return Response({
                'message': f'Successfully deleted {deleted_count} sample(s).',
                'deleted_count': deleted_count
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': f'Failed to delete samples: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


class FinishBulkDeleteView(APIView):
    """
    POST /api/finishes/bulk-delete/
    Body: { "finish_ids": [1, 2, 3, ...] }
    Deletes all finishes with the given IDs. Admin only.
    """
    permission_classes = [IsAuthenticated, IsAdmin]

    def post(self, request):
        finish_ids = request.data.get('finish_ids', [])
        if not finish_ids:
            return Response({'error': 'No finish IDs provided.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            qs = Finish.objects.filter(id__in=finish_ids)
            deleted_count = qs.count()
            qs.delete()
            return Response({
                'message': f'Successfully deleted {deleted_count} finish(es).',
                'deleted_count': deleted_count
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': f'Failed to delete finishes: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)


# ─── Store Management ViewSets & APIs ─────────────────────────────────────────

class StoreItemCategoryViewSet(viewsets.ModelViewSet):
    queryset = StoreItemCategory.objects.all()
    serializer_class = StoreItemCategorySerializer
    permission_classes = [IsAuthenticated]


class StoreItemViewSet(viewsets.ModelViewSet):
    queryset = StoreItem.objects.select_related('category').prefetch_related('rate_history').all()
    serializer_class = StoreItemSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        inward_sub = Subquery(
            StoreMaterialIn.objects.filter(item=OuterRef('pk'))
            .values('item')
            .annotate(total=Sum('qty'))
            .values('total')[:1],
            output_field=DecimalField(max_digits=12, decimal_places=2)
        )
        issued_sub = Subquery(
            StoreDailyIssue.objects.filter(item=OuterRef('pk'))
            .values('item')
            .annotate(total=Sum('qty'))
            .values('total')[:1],
            output_field=DecimalField(max_digits=12, decimal_places=2)
        )
        qs = StoreItem.objects.select_related('category').prefetch_related('rate_history').annotate(
            inward_qty_sum=Coalesce(inward_sub, Value(Decimal('0.00'), output_field=DecimalField(max_digits=12, decimal_places=2))),
            issued_qty_sum=Coalesce(issued_sub, Value(Decimal('0.00'), output_field=DecimalField(max_digits=12, decimal_places=2)))
        ).all()
        category = self.request.query_params.get('category')
        search = self.request.query_params.get('search')
        status_param = self.request.query_params.get('default_status')
        low_stock = self.request.query_params.get('low_stock')

        if category:
            qs = qs.filter(category_id=category)
        if search:
            qs = qs.filter(Q(item_code__icontains=search) | Q(item_name__icontains=search))
        if status_param:
            qs = qs.filter(default_status=status_param)
        
        if low_stock == 'true':
            low_ids = [item.id for item in qs if item.balance_stock_qty <= item.reorder_level]
            qs = qs.filter(id__in=low_ids)

        return qs

    @action(detail=True, methods=['post'], url_path='revise-rate')
    def revise_rate(self, request, pk=None):
        """
        Action to revise item rate & log comparison metrics.
        Body: { "new_rate": 22.00, "supplier_name": "basawa ent", "po_reference": "PO-102", "revision_reason": "Price hike" }
        """
        item = self.get_object()
        new_rate = Decimal(str(request.data.get('new_rate', '0.00')))
        supplier_name = request.data.get('supplier_name', '')
        po_reference = request.data.get('po_reference', '')
        revision_reason = request.data.get('revision_reason', 'Rate Revision')

        if new_rate <= Decimal('0.00'):
            return Response({'error': 'New rate must be greater than zero.'}, status=status.HTTP_400_BAD_REQUEST)

        old_rate = item.current_rate or item.base_rate
        diff = new_rate - item.base_rate
        pct_change = round((diff / item.base_rate * 100), 2) if item.base_rate > 0 else Decimal('0.00')

        # Create history record
        rate_log = StoreItemRateHistory.objects.create(
            item=item,
            old_rate=old_rate,
            new_rate=new_rate,
            rate_difference=diff,
            percentage_change=pct_change,
            supplier_name=supplier_name,
            po_reference=po_reference,
            revision_reason=revision_reason,
            updated_by=request.user if request.user.is_authenticated else None
        )

        item.current_rate = new_rate
        item.save()

        return Response({
            'message': f'Rate revised from ₹{old_rate} to ₹{new_rate} successfully.',
            'item': StoreItemSerializer(item, context={'request': request}).data,
            'rate_log': StoreItemRateHistorySerializer(rate_log).data
        })

    @action(detail=False, methods=['get'], url_path='rate-comparison')
    def rate_comparison(self, request):
        """
        Returns comparison matrix of Base Master Rate vs Current Effective Rate across all store items.
        """
        items = StoreItem.objects.all()
        comparison = []
        for item in items:
            diff = item.current_rate - item.base_rate
            pct = round((diff / item.base_rate * 100), 2) if item.base_rate > 0 else Decimal('0.00')
            latest_history = item.rate_history.first()

            comparison.append({
                'id': str(item.id),
                'item_code': item.item_code,
                'item_name': item.item_name,
                'unit': item.unit,
                'base_rate': float(item.base_rate),
                'current_rate': float(item.current_rate),
                'difference': float(diff),
                'percentage_change': float(pct),
                'latest_supplier': latest_history.supplier_name if latest_history else "",
                'latest_po': latest_history.po_reference if latest_history else "",
                'last_updated': latest_history.effective_date.strftime('%Y-%m-%d') if latest_history else item.updated_at.strftime('%Y-%m-%d'),
            })

        return Response(comparison)

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        """
        Generates and downloads a clean Excel import template for Store Items.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Store Items Import"

        headers = [
            "Item Code", "Item Name", "Category Name", "Unit", 
            "Base Rate", "Reorder Level", "Default Status", "Weight (kg)", "Remark"
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color="EA580C", end_color="EA580C", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        sample_rows = [
            ["IT101", "Fevicol SH 50kg Drum", "Adhesives & Glue", "kg", 185.00, 10.00, "charge", 50.0, "High strength wood glue"],
            ["IT102", "Sandpaper 120 Grit Sheet", "Abrasives & Sanding", "pcs", 12.50, 100.00, "charge", 0.05, "Waterproof silicon carbide paper"],
            ["IT103", "NC Clear Thinner 20L", "Chemicals & Thinners", "ltr", 110.00, 15.00, "charge", 18.0, "Fast drying lacquer thinner"],
            ["IT104", "Antique Brass Drop Handle 96mm", "Hardware & Fittings", "pcs", 45.00, 50.00, "charge", 0.12, "Cast brass vintage finish handle"],
            ["IT105", "Masking Tape 2 inch Roll", "Packaging & Tapes", "roll", 35.00, 25.00, "free", 0.1, "General masking tape for polishing"]
        ]

        for r in sample_rows:
            ws.append(r)

        col_widths = [15, 32, 22, 10, 15, 15, 16, 14, 35]
        for idx, width in enumerate(col_widths, 1):
            col_letter = get_column_letter(idx)
            ws.column_dimensions[col_letter].width = width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Store_Item_Master_Import_Template.xlsx'
        return response

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        """
        Imports Store Items from an uploaded Excel or CSV file.
        Checks for duplicates by item_code and item_name, skipping duplicate records.
        """
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'No file uploaded. Please upload a valid .xlsx, .xls, or .csv file.'}, status=status.HTTP_400_BAD_REQUEST)

        filename = file_obj.name.lower()
        rows_data = []

        try:
            if filename.endswith('.csv'):
                decoded_file = file_obj.read().decode('utf-8-sig')
                io_string = io.StringIO(decoded_file)
                reader = csv.DictReader(io_string)
                for r in reader:
                    rows_data.append(r)
            else:
                wb = openpyxl.load_workbook(file_obj, data_only=True)
                ws = wb.active
                headers = [str(cell.value or '').strip() for cell in ws[1]]
                
                header_map = {}
                for idx, h in enumerate(headers):
                    hl = h.lower().replace(' ', '_').replace('(', '').replace(')', '').replace('₹', '').replace('kg', '').strip('_')
                    header_map[idx] = hl

                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    row_dict = {}
                    for idx, val in enumerate(row):
                        key = header_map.get(idx)
                        if key:
                            row_dict[key] = val
                    rows_data.append(row_dict)
        except Exception as e:
            return Response({'detail': f'Error reading Excel file: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        if not rows_data:
            return Response({'detail': 'The uploaded Excel file contains no data rows.'}, status=status.HTTP_400_BAD_REQUEST)

        existing_codes = set(StoreItem.objects.values_list('item_code', flat=True))
        existing_names = set(StoreItem.objects.values_list('item_name', flat=True))
        categories_map = {c.name.lower(): c for c in StoreItemCategory.objects.all()}

        created_count = 0
        skipped_count = 0
        skipped_items = []

        for idx, r in enumerate(rows_data, start=2):
            item_code = str(r.get('item_code') or r.get('code') or f'IT{random.randint(100, 999)}').strip()
            item_name = str(r.get('item_name') or r.get('name') or '').strip()

            if not item_name:
                continue

            # Check Duplicate
            if item_code in existing_codes or item_name.lower() in [n.lower() for n in existing_names]:
                skipped_count += 1
                skipped_items.append({'row': idx, 'item_code': item_code, 'item_name': item_name, 'reason': 'Duplicate Item Code or Name'})
                continue

            category_name = str(r.get('category_name') or r.get('category') or '').strip()
            category_obj = None
            if category_name:
                category_obj = categories_map.get(category_name.lower())
                if not category_obj:
                    category_obj = StoreItemCategory.objects.create(name=category_name)
                    categories_map[category_name.lower()] = category_obj

            unit = str(r.get('unit') or 'pcs').strip()
            try:
                base_rate = Decimal(str(r.get('base_rate') or r.get('rate') or '0.00')).quantize(Decimal('0.01'))
            except Exception:
                base_rate = Decimal('0.00')

            try:
                reorder_level = Decimal(str(r.get('reorder_level') or '10.00')).quantize(Decimal('0.01'))
            except Exception:
                reorder_level = Decimal('10.00')

            default_status = str(r.get('default_status') or 'charge').strip().lower()
            if default_status not in ['charge', 'free']:
                default_status = 'charge'

            weight_val = None
            raw_w = r.get('weight') or r.get('weight_kg')
            if raw_w:
                try:
                    weight_val = Decimal(str(raw_w)).quantize(Decimal('0.01'))
                except Exception:
                    weight_val = None

            remark = str(r.get('remark') or r.get('remarks') or '').strip()

            new_item = StoreItem.objects.create(
                item_code=item_code,
                item_name=item_name,
                category=category_obj,
                unit=unit,
                base_rate=base_rate,
                current_rate=base_rate,
                reorder_level=reorder_level,
                default_status=default_status,
                weight=weight_val,
                remark=remark
            )

            existing_codes.add(item_code)
            existing_names.add(item_name)
            created_count += 1

        return Response({
            'detail': f'Import completed! {created_count} items created, {skipped_count} duplicate items skipped.',
            'imported': created_count,
            'duplicates_skipped': skipped_count,
            'total_processed': len(rows_data),
            'skipped_items': skipped_items
        })



class ContractorPersonViewSet(viewsets.ModelViewSet):
    queryset = ContractorPerson.objects.all()
    serializer_class = ContractorPersonSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = ContractorPerson.objects.all()
        contractor_id = self.request.query_params.get('contractor')
        if contractor_id:
            qs = qs.filter(contractor_id=contractor_id)
        return qs


class StorePurchaseOrderViewSet(viewsets.ModelViewSet):
    queryset = StorePurchaseOrder.objects.all()
    serializer_class = StorePurchaseOrderSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class StoreMaterialInViewSet(viewsets.ModelViewSet):
    queryset = StoreMaterialIn.objects.all()
    serializer_class = StoreMaterialInSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = StoreMaterialIn.objects.all()
        supplier_id = self.request.query_params.get('supplier')
        month = self.request.query_params.get('month')
        item_id = self.request.query_params.get('item')

        if supplier_id:
            qs = qs.filter(supplier_id=supplier_id)
        if month:
            qs = qs.filter(month_year__icontains=month)
        if item_id:
            qs = qs.filter(item_id=item_id)
        return qs

    def perform_create(self, serializer):
        instance = serializer.save(received_by=self.request.user)
        # Update item's current rate if bill rate is updated
        item = instance.item
        if instance.bill_rate > Decimal('0.00') and instance.bill_rate != item.current_rate:
            old_rate = item.current_rate or item.base_rate
            diff = instance.bill_rate - item.base_rate
            pct_change = round((diff / item.base_rate * 100), 2) if item.base_rate > 0 else Decimal('0.00')
            
            StoreItemRateHistory.objects.create(
                item=item,
                old_rate=old_rate,
                new_rate=instance.bill_rate,
                rate_difference=diff,
                percentage_change=pct_change,
                supplier_name=instance.supplier.name if instance.supplier else "",
                po_reference=instance.bill_no,
                revision_reason=f"Auto-updated from Material Inward #{instance.voucher_no}",
                updated_by=self.request.user
            )
            item.current_rate = instance.bill_rate
    def destroy(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            return Response({'detail': 'Only Admin users are authorized to void or delete store material in vouchers.'}, status=status.HTTP_403_FORBIDDEN)
        instance = self.get_object()
        reason = request.data.get('reason') or request.query_params.get('reason') or 'Admin Void Voucher'
        
        log_audit_event(
            user=request.user,
            action=AuditAction.DELETE,
            module_name='Store Management',
            model_name='StoreMaterialIn',
            object_id=instance.id,
            object_repr=f"Voided Inward Voucher #{instance.voucher_no} ({instance.item.item_name})",
            reason=reason,
            request=request
        )
        return super().destroy(request, *args, **kwargs)


class StoreDailyIssueViewSet(viewsets.ModelViewSet):
    queryset = StoreDailyIssue.objects.all()
    serializer_class = StoreDailyIssueSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = StoreDailyIssue.objects.all()
        contractor_id = self.request.query_params.get('contractor')
        month = self.request.query_params.get('month')
        status_param = self.request.query_params.get('status')
        unit_id = self.request.query_params.get('production_unit')

        if contractor_id:
            qs = qs.filter(contractor_id=contractor_id)
        if month:
            qs = qs.filter(month_year__icontains=month)
        if status_param:
            qs = qs.filter(status=status_param)
        if unit_id:
            qs = qs.filter(production_unit_id=unit_id)
        return qs

    def perform_create(self, serializer):
        issue = serializer.save(issued_by=self.request.user)
        item = issue.item
        if item and item.balance_stock_qty <= item.reorder_level:
            admin_and_store_users = User.objects.filter(role__in=['admin', 'store_manager'], is_active=True)
            for u in admin_and_store_users:
                Notification.objects.create(
                    recipient=u,
                    title=f"Low Stock Alert: {item.item_name} ({item.item_code})",
                    message=f"Store balance for {item.item_name} has dropped to {item.balance_stock_qty} {item.unit} (Reorder Level: {item.reorder_level} {item.unit}). Please reorder soon.",
                    link=f"/store-management"
                )

    def destroy(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            return Response({'detail': 'Only Admin users are authorized to void or delete store daily issue vouchers.'}, status=status.HTTP_403_FORBIDDEN)
        instance = self.get_object()
        reason = request.data.get('reason') or request.query_params.get('reason') or 'Admin Void Voucher'
        
        log_audit_event(
            user=request.user,
            action=AuditAction.DELETE,
            module_name='Store Management',
            model_name='StoreDailyIssue',
            object_id=instance.id,
            object_repr=f"Voided Outward Issue Voucher #{instance.voucher_no} ({instance.item.item_name})",
            reason=reason,
            request=request
        )
        return super().destroy(request, *args, **kwargs)


class StoreMaterialReturnViewSet(viewsets.ModelViewSet):
    queryset = StoreMaterialReturn.objects.all()
    serializer_class = StoreMaterialReturnSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = StoreMaterialReturn.objects.all()
        contractor_id = self.request.query_params.get('contractor')
        month = self.request.query_params.get('month')
        status_param = self.request.query_params.get('status')
        unit_id = self.request.query_params.get('production_unit')

        if contractor_id:
            qs = qs.filter(contractor_id=contractor_id)
        if month:
            qs = qs.filter(month_year__icontains=month)
        if status_param:
            qs = qs.filter(status=status_param)
        if unit_id:
            qs = qs.filter(production_unit_id=unit_id)
        return qs

    def perform_create(self, serializer):
        serializer.save(returned_by=self.request.user)

    def destroy(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            return Response({'detail': 'Only Admin users are authorized to void or delete store material return vouchers.'}, status=status.HTTP_403_FORBIDDEN)
        instance = self.get_object()
        reason = request.data.get('reason') or request.query_params.get('reason') or 'Admin Void Voucher'
        
        log_audit_event(
            user=request.user,
            action=AuditAction.DELETE,
            module_name='Store Management',
            model_name='StoreMaterialReturn',
            object_id=instance.id,
            object_repr=f"Voided Return Voucher #{instance.voucher_no} ({instance.item.item_name})",
            reason=reason,
            request=request
        )
        return super().destroy(request, *args, **kwargs)


class StoreRequisitionViewSet(viewsets.ModelViewSet):
    queryset = StoreRequisition.objects.select_related('requested_by', 'approved_by', 'item', 'production_unit').all()
    serializer_class = StoreRequisitionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        status_param = self.request.query_params.get('status')
        unit_id = self.request.query_params.get('production_unit')
        user = self.request.user

        if status_param:
            qs = qs.filter(status=status_param)
        if unit_id:
            qs = qs.filter(production_unit_id=unit_id)

        if user.role in ['supervisor', 'contractor']:
            qs = qs.filter(requested_by=user)
        return qs

    def perform_create(self, serializer):
        serializer.save(requested_by=self.request.user)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        if request.user.role not in ['admin', 'store_manager']:
            return Response({'detail': 'Only Admin or Store Manager can approve requisitions.'}, status=status.HTTP_403_FORBIDDEN)
        mrn = self.get_object()
        mrn.status = 'approved'
        mrn.approved_by = request.user
        mrn.save()
        return Response(StoreRequisitionSerializer(mrn).data)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        if request.user.role not in ['admin', 'store_manager']:
            return Response({'detail': 'Only Admin or Store Manager can reject requisitions.'}, status=status.HTTP_403_FORBIDDEN)
        mrn = self.get_object()
        mrn.status = 'rejected'
        mrn.rejection_reason = request.data.get('reason', 'Rejected by Store Management')
        mrn.save()
        return Response(StoreRequisitionSerializer(mrn).data)


class StoreStockAdjustmentViewSet(viewsets.ModelViewSet):
    queryset = StoreStockAdjustment.objects.select_related('item', 'logged_by', 'approved_by').all()
    serializer_class = StoreStockAdjustmentSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        status_param = self.request.query_params.get('status')
        if status_param:
            qs = qs.filter(status=status_param)
        return qs

    def perform_create(self, serializer):
        serializer.save(logged_by=self.request.user)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve(self, request, pk=None):
        if request.user.role != 'admin':
            return Response({'detail': 'Only Admin users can approve stock variance adjustments.'}, status=status.HTTP_403_FORBIDDEN)
        adj = self.get_object()
        adj.status = 'approved'
        adj.approved_by = request.user
        adj.admin_remark = request.data.get('remark', 'Approved by Admin')
        adj.save()

        log_audit_event(
            user=request.user,
            action=AuditAction.UPDATE,
            module_name='Store Management',
            model_name='StoreStockAdjustment',
            object_id=adj.id,
            object_repr=f"Approved Stock Adjustment #{adj.adjustment_no} ({adj.item.item_name}: {adj.quantity_delta})",
            reason=adj.admin_remark,
            request=request
        )
        return Response(StoreStockAdjustmentSerializer(adj).data)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject(self, request, pk=None):
        if request.user.role != 'admin':
            return Response({'detail': 'Only Admin users can reject stock variance adjustments.'}, status=status.HTTP_403_FORBIDDEN)
        adj = self.get_object()
        adj.status = 'rejected'
        adj.admin_remark = request.data.get('remark', 'Rejected by Admin')
        adj.save()
        return Response(StoreStockAdjustmentSerializer(adj).data)


class StoreStockSummaryView(APIView):
    """
    GET /api/store/stock-summary/
    Returns real-time aggregated inventory stats for Store Items matching Excel Sheet 1:
    - Total Items Count
    - Total Inward Stock Qty
    - Total Issued Stock Qty
    - Net Available Stock Qty
    - Total Inventory Valuation (₹)
    - Detailed list of item inventory records
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        items = StoreItem.objects.select_related('category').all()
        
        # Pre-aggregate totals across all items in 4 single bulk SQL queries instead of 4*N queries
        inward_map = dict(
            StoreMaterialIn.objects.values('item_id')
            .annotate(total=Sum('qty'))
            .values_list('item_id', 'total')
        )
        issued_map = dict(
            StoreDailyIssue.objects.values('item_id')
            .annotate(total=Sum('qty'))
            .values_list('item_id', 'total')
        )
        returned_map = dict(
            StoreMaterialReturn.objects.values('item_id')
            .annotate(total=Sum('qty'))
            .values_list('item_id', 'total')
        )
        adjustment_map = dict(
            StoreStockAdjustment.objects.filter(status='approved')
            .values('item_id')
            .annotate(total=Sum('quantity_delta'))
            .values_list('item_id', 'total')
        )

        summary_list = []
        tot_stock = Decimal('0.00')
        tot_issued = Decimal('0.00')
        tot_balance = Decimal('0.00')
        tot_valuation = Decimal('0.00')

        for item in items:
            stk_qty = inward_map.get(item.id) or Decimal('0.00')
            iss_qty = issued_map.get(item.id) or Decimal('0.00')
            ret_qty = returned_map.get(item.id) or Decimal('0.00')
            adj_qty = adjustment_map.get(item.id) or Decimal('0.00')

            bal_qty = (stk_qty + ret_qty + adj_qty) - iss_qty
            rate = item.current_rate or item.base_rate
            val = bal_qty * rate

            tot_stock += stk_qty
            tot_issued += iss_qty
            tot_balance += bal_qty
            tot_valuation += val

            summary_list.append({
                'id': str(item.id),
                'item_code': item.item_code,
                'item_name': item.item_name,
                'category_name': item.category.name if item.category else 'General',
                'stock_qty': float(stk_qty),
                'issued_qty': float(iss_qty),
                'balance_qty': float(bal_qty),
                'rate': float(rate),
                'unit': item.unit,
                'total_value': float(val),
                'default_status': item.default_status,
                'reorder_level': float(item.reorder_level),
                'is_low_stock': bal_qty <= item.reorder_level
            })

        return Response({
            'total_items_count': len(items),
            'total_stock_qty': float(tot_stock),
            'total_issued_qty': float(tot_issued),
            'total_balance_qty': float(tot_balance),
            'total_inventory_valuation': float(tot_valuation),
            'items': summary_list
        })


class MonthlyContractorBillingView(APIView):
    """
    GET /api/store/monthly-contractor-bill/?contractor=<id>&month=Jul-26
    Generates monthly store billing summary for a specific contractor:
    - Lists all Chargeable StoreDailyIssue records for that contractor in that month.
    - Aggregates Total Chargeable Amount, Total Items, and Non-chargeable summary.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        contractor_id = request.query_params.get('contractor')
        month = request.query_params.get('month')

        if not contractor_id:
            return Response({'error': 'contractor parameter is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            contractor = User.objects.get(id=contractor_id)
        except User.DoesNotExist:
            return Response({'error': 'Contractor not found.'}, status=status.HTTP_404_NOT_FOUND)

        qs = StoreDailyIssue.objects.filter(contractor=contractor)
        if month:
            qs = qs.filter(month_year__icontains=month)

        chargeable_qs = qs.filter(status=StoreItemStatus.CHARGE)
        non_chargeable_qs = qs.filter(status=StoreItemStatus.NON_CHARGE)

        chargeable_items = StoreDailyIssueSerializer(chargeable_qs, many=True).data
        non_chargeable_items = StoreDailyIssueSerializer(non_chargeable_qs, many=True).data

        total_chargeable_amt = sum(float(item['chargeable_total'] or 0) for item in chargeable_items)
        total_non_chargeable_amt = sum(float(item['non_chargeable_total'] or 0) for item in non_chargeable_items)

        return Response({
            'contractor': {
                'id': str(contractor.id),
                'username': contractor.username,
                'full_name': contractor.get_full_name() or contractor.username,
                'phone': contractor.phone,
            },
            'month_year': month or 'All Time',
            'chargeable_items': chargeable_items,
            'non_chargeable_items': non_chargeable_items,
            'total_chargeable_amount': total_chargeable_amt,
            'total_non_chargeable_amount': total_non_chargeable_amt,
            'total_issue_entries_count': len(chargeable_items) + len(non_chargeable_items)
        })


class DatabaseRelationshipsPDFView(APIView):
    """
    GET /api/tools/database-relationships-pdf/
    Returns a single landscape A3 PDF file displaying all ERP database tables,
    their Primary Keys, Foreign Keys, and directional connector lines.
    Automatically inspects Django models dynamically.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            pdf_bytes = generate_db_relationships_pdf()
            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="Database_Relationships.pdf"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            return response
        except Exception as err:
            print("Error generating Database Relationships PDF:", err)
            return Response(
                {'detail': f'Failed to generate database relationships PDF: {str(err)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ─── Global Audit Log System ──────────────────────────────────────────────────

def log_audit_event(user, action, module_name, model_name, object_id='', object_repr='', changes=None, file_info=None, reason=None, request=None):
    """
    Helper utility to log audit events into the unified AuditLog table.
    """
    from .middleware import get_client_ip, get_user_agent
    ip_addr = get_client_ip(request)
    user_agent = get_user_agent(request)
    
    username = 'System'
    role = 'system'
    if user and getattr(user, 'is_authenticated', False):
        username = user.get_full_name() or user.username
        role = getattr(user, 'role', 'user')

    return AuditLog.objects.create(
        user=user if user and getattr(user, 'is_authenticated', False) else None,
        username=username,
        user_role=role,
        ip_address=ip_addr,
        user_agent=user_agent[:500] if user_agent else '',
        action=action,
        module_name=module_name,
        model_name=model_name,
        object_id=str(object_id) if object_id else '',
        object_repr=str(object_repr)[:250] if object_repr else '',
        changes=changes or {},
        file_info=file_info or {},
        reason=reason or ''
    )


class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only viewset for inspecting global audit logs. Restricted to Admin role.
    Supports filtering by user, module, action, search query, and date range.
    """
    queryset = AuditLog.objects.select_related('user').all()
    serializer_class = AuditLogSerializer
    permission_classes = [IsAuthenticated, IsAdmin]
    pagination_class = OptionalPagination
    ordering = ['-timestamp']

    def get_queryset(self):
        qs = super().get_queryset()
        
        user_id = self.request.query_params.get('user')
        if user_id:
            qs = qs.filter(user_id=user_id)

        module_name = self.request.query_params.get('module')
        if module_name:
            qs = qs.filter(module_name__iexact=module_name)

        action = self.request.query_params.get('action')
        if action:
            qs = qs.filter(action=action)

        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(username__icontains=search) |
                Q(object_repr__icontains=search) |
                Q(module_name__icontains=search) |
                Q(model_name__icontains=search) |
                Q(reason__icontains=search) |
                Q(ip_address__icontains=search)
            )

        start_date = self.request.query_params.get('start_date')
        if start_date:
            qs = qs.filter(timestamp__date__gte=start_date)

        end_date = self.request.query_params.get('end_date')
        if end_date:
            qs = qs.filter(timestamp__date__lte=end_date)

        return qs


class HealthCheckView(APIView):
    """
    Ultra-lightweight keep-alive health check endpoint for UptimeRobot pinging.
    Allows unauthenticated requests to keep the Render container awake 24/7.
    """
    permission_classes = [AllowAny]

    def get(self, request):
        return Response({"status": "ok", "message": "ERP Backend is awake and healthy"})

